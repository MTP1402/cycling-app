# ═══════════════════════════════════════════════════════════════════
# CYCLING COACH API — main.py
#
# VERSION: 2.14.0  (2026-08-03)
# Check this against GET / on the live Railway URL before assuming
# a deploy has actually landed — the two should always match.
#
# CHANGELOG
#   2.14.0 (2026-08-03) — four pieces built together tonight, all
#                         extending the power-curve foundation from
#                         2.13.x or the existing flagged-duplicate
#                         system from 2.8.0/2.9.x:
#                         FTP ESTIMATE — new estimate_ftp_from_curve()
#                         (95% of best 20-min power, the standard field
#                         estimate). Shown as a banner on the ride-
#                         detail page only when a real estimate exists
#                         AND it differs from the profile's on-file FTP
#                         by 5%+, so it doesn't nag on every ride. Never
#                         auto-applies — always a suggestion with an
#                         explicit confirm step. New POST /profile/ftp
#                         for that confirm step: a targeted single-
#                         column UPDATE, deliberately separate from
#                         POST /profile, which re-saves the WHOLE form
#                         and would have silently wiped every other
#                         profile field if called with just ftp set.
#                         Caught a real bug in testing, not shipped: the
#                         first version also had a 60-minute-direct
#                         fallback for rides without a 20-min checkpoint
#                         but long enough for 60 — dead code, since
#                         compute_power_curve() finds the best window of
#                         each duration anywhere in the ride, so any
#                         ride long enough for a 3600s checkpoint
#                         necessarily has a 1200s one too. That branch
#                         could never fire; removed rather than left in.
#                         POWER CURVE COACHING TOOL — new 6th AI tool,
#                         get_power_curve, returning key checkpoints
#                         (5s/15s/30s/1/5/10/20/30/60-min, whichever the
#                         ride has) plus the FTP estimate when
#                         computable — so the coaching chat can discuss
#                         sustained power at a specific duration or
#                         whether a ride looks like a new FTP, beyond
#                         the basic 5s/15s/30s/5-min bests already in
#                         its context.
#                         FLAGGED-DUPLICATE BADGE — GET /coaching/memory
#                         now returns a `flagged` bool per dated_log
#                         entry (a correlated EXISTS check against
#                         possible_duplicate_of), so the Ride History
#                         list can show an in-context badge, not just
#                         the one dedicated Dashboard review card, which
#                         is easy to miss if a rider doesn't happen to
#                         scroll to it. Resolution itself is unchanged —
#                         same existing /rides/flagged, /rides/{id}/
#                         clear-review, DELETE /rides/{id}.
#                         TCX EXPORT — new GET /rides/{id}/export-tcx,
#                         built by build_tcx() from the ride's own
#                         stored raw stream data. GPX was considered and
#                         explicitly ruled out: GPX trackpoints require
#                         real lat/lon per spec, and this app has never
#                         captured GPS coordinates (FIT/Strava streams
#                         here only ever held power/HR/cadence/altitude/
#                         distance/speed) — a GPX with fake coordinates
#                         would misrepresent the actual route. TCX's
#                         schema makes Position genuinely optional, so
#                         this is a real, standards-compliant export
#                         (TrainingPeaks/Garmin Connect/WKO-readable)
#                         carrying everything except a route/map. TCX
#                         *import* stays a separate, larger follow-up —
#                         not attempted here.
#                         Tested before shipping: estimate_ftp_from_curve
#                         verified against the same synthetic power
#                         curve data used for compute_power_curve's own
#                         tests (20-min and 60-min branches both
#                         verified); ride-detail page rendered with a
#                         real profile_ftp comparison to confirm the
#                         banner only appears past the 5% threshold;
#                         generated TCX XML parsed back with a real XML
#                         parser to confirm it's well-formed, not just
#                         string-built; get_power_curve tool tested
#                         directly against synthetic stream data.
#   2.13.1 (2026-08-03) — fixed a real usability gap on the new Power
#                         Curve chart, reported immediately after trying
#                         it: the tooltip only fired on a pixel-perfect
#                         hit against an invisible (pointRadius:0) point
#                         sitting exactly on the line — workable with a
#                         careful mouse, effectively impossible with a
#                         finger on the phone app. Set interaction mode
#                         to 'nearest' with axis:'x' and intersect:false,
#                         so a hover or tap anywhere near a given duration
#                         finds the closest point along the x-axis rather
#                         than requiring an exact pixel match — standard
#                         fix for this exact class of problem, same
#                         approach the other ride-detail charts didn't
#                         need only because they're driven by the
#                         separate drag-select interaction instead of
#                         point-hover tooltips. Also widened each point's
#                         own hit target (pointHitRadius:20) as a second,
#                         complementary fix, and added a small visible
#                         dot on hover (pointHoverRadius:5) so there's
#                         confirmation of exactly which point is
#                         selected, not just a tooltip appearing.
#   2.13.0 (2026-08-03) — added the Power Curve chart to the ride-detail
#                         page (carried-over item from the handoff doc) —
#                         continuous best-average-power across every
#                         duration from 1s out to the full ride length,
#                         not just the four fixed checkpoints (5s/15s/30s/
#                         5-min) already shown on the long-term Dashboard
#                         trend charts, which are untouched by this. New
#                         compute_power_curve() helper: prefix-sum rolling-
#                         window max per checkpoint duration (a log-ish
#                         spaced set, dense at the short end and sparse at
#                         the long end, matching the standard shape this
#                         kind of curve is always drawn with) rather than
#                         a naive O(n^2) scan over every possible duration,
#                         which would be far too slow on a 3+ hour ride at
#                         request time.
#                         Deliberately treats a missing/None power reading
#                         as 0 watts (genuine coasting), NOT filtered out
#                         the way the existing p5/p15/p30/p300 checkpoints
#                         are (see parse_fit_bytes) — a power curve needs
#                         to reflect real sustained output including
#                         recovery dips, and that's the standard
#                         definition every comparable tool uses. This is a
#                         new, separate computation making its own
#                         deliberate choice, not a change to the existing
#                         fixed-checkpoint convention elsewhere in the
#                         file, which is intentionally left alone.
#                         Chart has its own log-scale duration x-axis with
#                         custom "5s/1m/5m/20m/1h" tick and tooltip
#                         formatting — a completely different domain from
#                         the mile-based charts above it, so it's
#                         deliberately NOT added to the shared drag-
#                         select-to-recompute chartIds group; dragging on
#                         a duration axis wouldn't mean the same thing as
#                         dragging on a mile-based one.
#                         Scoped to time-domain only, matching what Marc
#                         asked for tonight — no historical/rolling-window
#                         comparison overlay (e.g. "vs last 6 weeks," as
#                         seen in the reference screenshot) and no
#                         distance-based best-effort ladder (5mi/10K/10mi/
#                         20K + HR/elevation) — both stay explicitly on
#                         the backlog as separate, larger pieces.
#                         Hidden entirely (no card, no broken chart) on
#                         rides with fewer than 3 computable duration
#                         checkpoints — very short rides, or rides
#                         predating raw-stream storage.
#                         Tested before shipping: compute_power_curve()
#                         verified against synthetic power data with a
#                         known 60s all-out effort inside an otherwise
#                         steady ride — the 60s checkpoint correctly
#                         isolates the effort's true average, shorter
#                         checkpoints inside that window read higher (as
#                         expected, per-second exceeding the sustained
#                         60s average), and checkpoints longer than the
#                         effort correctly dilute toward the ride's
#                         overall average as the window grows past it.
#                         Generated Chart.js JS syntax-checked directly.
#   2.12.0 (2026-08-03) — fixed two real bugs on the ride-detail page's
#                         drag-select-to-recompute (built in 2.9.0, never
#                         browser-tested until now): every chart there was
#                         left on Chart.js's default CATEGORY x-axis
#                         (parallel labels/data arrays, no scales.x.type
#                         set) instead of a real LINEAR scale keyed to
#                         actual mile values. getValueForPixel() on a
#                         category scale returns the *index* into the
#                         label array, not the mile value the label
#                         displays — so a drag anywhere on the chart was
#                         silently computing stats from index numbers
#                         mislabeled as miles. Near the start of a ride
#                         the index and the real mile value happen to be
#                         close in magnitude, which is why dragging there
#                         looked plausible (confirmed live: a drag near
#                         mile 0 on a 42.44mi ride reported "44 mi
#                         selected" — an index span, not a real distance)
#                         while dragging further out clearly didn't work.
#                         Fixed by switching every chart to
#                         scales.x.type:'linear' with each dataset passed
#                         as real {x,y} point objects instead of a shared
#                         category-label array. computeSelectionStats()
#                         and pixelToValue() needed no logic changes —
#                         they were already written correctly for real
#                         mile values, they just weren't being fed them.
#                         Also fixed: the charts had no explicit height,
#                         so Chart.js's responsive sizing kept growing
#                         with nothing to anchor to, and only one chart
#                         fit on screen at a time. .dchart-wrap now has a
#                         fixed 260px height (matches the cap already
#                         used successfully on the main Dashboard's
#                         charts) with maintainAspectRatio:false so the
#                         canvas fills it instead of dictating its own
#                         size.
#                         Verified via a synthetic multi-point ride: drag
#                         selections at the start, middle, and end of the
#                         ride now all report a selected-mile span that
#                         matches the actual drag range and never exceeds
#                         total ride distance (the "44 mi selected on a
#                         42.44mi ride" symptom is gone); confirmed the
#                         same fix applies uniformly across all charts
#                         including the FIT-upload-only L/R balance chart.
#                         The stat-cards-update-in-place UX Marc asked
#                         for (folding the selection readout into the
#                         existing top stat cards instead of a separate
#                         line below) is still open — deliberately kept
#                         out of this fix to isolate the two confirmed
#                         bugs from a UI reshuffle; next up.
#   2.11.0 (2026-08-02) — closed the same gap on the OTHER import
#                         path: manual FIT uploads never checked
#                         activity type either, only Strava sync did
#                         (fixed in 2.10.0). FIT files carry a sport
#                         field — already extracted for virtual-ride
#                         detection, just never checked against non-
#                         cycling activities. /upload now rejects
#                         running/walking/hiking/swimming with a clear
#                         error; a missing/blank sport is still
#                         allowed, since some devices don't always
#                         populate it and there's no reason to assume
#                         the worst on an ambiguous case.
#                         Also added GET /rides/audit-slow-pace — a
#                         genuine way to find walks already sitting in
#                         the database from before this fix existed.
#                         The app never stored activity type on
#                         existing rows, so there's no direct lookup;
#                         this uses average pace instead (<6 mph) as a
#                         real, if imperfect, heuristic — walking pace
#                         is categorically different from even a slow
#                         ride. Verified directly against Marc's own
#                         real walk data (6 activities, 6.4mi, 2h13m —
#                         2.89 mph average) before shipping, and
#                         confirmed a genuinely hilly, slow ride (8mph)
#                         correctly stays unflagged. Surfaces
#                         candidates for review only — deletes nothing
#                         on its own.
#   2.10.0 (2026-08-02) — Strava sync has never filtered by activity
#                         type — it pulled and stored everything
#                         Strava returned (walks, runs, hikes, etc.)
#                         as if they were rides. Not related to the
#                         dedup bug fixed earlier the same day — a
#                         separate, longer-standing gap, found while
#                         accounting for a small remaining mileage
#                         difference (app showing 3,006.1mi/107 rides
#                         against Strava's own 2,999.6mi/100 activities
#                         after cleaning up the dedup bug's damage —
#                         Marc's own hypothesis that the extra ~7
#                         entries might be walks, not rides). Added
#                         CYCLING_ACTIVITY_TYPES, checking Strava's
#                         sport_type field first (more specific —
#                         GravelRide, MountainBikeRide, etc.), falling
#                         back to the older type field if not present.
#                         Prevents this going forward; does NOT
#                         retroactively identify anything already
#                         imported, since activity type was never
#                         stored on existing rows — no way to tell,
#                         after the fact, which past rides might have
#                         been misclassified. Manual inspection (short
#                         distance, odd pace, a telling name) is the
#                         only way to find any that already slipped
#                         through.
#   2.9.2 (2026-08-02) — added GET /rides/audit-duplicates, a one-time
#                         read-only diagnostic to find rides matching
#                         the exact pattern the v2.8.0 bug (fixed in
#                         2.9.1) created — a freshly re-synced ride
#                         paired with the older pre-existing ride it
#                         duplicated. Surfaces candidates for review;
#                         doesn't delete anything on its own.
#   2.9.1 (2026-08-02) — CRITICAL FIX: a real data-loss bug introduced
#                         by the v2.8.0 overlap-based dedup redesign.
#                         That version required the EXISTING candidate
#                         row to also have start_time/elapsed_h to be
#                         considered a possible duplicate at all — but
#                         start_time only began being populated
#                         partway through that same session, meaning
#                         essentially a rider's entire ride history
#                         predating it has start_time=NULL. The moment
#                         a newly-synced Strava activity had its own
#                         start_time (which is nearly always, since
#                         Strava provides it directly), the dedup
#                         check would find zero matches against ANY
#                         older row — not because no duplicate
#                         existed, but because the query structurally
#                         excluded rows lacking start_time from ever
#                         matching, with no fallback. Every re-sync of
#                         already-imported rides was silently creating
#                         full duplicates, completely unflagged.
#                         Confirmed in production: Strava itself
#                         showed 100 activities / 2,999.6 mi for the
#                         year; the app showed 141 rides / 3,924.6 mi
#                         — 41 extra rides, ~925 extra miles.
#                         Fixed by making the fallback decision
#                         per-candidate-row instead of once globally:
#                         the distance/duration check now explicitly
#                         runs against rows lacking start_time/
#                         elapsed_h (exactly the rows overlap can't
#                         evaluate), while rows that DO have full
#                         timing data are only ever compared via
#                         overlap — never re-evaluated by distance/
#                         duration, which is what would reintroduce
#                         the earlier false-positive bug from the same
#                         session. Re-tested all three critical cases
#                         together before shipping: the actual bug
#                         scenario (now caught via fallback, skipped
#                         silently — matching the original long-
#                         standing behavior for exact-match dupes),
#                         the earlier false-positive case (still
#                         correctly not flagged), and the original
#                         overlap scenario (staggered start + forgotten
#                         stop — still correctly flagged for review,
#                         not silently skipped).
#                         Does NOT retroactively clean up rides already
#                         wrongly duplicated before this fix — that's a
#                         separate, deliberate cleanup step.
#   2.9.0 (2026-07-29) — added drag-select-to-recompute on the ride-
#                         detail page, the last deferred piece from
#                         that original build. Drag across any chart
#                         and it recomputes stats for just that
#                         stretch — synced as a highlighted range
#                         across all charts at once, using Chart.js's
#                         own scale API (getValueForPixel /
#                         getPixelForValue) to convert between mouse
#                         position and the actual distance value,
#                         rather than tracking raw pixels.
#                         Stats are computed from the FULL-resolution
#                         stream data (now embedded separately from
#                         the ~400-point downsampled chart arrays),
#                         not the thinned data used for rendering —
#                         a short selection needs real precision, not
#                         chart-smoothing-level approximation. Uses
#                         the same non-zero-power averaging convention
#                         as the rest of the app for consistency.
#                         Tested thoroughly: the actual generated
#                         computeSelectionStats function (not a
#                         reimplementation) against a synthetic ride
#                         with a distinct climb, confirming the climb
#                         and a flat section produce clearly different,
#                         correct stats, and that a reversed drag
#                         (dragging right-to-left) gives identical
#                         results to a forward drag. The drag/mouse
#                         mechanics themselves can't be simulated in
#                         this environment — worth trying for real to
#                         confirm the feel, though the underlying
#                         logic is fully verified.
#   2.8.0 (2026-07-29) — replaced start-time-proximity dedup with a
#                         real time-window overlap check, per Marc's
#                         exact real scenario: starting two separate
#                         recording devices at different moments, and
#                         sometimes forgetting to stop one for hours.
#                         The previous fix (start times within 10 min)
#                         could still miss both a staggered start past
#                         10 min AND a forgotten-running device saved
#                         much later — neither necessarily keeps start
#                         times close. Now checks whether the two
#                         rides' actual [start, device-off] windows
#                         overlap at all, using elapsed_h (wall-clock
#                         time including pauses) to compute a real end
#                         time — correctly catches both cases
#                         regardless of how mismatched the reported
#                         distance/duration end up looking.
#                         Bigger behavior change: when an overlap is
#                         found, this no longer silently skips the
#                         import. It imports the ride anyway and flags
#                         it (possible_duplicate_of, a new column)
#                         against the ride it overlaps — the app isn't
#                         well-positioned to guess which of two
#                         genuinely independent recordings is the
#                         "right" one, so nothing gets silently
#                         dropped or silently merged; it surfaces for
#                         the rider to resolve. New GET /rides/flagged
#                         to list flagged pairs, POST /rides/{id}/
#                         clear-review to confirm two rides are
#                         genuinely separate (clears the flag, keeps
#                         both) — resolving as "wrong one" still just
#                         uses the existing DELETE /rides/{id}.
#                         Falls back to the original distance/duration-
#                         only check (still an automatic skip,
#                         unchanged) only when start_time or elapsed_h
#                         isn't available — nothing to compute a
#                         window from.
#                         Frontend piece to actually see and resolve
#                         flagged pairs NOT YET BUILT — this ships the
#                         detection/data side only.
#                         Tested thoroughly before shipping: the exact
#                         real scenario end-to-end (staggered start +
#                         14-hour forgotten stop, correctly flagged
#                         against the real ride, not silently
#                         dropped), confirmed this is a genuine
#                         improvement over the previous start-time-only
#                         fix (a 17-min stagger it would have missed),
#                         a genuinely separate same-day ride correctly
#                         NOT flagged, and placeholder/parameter
#                         alignment manually verified in both queries.
#   2.7.0 (2026-07-29) — fixed the actual duplicate-detection bug
#                         behind a real double-count (rides 1077/1078
#                         on 2026-07-29 — Karoo FIT uploaded to Strava
#                         plus a separate phone-recorded Strava
#                         activity of the same physical ride, imported
#                         as two rides, doubling mileage/elevation on
#                         the dashboard). The old check (distance
#                         within 0.5mi, duration within 6min) wasn't
#                         wide enough for two independently-recorded
#                         devices, whose measured distance/duration
#                         can genuinely drift more than that due to
#                         GPS accuracy and differing auto-pause
#                         behavior — even though it's unmistakably the
#                         same ride. Added a new start_time column
#                         (both FIT parsing and Strava sync already
#                         extracted this, just discarded everything
#                         but the date) and made start-time proximity
#                         (within 10 min) the deciding factor whenever
#                         it's available on both sides — start time
#                         barely drifts between devices regardless of
#                         GPS/pause differences. Falls back to the
#                         original distance/duration check only when
#                         start_time isn't available for at least one
#                         side (older data predating this fix).
#                         A real bug was caught and fixed during
#                         testing, not just in the final version:
#                         initially OR'd the new start-time check
#                         alongside the old distance/duration one, but
#                         testing specifically for false positives
#                         found that two genuinely different same-day
#                         rides with a coincidentally similar distance
#                         got wrongly flagged as duplicates — that
#                         risk existed in the original check too, OR'ing
#                         a second condition on top didn't fix it.
#                         Corrected so start-time is authoritative when
#                         available, not just additive. Re-verified
#                         against four cases after the fix: the real
#                         bug scenario (still caught), the false-
#                         positive case (now correctly NOT flagged),
#                         old data with no start_time on either side
#                         (unchanged fallback behavior), and mixed old/
#                         new data with start_time on only one side.
#   2.6.0 (2026-07-29) — added GET /rides/export: full ride history
#                         as a downloadable .xlsx workbook — every
#                         tracked field (distance, power, HR, cadence,
#                         sprint bests, elevation gain/loss, calories,
#                         TSS/IF, L/R balance, equipment, notes),
#                         formatted header row, frozen top row,
#                         auto-sized columns. NEW DEPENDENCY: requires
#                         openpyxl added to requirements.txt on
#                         GitHub — not something I can add myself,
#                         since I only ever have main.py and
#                         cycling_coach_app.html in front of me, not
#                         the actual requirements file. Deploy will
#                         fail with a missing-module error until
#                         that's added.
#                         Tested before shipping: full workbook
#                         generation and save, a realistic ride with
#                         many missing/None fields (handled without
#                         crashing), and a genuine round-trip check —
#                         re-opening the generated file to confirm
#                         it's valid, not just that writing it didn't
#                         error.
#   2.5.1 (2026-07-29) — the coaching chat can now directly answer
#                         "what's the date/time" — added get_local_now()
#                         alongside the date-only helper from the
#                         previous fix, and a clearly-labeled CURRENT
#                         DATE/TIME line at the very top of the system
#                         prompt (rider's real local time, not the
#                         server's), with an explicit instruction to
#                         answer confidently from it rather than guess.
#                         Refactored get_local_today() to derive from
#                         get_local_now() instead of duplicating the
#                         timezone lookup — verified this preserved
#                         identical behavior before shipping.
#   2.5.0 (2026-07-29) — fixed a real timezone bug: the coaching chat
#                         was reasoning about "today" using the
#                         server's UTC clock, not the rider's actual
#                         local time. Railway runs UTC; Houston is
#                         UTC-5 in summer, so anywhere from ~7pm to
#                         midnight local time, the server's date had
#                         already rolled to the next calendar day
#                         while the rider's actual day hadn't ended —
#                         exactly the "it thinks I'm already in
#                         tomorrow morning" symptom reported at 7:20pm
#                         evening. Added a real per-user timezone
#                         field on the profile (defaults to
#                         America/Chicago, correct for the current
#                         beta group) and a get_local_today() helper
#                         using zoneinfo, so daylight saving is handled
#                         automatically rather than a hardcoded offset
#                         that would silently drift wrong twice a
#                         year. Fixed both the coaching chat's date-
#                         relative reasoning and the dashboard's pace-
#                         vs-goal calculation, since both were built on
#                         the same server-UTC date.today() call.
#                         Tested against Marc's exact reported
#                         scenario (a specific UTC timestamp that
#                         reproduces the bug precisely), a winter date
#                         to confirm DST handling needs no special-
#                         casing, an invalid-timezone fallback so a bad
#                         value can't crash the whole page, and — most
#                         convincingly — run live at the actual current
#                         moment: the old behavior gives July 30
#                         (tomorrow), the fix gives July 29 (today),
#                         confirmed in real time, not simulated.
#   2.4.0 (2026-07-27) — two of the ride-detail page's deliberately-
#                         deferred pieces from earlier tonight, done
#                         together: added a left-right power balance
#                         chart (FIT-upload-only, same documented
#                         asymmetry as everywhere else L/R balance
#                         shows up — only appears when the data is
#                         genuinely present, tested against real data,
#                         an absent key, and a present-but-all-null
#                         key, since that last one is a real edge
#                         case). Also added GET /rides/by-date/{date}
#                         — bridges the coaching memory's dated log
#                         (which only has a date, no ride_id) through
#                         to a ride's detail page; Ride History cards
#                         are now clickable. If multiple rides share a
#                         date, prefers whichever one has stream data,
#                         since that gives an actually useful detail
#                         page instead of the "no chart data"
#                         fallback. History entries with no matching
#                         ride (several of tonight's seeded historical
#                         entries genuinely have none) get a clear
#                         inline message instead of a broken link.
#   2.3.1 (2026-07-27) — fixed the Sprint & Aerobic Power chart's
#                         tooltip order — it was showing 5-min at the
#                         top and 30s at the bottom (an explicit
#                         reversed sort, not just dataset order),
#                         backwards from the logical 5s/15s/30s/5-min
#                         reading order. Scoped precisely to this one
#                         chart — three other charts use the same
#                         generic reversed-sort pattern for unrelated
#                         reasons and were deliberately left untouched.
#   2.3.0 (2026-07-27) — first version of the ride-detail page (step 6
#                         of the ride data plan) — deliberately scoped
#                         to a first working slice rather than the
#                         full original spec: single-value stats
#                         (distance, elevation, power, HR, cadence,
#                         calories, TSS/IF, L/R balance, equipment
#                         used) plus four charts (altitude profile,
#                         power, heart rate, cadence) plotted against
#                         distance. Deferred to a later pass: drag-
#                         select-to-recompute interactivity, and
#                         linking from the Ride History cards (which
#                         needs a date-to-ride lookup with a graceful
#                         "no ride found" fallback, since several of
#                         tonight's seeded historical entries have no
#                         corresponding ride row at all). Built as a
#                         server-rendered HTML page shown in an
#                         iframe, same proven pattern as /dashboard,
#                         rather than adding a new client-side
#                         charting dependency to the app shell.
#                         Streams get downsampled for the charts
#                         (~400 points) since a multi-hour ride can be
#                         7,000+ raw samples — full resolution stays
#                         intact for the stats above, this only
#                         affects what gets plotted. Reachable via a
#                         new "View ride details" link after upload,
#                         using the ride_id already tracked from the
#                         equipment picker.
#                         Tested before shipping: downsampling target
#                         verified, full render with realistic stream
#                         data, graceful fallback for rides with no
#                         stream data (pre-dates raw storage), HTML-
#                         escaping verified against an injection
#                         attempt, mismatched-length stream arrays
#                         (a realistic case — different fields can
#                         have slightly different sample counts)
#                         handled without crashing, and the generated
#                         Chart.js JS syntax-checked directly.
#   2.2.0 (2026-07-27) — three pieces built together tonight:
#                         AI TOOL-USE SYSTEM (step 5 of the ride data
#                         plan) — 5 tools (metric-by-mile-range,
#                         push-segment detection, HR/power zone
#                         breakdown, combined numeric+keyword search
#                         across rides and coaching history with
#                         count-first/narrow-on-request behavior, and
#                         direct dated-log lookup), wired into both
#                         the post-upload assessment and the ongoing
#                         coaching chat via a proper tool-use loop
#                         (run_claude_with_tools). Caught and fixed
#                         two real bugs before shipping: ride_id was
#                         never being passed to get_coaching_summary
#                         at all (would have made the new tools
#                         uncallable during the post-upload deep-dive
#                         specifically), and a None system-prompt case
#                         that could have sent a malformed request.
#                         Added search_preferences as a 6th memory
#                         theme, same mechanism as the other five.
#                         PERSISTENT LAST-RIDE SYNOPSIS — the
#                         assessment shown once at upload was
#                         discarded after; now saved (coaching_synopsis
#                         column) and shown as a card at the top of
#                         the Dashboard until the next ride replaces
#                         it. HTML-escaped (verified against a raw
#                         script-tag test case before shipping).
#                         EQUIPMENT/SETUP TRACKING — new equipment
#                         roster table, captured via the profile
#                         interview (indoor trainer as its own entry,
#                         not just outdoor bike variants), picked from
#                         a dropdown per ride rather than typed fresh.
#                         Caught a real bug before it shipped: the
#                         interview's generic profile-field-saving
#                         loop would have tried to insert an
#                         "equipment" list as a literal column on the
#                         profiles table — handled separately now.
#                         Frontend: cycling_coach_app.html gets a new
#                         equipment picker card, shown after a
#                         successful FIT upload.
#                         Tested before shipping: unit tests on all
#                         four new tool algorithms against synthetic
#                         data, simulated tool-use loop control flow,
#                         three dashboard-render tests for the
#                         synopsis card (renders correctly, absent
#                         when there's nothing to show, only the most
#                         recent ride's synopsis appears), branching
#                         tests on the equipment-extraction logic, and
#                         a full integration check importing the whole
#                         module fresh with realistic combined data.
#   2.1.1 (2026-07-26) — added a dedicated /health endpoint, separate
#                         from / (which doubles as the version-check
#                         endpoint). Testing whether that overlap is
#                         related to the persistent stale-version-
#                         serving issue that's outlasted CDN checks,
#                         a manual restart, and two full redeploys.
#                         railway.json's healthcheckPath needs to be
#                         updated to /health to match — not just this
#                         file — for this test to mean anything.
#   2.1.0 (2026-07-24) — fixed the FIT-vs-Strava methodology mismatch
#                         found while testing (rides 874/1071, same
#                         ride, 90W vs 177W avg power and a 32-min
#                         duration gap). FIT uploads now use moving
#                         time (total_timer_time) as the primary
#                         duration, matching Strava sync's existing
#                         convention, with a fallback to elapsed time
#                         if a device doesn't report timer time.
#                         Elapsed time is kept as a new separate field
#                         (elapsed_h) on both paths rather than
#                         discarded, so stoppage time stays
#                         recoverable. avg_power on the FIT path is
#                         now computed from actual non-zero readings,
#                         same as Strava, instead of trusting the
#                         device's own session-level average (which
#                         includes coasting). Also added DELETE
#                         /rides/{id} — there was previously no way
#                         to remove a single bad ride, only the
#                         all-or-nothing /rides/clear.
#   2.0.1 (2026-07-23) — the new fields from 2.0.0 (L/R balance,
#                         elevation loss, calories, TSS/IF) were being
#                         computed and stored correctly but never
#                         actually threaded into what the AI sees —
#                         same class of gap as the 5-min-power miss
#                         from earlier today. Added to both the
#                         post-upload coaching summary and the ongoing
#                         coaching chat's recent-rides context (shown
#                         only when actually present, since most rides
#                         won't have L/R balance). Also fixed elevation
#                         loss printing as "-None ft" when a ride has
#                         no descent data rather than omitting it.
#   2.0.0 (2026-07-23) — RIDE DATA ARCHITECTURE, part 1 of the plan
#                         in Ride_Data_Architecture_Plan.md. Both FIT
#                         upload and Strava sync already read every
#                         data point in a ride to compute averages —
#                         they just threw the raw points away
#                         afterward. Now the raw per-second streams
#                         (power, HR, cadence, altitude, distance,
#                         speed, left-right balance where available)
#                         are kept in a new ride_streams table, not
#                         discarded. This is the foundation everything
#                         else in the plan depends on — nothing
#                         displayed differently yet, but the data now
#                         exists to build on. Also: elevation LOSS is
#                         now tracked alongside gain (wasn't before);
#                         calories, avg left-right balance, and
#                         device-reported TSS/intensity factor are
#                         captured where present; enhanced_altitude/
#                         enhanced_speed are preferred over the plain
#                         FIT fields when both exist (more precise,
#                         same underlying data). Left-right balance,
#                         torque effectiveness, and pedal smoothness
#                         are FIT-upload-only — Strava's public API
#                         doesn't expose those as streams, so Strava-
#                         synced rides won't have them regardless of
#                         what the original device recorded.
#                         Tested before shipping: enhanced-field
#                         preference, elevation gain/loss split, and
#                         stream-building logic verified against
#                         controlled fake data; JSONB serialization
#                         confirmed round-trips cleanly; dashboard
#                         re-rendered against both old- and new-shape
#                         ride records to confirm nothing broke for
#                         existing data.
#   1.9.4 (2026-07-23) — the new edit-log-entry endpoint used PATCH,
#                         and PowerShell's Invoke-RestMethod doesn't
#                         reliably form-encode a hashtable body for
#                         PATCH the way it does for POST — the date
#                         fix request came through with entry_date
#                         empty despite being provided. Switched to
#                         POST on the same path (DELETE stays as-is,
#                         no conflict) — same method that's worked
#                         cleanly for every other write in this app.
#   1.9.3 (2026-07-23) — added PATCH /coaching/memory/log/{id} (fix a
#                         wrong date or summary on an existing dated
#                         entry) and DELETE /coaching/memory/log/{id}
#                         — there was no way to correct or remove a
#                         bad entry before this. Immediate use case:
#                         a seeded entry got dated from a raw UTC
#                         conversation timestamp without converting to
#                         local time, landing it a day off from the
#                         actual ride.
#   1.9.2 (2026-07-23) — coaching_memory_log had no protection against
#                         the same date being logged twice — running a
#                         seed import twice (easy to do by accident)
#                         produced exact-topic duplicates, 18 dates
#                         doubled into 36 rows. Added a unique index
#                         on (user_id, entry_date) with an automatic
#                         one-time cleanup in init_db() that removes
#                         pre-existing duplicates (keeping the more
#                         recent of each pair) before the index locks
#                         in — self-healing on next deploy, no manual
#                         cleanup needed. All three places that write
#                         a dated entry (get_coaching_summary,
#                         coaching_chat, seed_memory) now upsert on
#                         that same date instead of blindly inserting,
#                         so this can't recur going forward — a
#                         second mention of an already-logged date
#                         updates that entry rather than duplicating
#                         it. Themes were already protected this way
#                         from the start; this brings the dated log
#                         up to the same standard.
#   1.9.1 (2026-07-22) — /coaching/memory/seed failed with "Could not
#                         parse extraction result as JSON" — the model
#                         wrapped its response in markdown code fences
#                         despite being told not to (a common LLM
#                         habit that shouldn't have been trusted away
#                         with an instruction alone). Added a shared
#                         extract_json_object() helper that pulls the
#                         JSON out regardless of fences or stray text
#                         around it, used everywhere the memory system
#                         parses AI-generated JSON — the seed endpoint,
#                         and both places /coaching/chat and /upload's
#                         summary extract memory updates.
#   1.9.0 (2026-07-22) — hybrid coaching memory: a dated log (one
#                         distilled entry per ride/episode actually
#                         worth remembering) plus five standing
#                         pattern threads (hydration/fueling, effort
#                         perception, recovery/readiness, environment,
#                         life context) that update over time rather
#                         than growing forever. Both /upload's summary
#                         and /coaching/chat now read this memory into
#                         context and can update it after every
#                         exchange, via a trailing JSON block the AI
#                         produces and the backend parses out (never
#                         shown to the user) — no separate API call
#                         needed for the update. New: GET /coaching/
#                         memory to see what's stored, POST /coaching/
#                         memory/seed to retroactively import
#                         historical conversation content using the
#                         same distillation approach, so this can
#                         start from real history instead of zero.
#                         Cost impact at current scale is negligible —
#                         a modest addition to what's already sent per
#                         message, no unbounded growth over time since
#                         old raw content isn't retained, only the
#                         compressed memory it produced.
#   1.8.0 (2026-07-22) — coaching got noticeably deeper and more
#                         curious, closer to a real coaching session:
#                         both /upload's post-ride summary and the
#                         ongoing /coaching/chat now include NP, max
#                         HR, and 5s/15s/30s/5-min power bests per
#                         ride (already computed, never exposed
#                         before). Coach now proactively draws on five
#                         themes when relevant and missing — hydration
#                         /fueling, effort vs. perceived exertion,
#                         recovery/readiness, environmental context,
#                         and life context shaping the ride — as
#                         judgment to apply, not a script to run
#                         through. New-ride discussion gets real depth
#                         instead of a length cap. Garbled voice-to-
#                         text gets flagged and clarified rather than
#                         silently guessed at.
#                         NOT included: push/surge detection, HR
#                         zone time-in-zone, start/end temperature —
#                         those need real second-by-second FIT
#                         parsing that doesn't exist yet, a separate
#                         feature. Long-range callbacks across past
#                         sessions still need the memory feature.
#   1.7.1 (2026-07-22) — Strava connect now forces the "Authorize as
#                         [Name]" confirmation screen every time
#                         (approval_prompt: auto -> force). Before
#                         this, if a browser already had a live
#                         Strava session, connecting silently reused
#                         it with no confirmation — risky on a shared
#                         or borrowed device. Now beta testers always
#                         see and confirm which Strava account is
#                         being connected.
#   1.7.0 (2026-07-19) — added 5-minute best power (p300), computed
#                         the same way as the existing 5s/15s/30s
#                         bests. New DB column, captured on both FIT
#                         upload and Strava sync. Shown on the Sprint
#                         Power chart as a separate purple line
#                         overlaid on the existing stacked bars —
#                         didn't touch the 5s/15s/30s stack itself,
#                         since 5-min is a different kind of metric
#                         (aerobic capacity, not anaerobic burst) and
#                         forcing it into the same stack would've
#                         misrepresented both.
#   1.6.0 (2026-07-19) — added 1M/3M/6M/YTD/All range buttons above
#                         the 7 zoomable charts. Buttons jump the
#                         zoom window via chart.zoomScale() — free
#                         pinch/pan still works from wherever a
#                         preset lands you, it's not a hard boundary.
#                         Split the old single charts-grid into
#                         "Per-Ride Trends" (its own section-header)
#                         plus the existing Coaching Analytics grid,
#                         each with its own range bar, both wired to
#                         all 7 charts so they stay in sync.
#   1.5.0 (2026-07-19) — pinch/scroll zoom + pan added to the 7
#                         per-ride charts (elevation, power, HR, and
#                         the 4 coaching-analytics charts) via
#                         chartjs-plugin-zoom. Weekly/monthly/ride-
#                         type charts untouched — not needed there.
#                         Each zoomable chart has a "reset zoom" link.
#   1.4.0 (2026-07-19) — added GET /admin/users (roster of everyone
#                         signed up, ride counts, Strava/profile
#                         status) — restricted to ADMIN_EMAILS
#   1.3.0 (2026-07-19) — added document import: POST /coaching/import,
#                         GET /coaching/imports, DELETE /coaching/
#                         imports/{id}. Text/markdown only (handoff
#                         docs, training notes) — deliberately NOT for
#                         medical records/lab results, enforced both
#                         in the UI copy and as an AI-level guardrail.
#                         Also trimmed coaching chat replies — was
#                         restating the rider's own numbers back to
#                         them before getting to the point.
#   1.2.1 (2026-07-19) — /coaching/chat now includes year-to-date
#                         totals and pace-vs-goal (it only had the
#                         last 5 individual rides before, so it
#                         couldn't answer "how am I doing on miles")
#   1.2.0 (2026-07-19) — added POST /coaching/chat: real post-ride
#                         coaching conversation (not the profile
#                         interview) — pulls profile + last 5 rides +
#                         recent notes into an ongoing chat
#   1.1.0 (2026-07-19) — Strava sync bounded to YEAR regardless of
#                         days_back (fixes 2024/2025 leak); dedup
#                         widened to +/-1 day tolerance to catch
#                         FIT-vs-Strava date drift; added GET /notes
#                         to verify saved personal notes actually saved
#   1.0.0                initial live build — dashboard, FIT upload,
#                         Strava OAuth + sync, AI profile interview
# ═══════════════════════════════════════════════════════════════════
APP_VERSION = "2.14.0"
ADMIN_EMAILS = {"mtpujol@gmail.com"}

from fastapi import FastAPI, UploadFile, File, HTTPException, Depends, Form
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import psycopg2
import psycopg2.extras
import hashlib
import secrets
import os
import json
import re
import html
import tempfile
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from collections import defaultdict
import httpx

app = FastAPI(title="Cycling Coach API")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

DATABASE_URL  = os.environ.get("DATABASE_URL", "")
ANTHROPIC_KEY       = os.environ.get("ANTHROPIC_API_KEY", "")
STRAVA_CLIENT_ID    = os.environ.get("STRAVA_CLIENT_ID", "266143")
STRAVA_CLIENT_SECRET= os.environ.get("STRAVA_CLIENT_SECRET", "")
STRAVA_REDIRECT_URI = "https://cycling-app-production.up.railway.app/strava/callback"
STRAVA_AUTH_URL     = "https://www.strava.com/oauth/authorize"
STRAVA_TOKEN_URL    = "https://www.strava.com/oauth/token"
security      = HTTPBearer()

ANNUAL_GOAL    = 6500
WEEKLY_TARGET  = 125
YEAR           = 2026

# Strava activity types that count as cycling — everything else (Walk,
# Run, Hike, Swim, etc.) gets skipped during sync. Covers both sport_type
# (Strava's newer, more specific field) and the older, simpler type field.
CYCLING_ACTIVITY_TYPES = {
    "Ride", "VirtualRide", "EBikeRide", "GravelRide", "MountainBikeRide", "Handcycle"
}

# Coaching memory themes — standing patterns tracked across sessions, separate
# from the dated log (which is per-ride/per-episode). Keys are the DB values;
# labels are what gets shown to the AI in context.
MEMORY_THEMES = {
    'hydration_fueling':     'Hydration & Fueling',
    'effort_perception':     'Effort & Perceived Exertion',
    'recovery_readiness':    'Recovery & Readiness',
    'environmental_context': 'Environmental Context',
    'life_context':          'Life Context',
    'search_preferences':    'Search & Filter Preferences',
}

def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = True
    return conn

def init_db():
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY, email TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL, password TEXT NOT NULL,
            token TEXT UNIQUE, created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS rides (
            id SERIAL PRIMARY KEY, user_id INTEGER REFERENCES users(id),
            ride_date DATE, name TEXT, dist_mi FLOAT, duration_h FLOAT,
            avg_power INTEGER, norm_power INTEGER, avg_hr INTEGER, max_hr INTEGER,
            avg_cadence INTEGER, max_cadence INTEGER,
            p5 INTEGER, p15 INTEGER, p30 INTEGER,
            elev_gain_ft FLOAT, ride_type TEXT DEFAULT 'General',
            is_virtual BOOLEAN DEFAULT FALSE, temp_c FLOAT,
            notes TEXT, created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS strava_tokens (
            id SERIAL PRIMARY KEY,
            user_id INTEGER UNIQUE REFERENCES users(id),
            athlete_id BIGINT,
            access_token TEXT,
            refresh_token TEXT,
            expires_at BIGINT,
            last_sync TIMESTAMP,
            created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS profiles (
            id SERIAL PRIMARY KEY,
            user_id INTEGER UNIQUE REFERENCES users(id),
            age INTEGER, weight_lbs FLOAT, location TEXT,
            fitness_level TEXT, ftp INTEGER,
            annual_goal_mi INTEGER, other_goals TEXT,
            health_notes TEXT, injuries TEXT,
            heat_tolerance TEXT, medical_clearance BOOLEAN DEFAULT FALSE,
            interview_complete BOOLEAN DEFAULT FALSE,
            raw_interview TEXT,
            updated_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS coaching_notes (
            id SERIAL PRIMARY KEY, user_id INTEGER REFERENCES users(id),
            note TEXT, created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS equipment (
            id SERIAL PRIMARY KEY, user_id INTEGER REFERENCES users(id),
            name TEXT, created_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(user_id, name)
        );
        CREATE TABLE IF NOT EXISTS imported_docs (
            id SERIAL PRIMARY KEY, user_id INTEGER REFERENCES users(id),
            filename TEXT, content TEXT, created_at TIMESTAMP DEFAULT NOW()
        );
        CREATE TABLE IF NOT EXISTS coaching_memory_log (
            id SERIAL PRIMARY KEY, user_id INTEGER REFERENCES users(id),
            entry_date DATE, summary TEXT, created_at TIMESTAMP DEFAULT NOW()
        );
        -- Clean up any pre-existing duplicate dates (keep the most recent per
        -- user+date) before locking in the uniqueness guarantee below. Safe to
        -- run every startup — a no-op once there's nothing left to dedupe.
        DELETE FROM coaching_memory_log a USING coaching_memory_log b
            WHERE a.id < b.id AND a.user_id = b.user_id AND a.entry_date = b.entry_date;
        CREATE UNIQUE INDEX IF NOT EXISTS idx_memory_log_user_date
            ON coaching_memory_log (user_id, entry_date);
        CREATE TABLE IF NOT EXISTS coaching_memory_themes (
            id SERIAL PRIMARY KEY, user_id INTEGER REFERENCES users(id),
            theme TEXT, content TEXT, updated_at TIMESTAMP DEFAULT NOW(),
            UNIQUE(user_id, theme)
        );
        CREATE TABLE IF NOT EXISTS ride_streams (
            id SERIAL PRIMARY KEY,
            ride_id INTEGER UNIQUE REFERENCES rides(id) ON DELETE CASCADE,
            streams JSONB,
            created_at TIMESTAMP DEFAULT NOW()
        );
        -- Add missing columns if upgrading from old schema
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS elev_gain_ft FLOAT;
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS ride_type TEXT DEFAULT 'General';
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS is_virtual BOOLEAN DEFAULT FALSE;
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS p300 INTEGER;
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS elev_loss_ft FLOAT;
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS calories INTEGER;
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS avg_lr_balance FLOAT;
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS training_stress_score FLOAT;
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS intensity_factor FLOAT;
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS elapsed_h FLOAT;
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS coaching_synopsis TEXT;
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS equipment_id INTEGER REFERENCES equipment(id);
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS start_time TIMESTAMP;
        ALTER TABLE rides ADD COLUMN IF NOT EXISTS possible_duplicate_of INTEGER REFERENCES rides(id);
        ALTER TABLE profiles ADD COLUMN IF NOT EXISTS timezone TEXT;
    """)
    cur.close(); conn.close()

def hash_password(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def get_local_now(tz_name=None):
    """The user's actual current local date and time, not the server's
    UTC clock. Same reasoning as get_local_today() — Railway runs UTC,
    Houston is UTC-5 in summer, so the raw server clock reads a
    different wall-clock time than the rider's actual one."""
    tz_name = tz_name or 'America/Chicago'
    try:
        return datetime.now(ZoneInfo(tz_name))
    except Exception:
        return datetime.now()

def get_local_today(tz_name=None):
    """The user's actual local calendar date, not the server's. Railway
    runs UTC — during Houston evening hours (roughly 7pm-midnight CDT),
    UTC has already rolled to the next calendar day, which was making
    the coaching chat think "today" was a day ahead of the rider's
    actual day. Defaults to America/Chicago (Houston) when a user
    hasn't set a timezone; falls back to server date only if the
    timezone name itself is somehow invalid."""
    return get_local_now(tz_name).date()

def extract_json_object(text):
    """Pull a JSON object out of AI output that may be wrapped in markdown code
    fences or have stray text around it (models don't always follow a "JSON
    only" instruction to the letter). Raises if nothing parseable is found."""
    match = re.search(r'\{.*\}', text, re.DOTALL)
    if not match:
        raise ValueError("No JSON object found in text")
    return json.loads(match.group(0))

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM users WHERE token = %s", (credentials.credentials,))
    user = cur.fetchone(); cur.close(); conn.close()
    if not user:
        raise HTTPException(status_code=401, detail="Invalid token")
    return user

def classify_ride(dist_mi, duration_h, avg_power, is_virtual):
    """Auto-classify ride type based on metrics."""
    if is_virtual:
        return 'General'
    if dist_mi and dist_mi >= 62:
        return 'Long Ride (100km+)'
    if avg_power and avg_power > 200:
        return 'Threshold'
    if duration_h and duration_h >= 3:
        return 'Aerobic Endurance'
    if dist_mi and dist_mi < 10:
        return 'Recovery/Rehab'
    return 'General'

def parse_fit_bytes(data):
    try:
        import fitparse
        with tempfile.NamedTemporaryFile(suffix='.fit', delete=False) as tmp:
            tmp.write(data); tmp_path = tmp.name
        ff = fitparse.FitFile(tmp_path)
        session = {}; records = []
        for msg in ff.get_messages():
            if msg.name == 'session':
                for f in msg.fields: session[f.name] = f.value
            elif msg.name == 'record':
                r = {}
                for f in msg.fields: r[f.name] = f.value
                records.append(r)
        os.unlink(tmp_path)

        # Prefer the "enhanced" altitude/speed fields when present — same
        # data, wider range and finer precision than the plain fields.
        def _alt(r):
            v = r.get('enhanced_altitude')
            return v if v is not None else r.get('altitude')
        def _spd(r):
            v = r.get('enhanced_speed')
            return v if v is not None else r.get('speed')
        def _ts(v):
            return v.isoformat() if hasattr(v, 'isoformat') else v

        powers   = [r['power']      for r in records if r.get('power')      and r['power'] > 0]
        cadences = [r['cadence']    for r in records if r.get('cadence')]
        alts     = [_alt(r) for r in records if _alt(r) is not None]

        def best_avg(vals, n):
            if not vals or len(vals) < n: return max(vals) if vals else None
            return round(max(sum(vals[i:i+n])/n for i in range(len(vals)-n+1)))

        np_val = None
        if powers and len(powers) > 30:
            smoothed = [sum(powers[max(0,i-29):i+1])/len(powers[max(0,i-29):i+1]) for i in range(len(powers))]
            np_val = round((sum(x**4 for x in smoothed)/len(smoothed))**0.25)

        # Elevation gain AND loss from altitude records (loss wasn't
        # captured before — descent matters for anywhere with real climbing)
        elev_gain_m = 0.0; elev_loss_m = 0.0
        if alts and len(alts) > 1:
            for i in range(1, len(alts)):
                diff = alts[i] - alts[i-1]
                if diff > 0: elev_gain_m += diff
                elif diff < 0: elev_loss_m += -diff
        elev_gain_ft = round(elev_gain_m * 3.28084, 0) if elev_gain_m else None
        elev_loss_ft = round(elev_loss_m * 3.28084, 0) if elev_loss_m else None

        # Raw per-second streams — kept instead of discarded after the
        # aggregates above are computed, so future tools/charts/AI queries
        # have something to work with. left_right_balance and the torque/
        # smoothness fields are included defensively — present only if the
        # power meter and head unit combination actually supports them.
        streams = {
            'timestamp':            [_ts(r.get('timestamp')) for r in records],
            'distance':              [r.get('distance') for r in records],
            'altitude':              [_alt(r) for r in records],
            'speed':                 [_spd(r) for r in records],
            'power':                 [r.get('power') for r in records],
            'heart_rate':            [r.get('heart_rate') for r in records],
            'cadence':               [r.get('cadence') for r in records],
            'left_right_balance':    [r.get('left_right_balance') for r in records],
            'left_torque_eff':       [r.get('left_torque_effectiveness') for r in records],
            'right_torque_eff':      [r.get('right_torque_effectiveness') for r in records],
            'left_pedal_smooth':     [r.get('left_pedal_smoothness') for r in records],
            'right_pedal_smooth':    [r.get('right_pedal_smoothness') for r in records],
        }
        # Session-level L/R balance average, if the field is present anywhere
        lr_vals = [v for v in streams['left_right_balance'] if v is not None]
        avg_lr_balance = round(sum(lr_vals) / len(lr_vals), 1) if lr_vals else None

        start   = session.get('start_time')
        dist    = session.get('total_distance')
        # Moving time (auto-pause excluded) as the primary duration — matches
        # Strava sync's convention (moving_time), so a ride's numbers don't
        # depend on which path it came in through. Elapsed time (full wall-
        # clock, stops included) is kept as a separate field rather than
        # discarded, so stoppage time is always recoverable if it matters.
        moving_seconds  = session.get('total_timer_time')
        elapsed_seconds = session.get('total_elapsed_time')
        sport      = str(session.get('sport', '')).lower()
        sub_sport  = str(session.get('sub_sport', '')).lower()
        # Check file_id for manufacturer (Zwift shows as manufacturer=zwift)
        manufacturer = ''
        for msg2 in ff.get_messages('file_id'):
            for f2 in msg2.fields:
                if f2.name == 'manufacturer' and f2.value:
                    manufacturer = str(f2.value).lower()
        is_virtual = (
            'virtual' in sport or 'indoor' in sport
            or 'virtual' in sub_sport
            or 'zwift' in manufacturer
            or 'zwift' in sport
        )

        dist_mi   = round(dist/1609.34, 2) if dist else None
        # Fall back to elapsed time if a device doesn't report timer time —
        # better an approximate duration than none at all.
        duration_h = round((moving_seconds or elapsed_seconds)/3600, 2) if (moving_seconds or elapsed_seconds) else None
        elapsed_h  = round(elapsed_seconds/3600, 2) if elapsed_seconds else None
        # Computed from actual non-zero readings, same as the Strava path —
        # not the device's own session-level average, which includes
        # coasting and made this number disagree with Strava-synced rides.
        avg_power_val = round(sum(powers)/len(powers)) if powers else None
        ride_type = classify_ride(dist_mi, duration_h, avg_power_val, is_virtual)

        return {
            'ride_date':   start.strftime('%Y-%m-%d') if hasattr(start,'strftime') else str(start)[:10],
            'start_time':  start.isoformat() if hasattr(start,'isoformat') else None,
            'sport':       sport,
            'name':        session.get('sport', 'Ride'),
            'dist_mi':     dist_mi,
            'duration_h':  duration_h,
            'avg_power':   avg_power_val,
            'norm_power':  session.get('normalized_power') or np_val,
            'avg_hr':      session.get('avg_heart_rate'),
            'max_hr':      session.get('max_heart_rate'),
            'avg_cadence': session.get('avg_cadence'),
            'max_cadence': max(cadences) if cadences else None,
            'p5':   best_avg(powers, 5),
            'p15':  best_avg(powers, 15),
            'p30':  best_avg(powers, 30),
            'p300': best_avg(powers, 300),
            'elev_gain_ft': elev_gain_ft,
            'elev_loss_ft': elev_loss_ft,
            'elapsed_h':    elapsed_h,
            'calories':     session.get('total_calories'),
            'avg_lr_balance': avg_lr_balance,
            'training_stress_score': session.get('training_stress_score'),
            'intensity_factor':      session.get('intensity_factor'),
            'ride_type':   ride_type,
            'is_virtual':  is_virtual,
            'temp_c':      session.get('avg_temperature'),
            'streams':     streams,
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail="Could not parse FIT file: " + str(e))

def build_full_dashboard(rides, name, annual_goal=None, user_timezone=None):
    """Build the full Cycling Analytics dashboard matching the local version."""
    import json as _json

    goal = annual_goal or ANNUAL_GOAL
    sorted_rides = sorted(rides, key=lambda r: r['ride_date'] if r.get('ride_date') else date.min)

    synopsis_card = ""
    if sorted_rides:
        latest = sorted_rides[-1]
        synopsis = latest.get('coaching_synopsis')
        if synopsis:
            latest_date = latest.get('ride_date')
            date_str = latest_date.strftime('%B %d, %Y') if hasattr(latest_date, 'strftime') else str(latest_date)
            synopsis_card = (
                "<div class='synopsis-card'>"
                + "<div class='synopsis-label'>Latest Ride Assessment</div>"
                + "<div class='synopsis-date'>" + html.escape(date_str) + " &nbsp;&middot;&nbsp; "
                + html.escape(str(latest.get('name', 'Ride'))) + "</div>"
                + "<div class='synopsis-text'>" + html.escape(synopsis) + "</div>"
                + "</div>"
            )

    def to_date(d):
        if isinstance(d, date): return d
        if isinstance(d, str): return date.fromisoformat(str(d)[:10])
        return None

    def dur_hrs(r):
        return float(r.get('duration_h') or 0)

    total_mi   = sum(float(r.get('dist_mi') or 0) for r in rides)
    total_hrs  = sum(dur_hrs(r) for r in rides)
    total_elev = sum(float(r.get('elev_gain_ft') or 0) for r in rides)
    virt_count = sum(1 for r in rides if r.get('is_virtual'))
    out_count  = len(rides) - virt_count

    week_start = date(YEAR-1, 12, 29)
    weeks = {}
    for i in range(53):
        ws = week_start + timedelta(weeks=i)
        weeks[ws] = 0.0
    for r in sorted_rides:
        d = to_date(r.get('ride_date'))
        if not d: continue
        mon = d - timedelta(days=d.weekday())
        if mon in weeks:
            weeks[mon] = weeks.get(mon, 0) + float(r.get('dist_mi') or 0)
    week_labels = [k.strftime('%b %d') for k in sorted(weeks)]
    week_miles  = [round(weeks[k], 1) for k in sorted(weeks)]
    week_target = [WEEKLY_TARGET] * len(week_labels)
    cum_actual  = []
    cum_target  = []
    running = 0.0
    for i, m in enumerate(week_miles):
        running += m
        cum_actual.append(round(running, 1))
        cum_target.append(round(goal * (i+1) / 53, 1))

    mo_names = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
    mo_keys  = [f'{YEAR}-{m:02d}' for m in range(1,13)]
    mo_mi    = defaultdict(float)
    mo_hrs   = defaultdict(float)
    for r in sorted_rides:
        d = to_date(r.get('ride_date'))
        if not d: continue
        mk = d.strftime('%Y-%m')
        mo_mi[mk]  += float(r.get('dist_mi') or 0)
        mo_hrs[mk] += dur_hrs(r)
    month_miles = [round(mo_mi[k], 1) for k in mo_keys]
    month_hours = [round(mo_hrs[k], 1) for k in mo_keys]

    all_types = ['General','Casual/Social','Aerobic Endurance',
                 'Long Ride (100km+)','Threshold','Hard/Intervals','Recovery/Rehab']
    type_mi  = defaultdict(float)
    type_hrs = defaultdict(float)
    for r in sorted_rides:
        rt = r.get('ride_type') or 'General'
        type_mi[rt]  += float(r.get('dist_mi') or 0)
        type_hrs[rt] += dur_hrs(r)
    rtype_mi_vals  = [round(type_mi[t], 1)  for t in all_types]
    rtype_hrs_vals = [round(type_hrs[t], 1) for t in all_types]

    virt_mi  = round(sum(float(r.get('dist_mi') or 0) for r in rides if r.get('is_virtual')), 1)
    out_mi   = round(sum(float(r.get('dist_mi') or 0) for r in rides if not r.get('is_virtual')), 1)
    virt_hrs = round(sum(dur_hrs(r) for r in rides if r.get('is_virtual')), 1)
    out_hrs  = round(sum(dur_hrs(r) for r in rides if not r.get('is_virtual')), 1)

    ride_dates = []
    ride_dates_iso = []
    ride_power = []
    ride_hr    = []
    ride_elev  = []
    for r in sorted_rides:
        d = to_date(r.get('ride_date'))
        if not d: continue
        ride_dates.append(d.strftime('%b %d'))
        ride_dates_iso.append(d.isoformat())
        ride_power.append(r.get('avg_power'))
        ride_hr.append(r.get('avg_hr'))
        ride_elev.append(float(r.get('elev_gain_ft') or 0))

    coach_rides  = [r for r in sorted_rides if float(r.get('dist_mi') or 0) >= 5]
    coach_dates  = [to_date(r['ride_date']).strftime('%b %d') for r in coach_rides if to_date(r.get('ride_date'))]
    coach_dates_iso = [to_date(r['ride_date']).isoformat() for r in coach_rides if to_date(r.get('ride_date'))]
    coach_avgpwr = [r.get('avg_power')   for r in coach_rides]
    coach_np     = [r.get('norm_power')  for r in coach_rides]
    coach_avghr  = [r.get('avg_hr')      for r in coach_rides]
    coach_maxhr  = [r.get('max_hr')      for r in coach_rides]
    coach_avgcad = [r.get('avg_cadence') for r in coach_rides]
    coach_maxcad = [r.get('max_cadence') for r in coach_rides]
    coach_p5     = [r.get('p5')  for r in coach_rides]
    coach_p10    = [r.get('p15') for r in coach_rides]
    coach_p20    = [r.get('p30') for r in coach_rides]
    coach_p300   = [r.get('p300') for r in coach_rides]
    coach_p10_mid = [(int(p10 or 0) - int(p20 or 0)) if p10 and p20 else None for p10,p20 in zip(coach_p10, coach_p20)]
    coach_p5_top  = [(int(p5  or 0) - int(p10 or 0)) if p5  and p10 else None for p5, p10 in zip(coach_p5,  coach_p10)]

    today        = get_local_today(user_timezone)
    day_of_year  = today.timetuple().tm_yday
    pace_mi      = round(goal * day_of_year / 366, 1)
    pace_diff    = round(total_mi - pace_mi, 1)
    pace_ahead   = pace_diff >= 0
    pct_complete = round(total_mi / goal * 100, 1) if goal else 0
    remaining    = round(goal - total_mi, 1)

    def j(v): return _json.dumps(v)

    pace_color = '#27AE60' if pace_ahead else '#E67E22'
    pace_word  = 'ahead of' if pace_ahead else 'behind'
    pace_bg    = '#E8F8F0' if pace_ahead else '#FEF0E8'

    if goal:
        goal_cards = (
            "<div class='stat-card'><div class='label'>Remaining</div>"
            + "<div class='value'>" + str(round(remaining,1)) + "</div>"
            + "<div class='sub'>miles to " + str(goal) + "</div></div>"
            + "<div class='stat-card " + ("green" if pace_ahead else "orange") + "'>"
            + "<div class='label'>Pace</div>"
            + "<div class='value'>" + str(abs(pace_diff)) + "</div>"
            + "<div class='sub'>miles " + pace_word + " pace</div></div>"
        )
        goal_subtitle = " &nbsp;&middot;&nbsp; Goal: " + str(goal) + " miles"
        goal_progress = (
            "<div class='progress-wrap'>"
            + "<div class='progress-label'>"
            + "<span><strong>" + str(round(total_mi,1)) + " mi</strong> completed</span>"
            + "<span>Goal: <strong>" + str(goal) + " mi</strong></span>"
            + "</div>"
            + "<div class='progress-bar-bg'>"
            + "<div class='progress-bar-fill' style='width:" + str(min(pct_complete,100)) + "%'></div>"
            + "</div></div>"
        )
    else:
        goal_cards = ""
        goal_subtitle = ""
        goal_progress = ""

    js_weekly = (
        "barChart('weeklyBar'," + j(week_labels) + ","
        "[{label:'Miles',data:" + j(week_miles) + ",backgroundColor:BLUE+'CC'},"
        "{label:'Target (" + str(WEEKLY_TARGET) + ")',data:" + j(week_target) + ",type:'line',"
        "borderColor:ORANGE,borderDash:[6,3],borderWidth:2,pointRadius:0,fill:false}]);"
    )
    js_cumul = (
        "lineChart('cumulativeLine'," + j(week_labels) + ","
        "[{label:'Actual',data:" + j(cum_actual) + ",borderColor:BLUE,backgroundColor:BLUE+'20',fill:true},"
        "{label:'Target pace',data:" + j(cum_target) + ",borderColor:ORANGE,borderDash:[6,3],borderWidth:2,pointRadius:0}]);"
    )
    js_mo_mi = "barChart('monthlyBar'," + j(mo_names) + ",[{label:'Miles',data:" + j(month_miles) + ",backgroundColor:BLUE+'CC'}]);"
    js_mo_hr = "barChart('monthlyHours'," + j(mo_names) + ",[{label:'Hours',data:" + j(month_hours) + ",backgroundColor:PURPLE+'CC'}]);"
    js_rt_mi = "barChart('rtypeMiles'," + j(all_types) + ",[{label:'Miles',data:" + j(rtype_mi_vals) + ",backgroundColor:TYPE_COLORS.map(c=>c+'CC')}],{indexAxis:'y'});"
    js_rt_hr = "barChart('rtypeHours'," + j(all_types) + ",[{label:'Hours',data:" + j(rtype_hrs_vals) + ",backgroundColor:TYPE_COLORS.map(c=>c+'CC')}],{indexAxis:'y'});"
    max_mi = max(out_mi + virt_mi, 1)
    max_hr = max(out_hrs + virt_hrs, 1)
    pct_om = round(out_mi / max_mi * 100, 1)
    pct_vm = max(round(virt_mi / max_mi * 100, 1), 2)
    pct_oh = round(out_hrs / max_hr * 100, 1)
    pct_vh = max(round(virt_hrs / max_hr * 100, 1), 2)
    _om = str(out_mi); _vm = str(virt_mi); _oh = str(out_hrs); _vh = str(virt_hrs)
    _pm = str(round(max_mi/2)); _ph = str(round(max_hr/2))
    _tm = str(round(max_mi)); _th = str(round(max_hr))
    virt_html = (
        "<div style='font-size:11px;color:#555;margin-bottom:8px;'>"
        + "<span style='display:inline-flex;align-items:center;gap:4px;margin-right:12px;'>"
        + "<span style='width:10px;height:10px;border-radius:2px;background:#27AE60;display:inline-block;'></span>Outdoor</span>"
        + "<span style='display:inline-flex;align-items:center;gap:4px;'>"
        + "<span style='width:10px;height:10px;border-radius:2px;background:#9B59B6;display:inline-block;'></span>Virtual</span></div>"
        + "<div style='font-size:10px;color:#999;display:flex;justify-content:space-between;padding-left:84px;margin-bottom:2px;'>"
        + "<span>0</span><span>" + _pm + "</span>"
        + "<span style='color:#1F4E79;font-weight:600;'>Miles</span>"
        + "<span>" + _tm + "</span></div>"
        + "<div style='border-left:1.5px solid #ccc;margin-left:84px;padding:2px 0;'>"
        + "<div style='display:flex;align-items:center;position:relative;height:26px;margin-bottom:5px;'>"
        + "<span style='position:absolute;left:-88px;font-size:11px;color:#555;width:84px;text-align:right;padding-right:6px;'>Outdoor mi</span>"
        + "<div style='height:22px;width:" + str(pct_om) + "%;background:#27AE60CC;border-radius:0 4px 4px 0;display:flex;align-items:center;padding:0 8px;'>"
        + "<span style='font-size:11px;font-weight:600;color:#fff;'>" + _om + "</span></div></div>"
        + "<div style='display:flex;align-items:center;position:relative;height:26px;margin-bottom:5px;'>"
        + "<span style='position:absolute;left:-88px;font-size:11px;color:#555;width:84px;text-align:right;padding-right:6px;'>Virtual mi</span>"
        + "<div style='height:22px;width:" + str(pct_vm) + "%;background:#9B59B6CC;border-radius:0 4px 4px 0;display:flex;align-items:center;padding:0 8px;min-width:38px;'>"
        + "<span style='font-size:11px;font-weight:600;color:#fff;'>" + _vm + "</span></div></div></div>"
        + "<div style='height:10px;'></div>"
        + "<div style='font-size:10px;color:#999;display:flex;justify-content:space-between;padding-left:84px;margin-bottom:2px;'>"
        + "<span>0</span><span>" + _ph + "</span>"
        + "<span style='color:#1F4E79;font-weight:600;'>Hours</span>"
        + "<span>" + _th + "</span></div>"
        + "<div style='border-left:1.5px solid #ccc;margin-left:84px;padding:2px 0;'>"
        + "<div style='display:flex;align-items:center;position:relative;height:26px;margin-bottom:5px;'>"
        + "<span style='position:absolute;left:-88px;font-size:11px;color:#555;width:84px;text-align:right;padding-right:6px;'>Outdoor hr</span>"
        + "<div style='height:22px;width:" + str(pct_oh) + "%;background:#27AE60CC;border-radius:0 4px 4px 0;display:flex;align-items:center;padding:0 8px;'>"
        + "<span style='font-size:11px;font-weight:600;color:#fff;'>" + _oh + "</span></div></div>"
        + "<div style='display:flex;align-items:center;position:relative;height:26px;'>"
        + "<span style='position:absolute;left:-88px;font-size:11px;color:#555;width:84px;text-align:right;padding-right:6px;'>Virtual hr</span>"
        + "<div style='height:22px;width:" + str(pct_vh) + "%;background:#9B59B6CC;border-radius:0 4px 4px 0;display:flex;align-items:center;padding:0 8px;min-width:36px;'>"
        + "<span style='font-size:11px;font-weight:600;color:#fff;'>" + _vh + "</span></div></div></div>"
    )
    js_virt = ""
    range_bar_html = (
        "<div class='range-bar'><span>Range:</span>"
        + "<button class='range-btn' data-range='30' onclick=\"setRange('30')\">1M</button>"
        + "<button class='range-btn' data-range='90' onclick=\"setRange('90')\">3M</button>"
        + "<button class='range-btn' data-range='182' onclick=\"setRange('182')\">6M</button>"
        + "<button class='range-btn' data-range='ytd' onclick=\"setRange('ytd')\">YTD</button>"
        + "<button class='range-btn active' data-range='all' onclick=\"setRange('all')\">All</button>"
        + "</div>"
    )
    js_elev  = "barChart('elevBar'," + j(ride_dates) + ",[{label:'Elev Gain (ft)',data:" + j(ride_elev) + ",backgroundColor:ORANGE+'CC'}],{zoomable:true});"
    js_pwr   = "lineChart('powerLine'," + j(ride_dates) + ",[{label:'Avg Power (W)',data:" + j(ride_power) + ",borderColor:RED,backgroundColor:RED+'20',fill:false,spanGaps:true}],{zoomable:true});"
    js_hr    = "lineChart('hrLine'," + j(ride_dates) + ",[{label:'Avg HR (bpm)',data:" + j(ride_hr) + ",borderColor:'#E91E63',backgroundColor:'#E91E6320',fill:false,spanGaps:true}],{zoomable:true});"

    js_coach_pwr = (
        "lineChart('coachPower'," + j(coach_dates) + ","
        "[{label:'Avg Power (W)',data:" + j(coach_avgpwr) + ",borderColor:BLUE,backgroundColor:BLUE+'20',fill:false,spanGaps:true},"
        "{label:'Norm Power (W)',data:" + j(coach_np) + ",borderColor:'#1a5276',borderDash:[6,3],borderWidth:2,pointRadius:2,fill:false,spanGaps:true}],"
        "{zoomable:true,plugins:{tooltip:{mode:'index',intersect:false,itemSort:function(a,b){return b.datasetIndex-a.datasetIndex;}}}});"
    )
    js_coach_hr = (
        "lineChart('coachHR'," + j(coach_dates) + ","
        "[{label:'Avg HR (bpm)',data:" + j(coach_avghr) + ",borderColor:'#E91E63',backgroundColor:'#E91E6320',fill:false,spanGaps:true},"
        "{label:'Max HR (bpm)',data:" + j(coach_maxhr) + ",borderColor:'#880e4f',borderDash:[6,3],borderWidth:2,pointRadius:2,fill:false,spanGaps:true}],"
        "{zoomable:true,plugins:{tooltip:{mode:'index',intersect:false,itemSort:function(a,b){return b.datasetIndex-a.datasetIndex;}}}});"
    )
    js_coach_cad = (
        "lineChart('coachCad'," + j(coach_dates) + ","
        "[{label:'Avg Cadence (rpm)',data:" + j(coach_avgcad) + ",borderColor:'#E67E22',backgroundColor:'#E67E2220',fill:false,spanGaps:true},"
        "{label:'Max Cadence (rpm)',data:" + j(coach_maxcad) + ",borderColor:'#784212',borderDash:[6,3],borderWidth:2,pointRadius:2,fill:false,spanGaps:true}],"
        "{zoomable:true,plugins:{tooltip:{mode:'index',intersect:false,itemSort:function(a,b){return b.datasetIndex-a.datasetIndex;}}}});"
    )
    js_sprint_p5  = j(coach_p5)
    js_sprint_p10 = j(coach_p10)
    js_sprint_p20 = j(coach_p20)
    js_sprint_p300 = j(coach_p300)
    js_coach_sprint = (
        "var _p5=" + js_sprint_p5 + ",_p10=" + js_sprint_p10 + ",_p20=" + js_sprint_p20 + ",_p300=" + js_sprint_p300 + ";"
        "barChart('coachSprint'," + j(coach_dates) + ","
        "[{label:'30s base',data:" + j(coach_p20) + ",backgroundColor:'#27AE60CC',stack:'s'},"
        "{label:'15s mid',data:" + j(coach_p10_mid) + ",backgroundColor:'#2E75B6CC',stack:'s'},"
        "{label:'5s burst',data:" + j(coach_p5_top) + ",backgroundColor:'#E67E22CC',stack:'s'},"
        "{label:'Best 5-min Power (W)',data:" + j(coach_p300) + ",type:'line',borderColor:PURPLE,"
        "backgroundColor:PURPLE+'20',borderWidth:2,pointRadius:2,fill:false,spanGaps:true,order:0}],"
        "{scales:{y:{stacked:true,beginAtZero:true},x:{stacked:true,ticks:{maxRotation:45}}},"
        "zoomable:true,"
        "plugins:{tooltip:{mode:'index',intersect:false,"
        "itemSort:function(a,b){var order={2:0,1:1,0:2,3:3};return order[a.datasetIndex]-order[b.datasetIndex];},"
        "callbacks:{label:function(ctx){"
        "var i=ctx.dataIndex;"
        "if(ctx.datasetIndex===3)return '5-min: '+(_p300[i]||'-')+'W';"
        "if(ctx.datasetIndex===2)return '5s: '+(_p5[i]||'-')+'W';"
        "if(ctx.datasetIndex===1)return '15s: '+(_p10[i]||'-')+'W';"
        "return '30s: '+(_p20[i]||'-')+'W';}}}}});"
    )

    return (
        "<!DOCTYPE html><html lang='en'><head>"
        + "<meta charset='UTF-8'>"
        + "<meta name='viewport' content='width=device-width,initial-scale=1.0'>"
        + "<title>" + name + "'s Cycling Dashboard " + str(YEAR) + "</title>"
        + "<script src='https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js'></script>"
        + "<script src='https://cdn.jsdelivr.net/npm/hammerjs@2.0.8/hammer.min.js'></script>"
        + "<script src='https://cdn.jsdelivr.net/npm/chartjs-plugin-zoom@2.2.0/dist/chartjs-plugin-zoom.min.js'></script>"
        + "<style>"
        + ":root{--blue:#1F4E79;--blue2:#2E75B6;--blue3:#D6E4F0;--green:#27AE60;--orange:#E67E22;--red:#E74C3C;--purple:#9B59B6;--grey:#F5F7FA;--text:#2C3E50;--card:#FFFFFF;}"
        + "*{box-sizing:border-box;margin:0;padding:0;}"
        + "body{font-family:'Segoe UI',Arial,sans-serif;background:var(--grey);color:var(--text);padding:20px;}"
        + "h1{color:var(--blue);font-size:1.6rem;margin-bottom:4px;}"
        + ".subtitle{color:#666;font-size:0.9rem;margin-bottom:20px;}"
        + ".stats-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin-bottom:20px;}"
        + ".stat-card{background:var(--card);border-radius:10px;padding:14px 16px;box-shadow:0 2px 6px rgba(0,0,0,0.07);border-left:4px solid var(--blue2);}"
        + ".synopsis-card{background:var(--card);border-radius:10px;padding:16px 20px;box-shadow:0 2px 6px rgba(0,0,0,0.07);border-left:4px solid var(--blue2);margin-bottom:20px;}"
        + ".synopsis-card .synopsis-label{font-size:0.7rem;font-weight:700;text-transform:uppercase;letter-spacing:.05em;color:var(--blue2);margin-bottom:6px;}"
        + ".synopsis-card .synopsis-date{font-size:0.75rem;color:#888;margin-bottom:10px;}"
        + ".synopsis-card .synopsis-text{font-size:0.9rem;line-height:1.6;color:#333;white-space:pre-wrap;}"
        + ".stat-card .label{font-size:0.7rem;color:#888;text-transform:uppercase;letter-spacing:.05em;}"
        + ".stat-card .value{font-size:1.5rem;font-weight:700;color:var(--blue);margin:4px 0 2px;}"
        + ".stat-card .sub{font-size:0.75rem;color:#666;}"
        + ".stat-card.green{border-left-color:var(--green);}"
        + ".stat-card.orange{border-left-color:var(--orange);}"
        + ".stat-card.purple{border-left-color:var(--purple);}"
        + ".progress-wrap{background:var(--card);border-radius:10px;padding:14px 18px;box-shadow:0 2px 6px rgba(0,0,0,0.07);margin-bottom:20px;}"
        + ".progress-label{display:flex;justify-content:space-between;font-size:0.82rem;color:#666;margin-bottom:6px;}"
        + ".progress-bar-bg{background:#E0E0E0;border-radius:8px;height:16px;overflow:hidden;}"
        + ".progress-bar-fill{height:100%;border-radius:8px;background:linear-gradient(90deg,var(--blue2),var(--green));}"
        + ".charts-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(min(100%,480px),1fr));gap:16px;}"
        + ".chart-card{background:var(--card);border-radius:10px;padding:16px 18px;box-shadow:0 2px 6px rgba(0,0,0,0.07);}"
        + ".chart-card h3{font-size:0.88rem;color:var(--blue);margin-bottom:12px;font-weight:600;}"
        + ".chart-card canvas{max-height:260px;}"
        + ".section-header{margin:24px 0 10px;padding-bottom:6px;border-bottom:2px solid var(--blue3);}"
        + ".section-header h2{color:var(--blue);font-size:1rem;font-weight:700;}"
        + ".section-header p{font-size:0.75rem;color:#888;margin-top:3px;}"
        + ".range-bar{display:flex;gap:6px;align-items:center;margin:4px 0 14px;flex-wrap:wrap;}"
        + ".range-bar span{font-size:11px;color:#888;margin-right:2px;}"
        + ".range-btn{padding:5px 12px;border:1px solid #ddd;border-radius:14px;font-size:11px;background:#fff;color:#555;cursor:pointer;}"
        + ".range-btn.active{background:var(--blue2);color:#fff;border-color:var(--blue2);}"
        + ".footer{text-align:center;color:#aaa;font-size:0.75rem;margin-top:20px;padding-top:12px;border-top:1px solid #eee;}"
        + "@media(max-width:600px){.charts-grid{grid-template-columns:1fr;}.stats-grid{grid-template-columns:repeat(2,1fr);}}"
        + "</style></head><body>"

        + "<h1>&#x1F6B4; " + name + "'s Cycling Dashboard " + str(YEAR) + "</h1>"
        + "<p class='subtitle'>Updated " + date.today().strftime('%B %d, %Y') + " &nbsp;&middot;&nbsp; " + str(len(rides)) + " rides" + goal_subtitle + "</p>"

        + synopsis_card

        + "<div class='stats-grid'>"
        + "<div class='stat-card green'><div class='label'>Year Total</div>"
        + "<div class='value'>" + str(round(total_mi,1)) + "</div>"
        + "<div class='sub'>miles" + (" &nbsp;(" + str(pct_complete) + "% of goal)" if goal else "") + "</div></div>"

        + (goal_cards if goal_cards else "")

        + "<div class='stat-card'><div class='label'>Hours in Saddle</div>"
        + "<div class='value'>" + str(round(total_hrs,1)) + "</div>"
        + "<div class='sub'>hours total</div></div>"

        + "<div class='stat-card'><div class='label'>Total Rides</div>"
        + "<div class='value'>" + str(len(rides)) + "</div>"
        + "<div class='sub'>" + str(virt_count) + " virtual &nbsp;&middot;&nbsp; " + str(out_count) + " outdoor</div></div>"

        + "<div class='stat-card purple'><div class='label'>Elevation</div>"
        + "<div class='value'>" + str(int(total_elev)) + "</div>"
        + "<div class='sub'>feet climbed total</div></div>"
        + "</div>"

        + (goal_progress if goal_progress else "")

        + "<div class='charts-grid'>"
        + "<div class='chart-card'><h3>&#x1F4C5; Weekly Mileage vs " + str(WEEKLY_TARGET) + "-Mile Target</h3><canvas id='weeklyBar'></canvas></div>"
        + "<div class='chart-card'><h3>&#x1F4C8; Cumulative Miles vs Annual Target</h3><canvas id='cumulativeLine'></canvas></div>"
        + "<div class='chart-card'><h3>&#x1F4C6; Monthly Miles</h3><canvas id='monthlyBar'></canvas></div>"
        + "<div class='chart-card'><h3>&#x23F1; Hours in the Saddle by Month</h3><canvas id='monthlyHours'></canvas></div>"
        + "<div class='chart-card'><h3>&#x1F3F7; Ride Type &#x2014; Miles</h3><canvas id='rtypeMiles'></canvas></div>"
        + "<div class='chart-card'><h3>&#x1F3F7; Ride Type &#x2014; Hours</h3><canvas id='rtypeHours'></canvas></div>"
        + "<div class='chart-card'><h3>&#x1F7E3; Virtual vs Outdoor</h3>" + virt_html + "</div>"
        + "</div>"

        + "<div class='section-header'>"
        + "<h2>&#x1F4C8; Per-Ride Trends</h2>"
        + "<p>Pinch or scroll to zoom, tap &#8635; to reset</p>"
        + "</div>"
        + range_bar_html
        + "<div class='charts-grid'>"
        + "<div class='chart-card'><h3>&#x26F0; Elevation Gain per Ride (ft) <a href='#' onclick=\"resetZoom('elevBar');return false;\" style='float:right;font-size:10px;color:#888;font-weight:400;text-decoration:none;'>&#8635; reset zoom</a></h3><canvas id='elevBar'></canvas></div>"
        + "<div class='chart-card'><h3>&#x26A1; Average Power per Ride (W) <a href='#' onclick=\"resetZoom('powerLine');return false;\" style='float:right;font-size:10px;color:#888;font-weight:400;text-decoration:none;'>&#8635; reset zoom</a></h3><canvas id='powerLine'></canvas></div>"
        + "<div class='chart-card'><h3>&#x2764; Average Heart Rate per Ride (bpm) <a href='#' onclick=\"resetZoom('hrLine');return false;\" style='float:right;font-size:10px;color:#888;font-weight:400;text-decoration:none;'>&#8635; reset zoom</a></h3><canvas id='hrLine'></canvas></div>"
        + "</div>"

        + "<div class='section-header'>"
        + "<h2>&#x1F3C6; Coaching Analytics &#x2014; Power &middot; Heart Rate &middot; Cadence &middot; Sprint Power</h2>"
        + "<p>Solid line = average &nbsp;&middot;&nbsp; Dashed line = max/normalized &nbsp;&middot;&nbsp; All rides &#x2265; 5 miles &nbsp;&middot;&nbsp; Pinch or scroll to zoom, tap &#8635; to reset</p>"
        + "</div>"
        + range_bar_html
        + "<div class='charts-grid'>"
        + "<div class='chart-card'><h3>&#x26A1; Avg Power vs Normalized Power (W) <a href='#' onclick=\"resetZoom('coachPower');return false;\" style='float:right;font-size:10px;color:#888;font-weight:400;text-decoration:none;'>&#8635; reset zoom</a></h3><canvas id='coachPower'></canvas></div>"
        + "<div class='chart-card'><h3>&#x2764;&#xFE0F; Avg HR vs Max HR (bpm) <a href='#' onclick=\"resetZoom('coachHR');return false;\" style='float:right;font-size:10px;color:#888;font-weight:400;text-decoration:none;'>&#8635; reset zoom</a></h3><canvas id='coachHR'></canvas></div>"
        + "<div class='chart-card'><h3>&#x1F504; Avg Cadence vs Max Cadence (rpm) <a href='#' onclick=\"resetZoom('coachCad');return false;\" style='float:right;font-size:10px;color:#888;font-weight:400;text-decoration:none;'>&#8635; reset zoom</a></h3><canvas id='coachCad'></canvas></div>"
        + "<div class='chart-card'><h3>&#x1F3CE;&#xFE0F; Sprint &amp; Aerobic Power &#x2014; 5s / 15s / 30s / 5-min Best (W) <a href='#' onclick=\"resetZoom('coachSprint');return false;\" style='float:right;font-size:10px;color:#888;font-weight:400;text-decoration:none;'>&#8635; reset zoom</a></h3><canvas id='coachSprint'></canvas></div>"
        + "</div>"

        + "<p class='footer'>Generated by Cycling Coach &nbsp;&middot;&nbsp; " + date.today().strftime('%Y-%m-%d') + "</p>"

        + "<script>"
        + "const BLUE='#2E75B6',GREEN='#27AE60',ORANGE='#E67E22',RED='#E74C3C',PURPLE='#9B59B6',GREY='#95A5A6';"
        + "const TYPE_COLORS=[GREY,ORANGE,BLUE,GREEN,'#F39C12',RED,GREEN];"
        + "Chart.defaults.font.family=\"'Segoe UI',Arial,sans-serif\";"
        + "Chart.defaults.font.size=11;Chart.defaults.color='#555';"
        + "const ZOOM_CONFIG={pan:{enabled:true,mode:'x'},"
        + "zoom:{wheel:{enabled:true},pinch:{enabled:true},mode:'x'},"
        + "limits:{x:{minRange:3}}};"
        + "const CHARTS={};"
        + "function resetZoom(id){if(CHARTS[id])CHARTS[id].resetZoom();}"
        + "const RIDE_DATE_ISO={"
        + "elevBar:" + j(ride_dates_iso) + ","
        + "powerLine:" + j(ride_dates_iso) + ","
        + "hrLine:" + j(ride_dates_iso) + ","
        + "coachPower:" + j(coach_dates_iso) + ","
        + "coachHR:" + j(coach_dates_iso) + ","
        + "coachCad:" + j(coach_dates_iso) + ","
        + "coachSprint:" + j(coach_dates_iso)
        + "};"
        + "function setRange(range){"
        + "var ids=['elevBar','powerLine','hrLine','coachPower','coachHR','coachCad','coachSprint'];"
        + "ids.forEach(function(id){"
        + "var chart=CHARTS[id];if(!chart)return;"
        + "if(range==='all'){chart.resetZoom();return;}"
        + "var dates=RIDE_DATE_ISO[id];if(!dates||dates.length<3)return;"
        + "var cutoff;"
        + "if(range==='ytd'){cutoff=new Date('" + str(YEAR) + "-01-01T00:00:00');}"
        + "else{var lastDate=new Date(dates[dates.length-1]+'T12:00:00');"
        + "cutoff=new Date(lastDate);cutoff.setDate(cutoff.getDate()-parseInt(range));}"
        + "var startIdx=dates.findIndex(function(d){return new Date(d+'T12:00:00')>=cutoff;});"
        + "if(startIdx===-1)startIdx=0;"
        + "if(startIdx>dates.length-3)startIdx=Math.max(0,dates.length-3);"
        + "chart.zoomScale('x',{min:startIdx,max:dates.length-1},'default');"
        + "});"
        + "document.querySelectorAll('.range-btn').forEach(function(b){"
        + "b.classList.toggle('active',b.dataset.range===range);});}"
        + "function barChart(id,labels,datasets,opts){"
        + "opts=opts||{};"
        + "const zoomable=opts.zoomable;"
        + "const plugins=Object.assign({legend:{display:datasets.length>1}},opts.plugins||{});"
        + "if(zoomable)plugins.zoom=ZOOM_CONFIG;"
        + "const finalOpts=Object.assign({responsive:true,"
        + "scales:{y:{beginAtZero:true},x:{ticks:{maxRotation:45}}}},opts,{plugins:plugins});"
        + "CHARTS[id]=new Chart(document.getElementById(id),{type:'bar',data:{labels:labels,datasets:datasets},options:finalOpts});"
        + "return CHARTS[id];}"
        + "function lineChart(id,labels,datasets,opts){"
        + "opts=opts||{};"
        + "const zoomable=opts.zoomable;"
        + "const plugins=Object.assign({legend:{display:datasets.length>1}},opts.plugins||{});"
        + "if(zoomable)plugins.zoom=ZOOM_CONFIG;"
        + "const finalOpts=Object.assign({responsive:true,"
        + "scales:{y:{beginAtZero:false},x:{ticks:{maxRotation:45}}},"
        + "elements:{point:{radius:2},line:{tension:0.3}}},opts,{plugins:plugins});"
        + "CHARTS[id]=new Chart(document.getElementById(id),{type:'line',data:{labels:labels,datasets:datasets},options:finalOpts});"
        + "return CHARTS[id];}"
        + js_weekly
        + js_cumul
        + js_mo_mi
        + js_mo_hr
        + js_rt_mi
        + js_rt_hr
        + js_elev
        + js_pwr
        + js_hr
        + js_coach_pwr
        + js_coach_hr
        + js_coach_cad
        + js_coach_sprint
        + "</script></body></html>"
    )

# ── Coaching AI Tools ─────────────────────────────────────────────────────
# A small number of general, reusable tools rather than one per possible
# question. The AI decides which to call, with what parameters, and how
# many times — these stay fixed, tested, narrow; no open-ended queries
# get written on the fly by the model itself.

COACHING_TOOLS = [
    {
        "name": "get_ride_metric_range",
        "description": "Get the average/min/max for a specific metric (power, heart_rate, cadence, altitude, speed) over a specific mile range within one ride. Use for questions about a specific part of a ride, like 'what was my power between mile 8 and 10' or 'what was the gradient on that climb'. Only works for rides that have detailed stream data (uploaded/synced since the raw-data storage was added — older rides may not have this).",
        "input_schema": {
            "type": "object",
            "properties": {
                "ride_id": {"type": "integer", "description": "The ride's ID"},
                "metric": {"type": "string", "enum": ["power", "heart_rate", "cadence", "altitude", "speed"]},
                "start_mile": {"type": "number", "description": "Start of the range, in miles into the ride"},
                "end_mile": {"type": "number", "description": "End of the range, in miles into the ride"}
            },
            "required": ["ride_id", "metric", "start_mile", "end_mile"]
        }
    },
    {
        "name": "find_push_segments",
        "description": "Detect discrete hard-effort surges within one ride from its power data — sustained periods of elevated power, each with location (mile marker), duration, and average/max power. Use when asked to identify hard efforts or pushes within a specific ride, similar to identifying distinct climbs or surges.",
        "input_schema": {
            "type": "object",
            "properties": {"ride_id": {"type": "integer"}},
            "required": ["ride_id"]
        }
    },
    {
        "name": "get_zone_breakdown",
        "description": "Get time-in-zone breakdown for one ride — either power zones (based on FTP, needs FTP set in profile) or heart rate zones (estimated from age via 220-age, needs age set in profile). Use when asked about zone distribution or how much time was spent at different intensities.",
        "input_schema": {
            "type": "object",
            "properties": {
                "ride_id": {"type": "integer"},
                "zone_type": {"type": "string", "enum": ["power", "heart_rate"]}
            },
            "required": ["ride_id", "zone_type"]
        }
    },
    {
        "name": "search_rides_and_history",
        "description": "Search across all rides and coaching history using any combination of elevation range, distance range, date range, and a keyword searched against the coaching memory's dated log. Always returns a count plus a small preview (up to 10 examples) — never the full list. If the count is large, that's a cue to narrow further (ask the rider for more specifics, or search again with tighter constraints) rather than dump everything. Use to find specific rides matching criteria, or to search past coaching discussions by topic.",
        "input_schema": {
            "type": "object",
            "properties": {
                "elevation_min_ft": {"type": "number"},
                "elevation_max_ft": {"type": "number"},
                "distance_min_mi": {"type": "number"},
                "distance_max_mi": {"type": "number"},
                "date_start": {"type": "string", "description": "YYYY-MM-DD"},
                "date_end": {"type": "string", "description": "YYYY-MM-DD"},
                "keyword": {"type": "string", "description": "Search term matched against dated coaching log entries"}
            },
            "required": []
        }
    },
    {
        "name": "get_dated_log_entry",
        "description": "Get the full coaching memory synopsis for one specific date. Use once a specific ride/date has been identified (e.g. via search_rides_and_history) and the rider wants the actual discussion/synopsis for that day — this is a direct lookup, not a search.",
        "input_schema": {
            "type": "object",
            "properties": {"date": {"type": "string", "description": "YYYY-MM-DD"}},
            "required": ["date"]
        }
    },
    {
        "name": "get_power_curve",
        "description": "Get one ride's power curve — best-average-power at key durations (5s, 15s, 30s, 1min, 5min, 10min, 20min, 30min, 60min, whichever the ride is long enough to have), plus a rough FTP estimate derived from it when computable. Use for questions that go beyond the basic 5s/15s/30s/5-min bests already in context — anything about sustained power at a specific duration, whether a ride looks like a new FTP, or overall power profile shape. Only works for rides with detailed stream data (uploaded/synced since raw-data storage was added).",
        "input_schema": {
            "type": "object",
            "properties": {"ride_id": {"type": "integer"}},
            "required": ["ride_id"]
        }
    }
]

def tool_get_ride_metric_range(user_id, ride_id, metric, start_mile, end_mile):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id FROM rides WHERE id=%s AND user_id=%s", (ride_id, user_id))
    if not cur.fetchone():
        cur.close(); conn.close()
        return {"error": "Ride not found or doesn't belong to this rider"}
    cur.execute("SELECT streams FROM ride_streams WHERE ride_id=%s", (ride_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row or not row.get('streams'):
        return {"error": "No detailed stream data available for this ride (only rides uploaded/synced since raw-data storage was added have this)"}
    streams = row['streams']
    distances = streams.get('distance')
    if not distances:
        return {"error": "No distance data available to locate the mile range for this ride"}
    distances_mi = [d/1609.34 if d is not None else None for d in distances]
    metric_vals = streams.get(metric)
    if not metric_vals:
        return {"error": f"No {metric} data available for this ride"}
    matched = [
        metric_vals[i] for i in range(min(len(distances_mi), len(metric_vals)))
        if distances_mi[i] is not None and start_mile <= distances_mi[i] <= end_mile and metric_vals[i] is not None
    ]
    if not matched:
        return {"error": f"No {metric} data found between mile {start_mile} and {end_mile} on this ride"}
    return {
        "metric": metric, "start_mile": start_mile, "end_mile": end_mile,
        "avg": round(sum(matched)/len(matched), 1),
        "min": round(min(matched), 1), "max": round(max(matched), 1),
        "sample_count": len(matched)
    }

def tool_find_push_segments(user_id, ride_id):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM rides WHERE id=%s AND user_id=%s", (ride_id, user_id))
    ride = cur.fetchone()
    if not ride:
        cur.close(); conn.close()
        return {"error": "Ride not found or doesn't belong to this rider"}
    cur.execute("SELECT streams FROM ride_streams WHERE ride_id=%s", (ride_id,))
    srow = cur.fetchone()
    cur.execute("SELECT ftp FROM profiles WHERE user_id=%s", (user_id,))
    prow = cur.fetchone()
    cur.close(); conn.close()
    if not srow or not srow.get('streams'):
        return {"error": "No detailed stream data available for this ride"}
    streams = srow['streams']
    powers = streams.get('power'); distances = streams.get('distance')
    if not powers or not distances:
        return {"error": "No power/distance data available for this ride"}

    ftp = prow['ftp'] if prow and prow.get('ftp') else None
    avg_power = ride.get('avg_power') or 150
    threshold = ftp if ftp else round(avg_power * 1.3)
    min_duration_samples = 30  # ~30s, assuming ~1 sample/sec recording

    segments = []
    i = 0; n = min(len(powers), len(distances))
    while i < n:
        if powers[i] is not None and powers[i] >= threshold:
            start_i = i
            while i < n and powers[i] is not None and powers[i] >= threshold:
                i += 1
            end_i = i
            duration = end_i - start_i
            if duration >= min_duration_samples:
                seg_powers = [p for p in powers[start_i:end_i] if p is not None]
                seg_dist = distances[start_i] if distances[start_i] is not None else None
                segments.append({
                    "start_mile": round(seg_dist/1609.34, 1) if seg_dist else None,
                    "duration_seconds": duration,
                    "avg_power": round(sum(seg_powers)/len(seg_powers)) if seg_powers else None,
                    "max_power": round(max(seg_powers)) if seg_powers else None
                })
        else:
            i += 1
    return {"ride_id": ride_id, "threshold_watts": threshold, "segments_found": len(segments), "segments": segments[:10]}

def tool_get_zone_breakdown(user_id, ride_id, zone_type):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id FROM rides WHERE id=%s AND user_id=%s", (ride_id, user_id))
    if not cur.fetchone():
        cur.close(); conn.close()
        return {"error": "Ride not found or doesn't belong to this rider"}
    cur.execute("SELECT ftp, age FROM profiles WHERE user_id=%s", (user_id,))
    prow = cur.fetchone()
    cur.execute("SELECT streams FROM ride_streams WHERE ride_id=%s", (ride_id,))
    srow = cur.fetchone()
    cur.close(); conn.close()
    if not srow or not srow.get('streams'):
        return {"error": "No detailed stream data available for this ride"}
    streams = srow['streams']

    if zone_type == 'power':
        ftp = prow['ftp'] if prow and prow.get('ftp') else None
        if not ftp:
            return {"error": "No FTP set in profile - can't compute power zones"}
        vals = [p for p in (streams.get('power') or []) if p is not None]
        if not vals:
            return {"error": "No power data available for this ride"}
        zones = {"Z1 (<55%)":0,"Z2 (56-75%)":0,"Z3 (76-90%)":0,"Z4 (91-105%)":0,"Z5 (106-120%)":0,"Z6 (121-150%)":0,"Z7 (>150%)":0}
        for p in vals:
            pct = p/ftp*100
            if pct < 55: zones["Z1 (<55%)"] += 1
            elif pct < 76: zones["Z2 (56-75%)"] += 1
            elif pct < 91: zones["Z3 (76-90%)"] += 1
            elif pct < 106: zones["Z4 (91-105%)"] += 1
            elif pct < 121: zones["Z5 (106-120%)"] += 1
            elif pct < 151: zones["Z6 (121-150%)"] += 1
            else: zones["Z7 (>150%)"] += 1
        total = len(vals)
        return {"zone_type": "power", "ftp_used": ftp, "total_samples": total,
                "zones_pct": {k: round(v/total*100, 1) for k, v in zones.items()}}
    else:
        age = prow['age'] if prow and prow.get('age') else None
        if not age:
            return {"error": "No age set in profile - can't estimate HR zones"}
        est_max_hr = 220 - age
        vals = [h for h in (streams.get('heart_rate') or []) if h is not None]
        if not vals:
            return {"error": "No heart rate data available for this ride"}
        zones = {"Z1 (<60%)":0,"Z2 (60-70%)":0,"Z3 (70-80%)":0,"Z4 (80-90%)":0,"Z5 (>90%)":0}
        for h in vals:
            pct = h/est_max_hr*100
            if pct < 60: zones["Z1 (<60%)"] += 1
            elif pct < 70: zones["Z2 (60-70%)"] += 1
            elif pct < 80: zones["Z3 (70-80%)"] += 1
            elif pct < 90: zones["Z4 (80-90%)"] += 1
            else: zones["Z5 (>90%)"] += 1
        total = len(vals)
        return {"zone_type": "heart_rate", "estimated_max_hr": est_max_hr,
                "note": "max HR estimated from age via 220-age, not directly measured",
                "total_samples": total, "zones_pct": {k: round(v/total*100, 1) for k, v in zones.items()}}

def tool_search_rides_and_history(user_id, elevation_min_ft=None, elevation_max_ft=None,
                                    distance_min_mi=None, distance_max_mi=None,
                                    date_start=None, date_end=None, keyword=None):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    numeric_given = any(v is not None for v in [elevation_min_ft, elevation_max_ft, distance_min_mi, distance_max_mi, date_start, date_end])

    numeric_rows = []; numeric_dates = None
    if numeric_given:
        conditions = ["user_id=%s"]; params = [user_id]
        if elevation_min_ft is not None: conditions.append("elev_gain_ft >= %s"); params.append(elevation_min_ft)
        if elevation_max_ft is not None: conditions.append("elev_gain_ft <= %s"); params.append(elevation_max_ft)
        if distance_min_mi is not None: conditions.append("dist_mi >= %s"); params.append(distance_min_mi)
        if distance_max_mi is not None: conditions.append("dist_mi <= %s"); params.append(distance_max_mi)
        if date_start is not None: conditions.append("ride_date >= %s"); params.append(date_start)
        if date_end is not None: conditions.append("ride_date <= %s"); params.append(date_end)
        cur.execute(f"SELECT ride_date, name, dist_mi, elev_gain_ft FROM rides WHERE {' AND '.join(conditions)} ORDER BY ride_date DESC", params)
        numeric_rows = [dict(r) for r in cur.fetchall()]
        numeric_dates = {str(r['ride_date']) for r in numeric_rows}

    keyword_rows = []; keyword_dates = None
    if keyword:
        cur.execute("SELECT entry_date, summary FROM coaching_memory_log WHERE user_id=%s AND summary ILIKE %s ORDER BY entry_date DESC",
                    (user_id, f"%{keyword}%"))
        keyword_rows = [dict(r) for r in cur.fetchall()]
        keyword_dates = {str(r['entry_date']) for r in keyword_rows}
    cur.close(); conn.close()

    if numeric_dates is not None and keyword_dates is not None:
        final_dates = numeric_dates & keyword_dates
    elif numeric_dates is not None:
        final_dates = numeric_dates
    elif keyword_dates is not None:
        final_dates = keyword_dates
    else:
        return {"error": "No search criteria provided - specify at least one filter (elevation, distance, date range, or keyword)"}

    preview = []
    for d in sorted(final_dates, reverse=True)[:10]:
        entry = {"date": d}
        ride_match = next((r for r in numeric_rows if str(r['ride_date']) == d), None)
        if ride_match:
            entry["name"] = ride_match.get('name')
            entry["dist_mi"] = ride_match.get('dist_mi')
            entry["elev_gain_ft"] = ride_match.get('elev_gain_ft')
        kw_match = next((r for r in keyword_rows if str(r['entry_date']) == d), None)
        if kw_match:
            entry["summary_snippet"] = (kw_match.get('summary') or '')[:150]
        preview.append(entry)
    return {"count": len(final_dates), "preview": preview}

def tool_get_dated_log_entry(user_id, date):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT entry_date, summary FROM coaching_memory_log WHERE user_id=%s AND entry_date=%s", (user_id, date))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        return {"error": f"No coaching log entry found for {date}"}
    return {"date": str(row['entry_date']), "summary": row['summary']}

def tool_get_power_curve(user_id, ride_id):
    """Relies on compute_power_curve()/estimate_ftp_from_curve()/
    fmt_duration_label(), all defined later in this file near
    build_ride_detail_html — fine at call time (this only runs during
    request handling, well after the whole module has loaded), just
    not colocated with the rest of the tool functions above."""
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id FROM rides WHERE id=%s AND user_id=%s", (ride_id, user_id))
    if not cur.fetchone():
        cur.close(); conn.close()
        return {"error": "Ride not found or doesn't belong to this rider"}
    cur.execute("SELECT streams FROM ride_streams WHERE ride_id=%s", (ride_id,))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row or not row.get('streams') or not row['streams'].get('power'):
        return {"error": "No detailed power stream data available for this ride"}
    curve = compute_power_curve(row['streams']['power'])
    if not curve:
        return {"error": "Not enough power data to compute a curve for this ride"}
    key_durations = [5, 15, 30, 60, 300, 600, 1200, 1800, 3600]
    by_dur = {c['duration_s']: c['watts'] for c in curve}
    checkpoints = {fmt_duration_label(d): by_dur[d] for d in key_durations if d in by_dur}
    result = {"ride_id": ride_id, "power_curve_watts": checkpoints}
    ftp_est = estimate_ftp_from_curve(curve)
    if ftp_est:
        result["estimated_ftp"] = {"watts": ftp_est['watts'], "basis": ftp_est['basis']}
    return result

def execute_coaching_tool(name, tool_input, user_id):
    try:
        if name == "get_ride_metric_range":
            return tool_get_ride_metric_range(user_id, **tool_input)
        elif name == "find_push_segments":
            return tool_find_push_segments(user_id, **tool_input)
        elif name == "get_zone_breakdown":
            return tool_get_zone_breakdown(user_id, **tool_input)
        elif name == "search_rides_and_history":
            return tool_search_rides_and_history(user_id, **tool_input)
        elif name == "get_dated_log_entry":
            return tool_get_dated_log_entry(user_id, **tool_input)
        elif name == "get_power_curve":
            return tool_get_power_curve(user_id, **tool_input)
        else:
            return {"error": f"Unknown tool: {name}"}
    except Exception as e:
        return {"error": f"Tool execution failed: {str(e)}"}

async def run_claude_with_tools(system_prompt, messages, user_id, max_tokens=900, max_iterations=5):
    """Runs a Claude conversation with tool-use support, looping until a
    final text-only response (no more tool calls) or max_iterations is hit.
    Returns the final reply text. The tools stay fixed and narrow — the
    model only ever decides which to call and with what parameters."""
    for _ in range(max_iterations):
        body = {"model": "claude-sonnet-4-6", "max_tokens": max_tokens,
                "messages": messages, "tools": COACHING_TOOLS}
        if system_prompt:
            body["system"] = system_prompt
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01"},
                json=body,
                timeout=45
            )
            data = resp.json()
        content_blocks = data.get('content', [])
        tool_use_blocks = [b for b in content_blocks if b.get('type') == 'tool_use']
        if not tool_use_blocks:
            return ''.join(b.get('text', '') for b in content_blocks if b.get('type') == 'text')
        messages.append({"role": "assistant", "content": content_blocks})
        tool_results = []
        for tb in tool_use_blocks:
            result = execute_coaching_tool(tb['name'], tb.get('input', {}), user_id)
            tool_results.append({"type": "tool_result", "tool_use_id": tb['id'], "content": json.dumps(result)})
        messages.append({"role": "user", "content": tool_results})
    return "I wasn't able to finish gathering everything needed to answer that fully — try asking a bit more specifically?"

async def get_coaching_summary(user, metrics, ride_id=None):
    if not ANTHROPIC_KEY:
        return "AI coaching unavailable."
    try:
        conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM rides WHERE user_id=%s ORDER BY ride_date DESC LIMIT 10", (user['id'],))
        recent = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT note FROM coaching_notes WHERE user_id=%s ORDER BY created_at DESC LIMIT 5", (user['id'],))
        notes = [r['note'] for r in cur.fetchall()]
        cur.execute("SELECT * FROM profiles WHERE user_id=%s", (user['id'],))
        profile = cur.fetchone()
        cur.execute("SELECT entry_date, summary FROM coaching_memory_log WHERE user_id=%s ORDER BY entry_date DESC LIMIT 30", (user['id'],))
        memory_log = [dict(r) for r in cur.fetchall()]
        cur.execute("SELECT theme, content FROM coaching_memory_themes WHERE user_id=%s", (user['id'],))
        memory_themes = {r['theme']: r['content'] for r in cur.fetchall()}
        cur.close(); conn.close()

        profile_ctx = ""
        if profile:
            profile_ctx = "RIDER PROFILE:\n"
            if profile.get('age'):          profile_ctx += "- Age: " + str(profile['age']) + "\n"
            if profile.get('weight_lbs'):   profile_ctx += "- Weight: " + str(profile['weight_lbs']) + " lbs\n"
            if profile.get('location'):     profile_ctx += "- Location: " + str(profile['location']) + "\n"
            if profile.get('fitness_level'):profile_ctx += "- Fitness: " + str(profile['fitness_level']) + "\n"
            if profile.get('ftp'):          profile_ctx += "- FTP: " + str(profile['ftp']) + "W\n"
            if profile.get('annual_goal_mi'):profile_ctx += "- Annual goal: " + str(profile['annual_goal_mi']) + " mi\n"
            if profile.get('other_goals'):  profile_ctx += "- Goals: " + str(profile['other_goals']) + "\n"
            if profile.get('health_notes'): profile_ctx += "- Health: " + str(profile['health_notes']) + "\n"
            if profile.get('injuries'):     profile_ctx += "- Injuries: " + str(profile['injuries']) + "\n"
            if profile.get('heat_tolerance'):profile_ctx += "- Heat tolerance: " + str(profile['heat_tolerance']) + "\n"

        recent_ctx = ""
        if recent:
            recent_ctx = "RECENT RIDES (last 10):\n"
            for r in recent[:5]:
                recent_ctx += "- " + str(r.get('ride_date',''))[:10] + ": " + str(r.get('dist_mi','')) + "mi, HR " + str(r.get('avg_hr','')) + ", pwr " + str(r.get('avg_power','')) + "W\n"

        memory_log_ctx = ""
        if memory_log:
            memory_log_ctx = "\nCOACHING MEMORY — DATED LOG (most recent first):\n"
            for m in memory_log:
                memory_log_ctx += "- " + str(m['entry_date']) + ": " + str(m['summary']) + "\n"

        memory_themes_ctx = ""
        if memory_themes:
            memory_themes_ctx = "\nCOACHING MEMORY — STANDING PATTERNS:\n"
            for key, label in MEMORY_THEMES.items():
                if memory_themes.get(key):
                    memory_themes_ctx += "- " + label + ": " + str(memory_themes[key]) + "\n"

        prompt = (
            profile_ctx + "\n"
            + ("PERSONAL NOTES: " + "; ".join(notes) + "\n\n" if notes else "")
            + recent_ctx + memory_log_ctx + memory_themes_ctx + "\n"
            + "LATEST RIDE:\n"
            + ("- Ride ID: " + str(ride_id) + " (use this for get_ride_metric_range, find_push_segments, or get_zone_breakdown if a deeper look at this specific ride would help)\n" if ride_id else "")
            + "- Date: " + str(metrics.get('ride_date','')) + "\n"
            + "- Distance: " + str(metrics.get('dist_mi','')) + " mi\n"
            + "- Avg power: " + str(metrics.get('avg_power','')) + "W, NP: " + str(metrics.get('norm_power','')) + "W\n"
            + "- Sprint/aerobic bests — 5s: " + str(metrics.get('p5','')) + "W, 15s: " + str(metrics.get('p15','')) + "W, "
            + "30s: " + str(metrics.get('p30','')) + "W, 5-min: " + str(metrics.get('p300','')) + "W\n"
            + "- Avg HR: " + str(metrics.get('avg_hr','')) + " bpm, Max HR: " + str(metrics.get('max_hr','')) + " bpm\n"
            + "- Cadence: " + str(metrics.get('avg_cadence','')) + " rpm\n"
            + "- Elevation: +" + str(metrics.get('elev_gain_ft','?')) + " ft" + (" / -" + str(metrics.get('elev_loss_ft')) + " ft" if metrics.get('elev_loss_ft') is not None else "") + "\n"
            + ("- Left/right power balance: " + str(metrics.get('avg_lr_balance')) + "% right\n" if metrics.get('avg_lr_balance') is not None else "")
            + ("- Training stress score: " + str(metrics.get('training_stress_score')) + ", Intensity factor: " + str(metrics.get('intensity_factor')) + "\n" if metrics.get('training_stress_score') else "")
            + ("- Calories: " + str(metrics.get('calories')) + "\n" if metrics.get('calories') else "")
            + "- Temp: " + str(metrics.get('temp_c','')) + "C\n\n"
            + "Give a real coaching assessment of this ride — not a fixed length, whatever the "
            + "data actually supports. Reference specific numbers (compare 5-min power and NP "
            + "to their FTP if known — that's often the most telling comparison). Reference their "
            + "specific situation — age, recovery status, heat, goals. If they mentioned recent "
            + "illness or injury, factor that in. If pre/post-ride weight, fluid intake, or food "
            + "during the ride isn't in their notes, mention briefly that logging it (in Post-Ride "
            + "Debrief or the Coaching tab) would sharpen future hydration/fueling feedback — don't "
            + "belabor it, one line is enough. End with one specific, actionable thing for next time.\n\n"
            + "After your assessment, on a new line, write exactly MEMORY_UPDATE: with nothing else "
            + "on that line, then on the next line a single JSON object (only JSON, no markdown "
            + "fences) shaped like: {\"dated_entry\": {\"date\":\"YYYY-MM-DD\",\"summary\":\"...\"} "
            + "or null, \"theme_updates\": {\"hydration_fueling\":\"...\" or null, "
            + "\"effort_perception\":\"...\" or null, \"recovery_readiness\":\"...\" or null, "
            + "\"environmental_context\":\"...\" or null, \"life_context\":\"...\" or null, "
            + "\"search_preferences\":\"...\" or null}}. "
            + "Include a dated_entry for this ride if anything is worth remembering later. Only "
            + "fill in a theme_update for a theme this ride actually informs; each one you include "
            + "must be the FULL updated pattern (folding in what's new with what's shown above), "
            + "not just the new piece. Keep every summary/update to 2-3 sentences, distilled."
        )
        messages = [{"role": "user", "content": prompt}]
        full_text = await run_claude_with_tools(None, messages, user['id'], max_tokens=800)

        assessment_text = full_text
        if 'MEMORY_UPDATE:' in full_text:
            parts = full_text.split('MEMORY_UPDATE:', 1)
            assessment_text = parts[0].strip()
            try:
                mem_update = extract_json_object(parts[1].strip())
            except Exception:
                mem_update = {}
            if mem_update:
                conn2 = get_db(); cur2 = conn2.cursor()
                de = mem_update.get('dated_entry')
                if de and de.get('date') and de.get('summary'):
                    cur2.execute("""
                        INSERT INTO coaching_memory_log (user_id, entry_date, summary)
                        VALUES (%s,%s,%s)
                        ON CONFLICT (user_id, entry_date) DO UPDATE SET summary=EXCLUDED.summary
                    """, (user['id'], de['date'], de['summary']))
                tu = mem_update.get('theme_updates') or {}
                for theme_key in MEMORY_THEMES:
                    content = tu.get(theme_key)
                    if content:
                        cur2.execute("""
                            INSERT INTO coaching_memory_themes (user_id, theme, content, updated_at)
                            VALUES (%s,%s,%s,NOW())
                            ON CONFLICT (user_id, theme) DO UPDATE SET content=EXCLUDED.content, updated_at=NOW()
                        """, (user['id'], theme_key, content))
                cur2.close(); conn2.close()

        return assessment_text
    except Exception as e:
        return "Coaching summary unavailable: " + str(e)

@app.on_event("startup")
def startup():
    init_db()

@app.get("/")
def root():
    return {"status": "Cycling Coach API running", "version": APP_VERSION}

@app.get("/health")
def health():
    """Dedicated healthcheck endpoint, separate from / (which is also used
    to verify the deployed version). Testing whether that overlap is
    related to the stale-version-serving issue — see Railway support
    thread and station.railway.com discussion for context."""
    return {"status": "ok"}

@app.post("/register")
def register(email: str = Form(...), name: str = Form(...), password: str = Form(...)):
    conn = get_db(); cur = conn.cursor()
    try:
        token = secrets.token_hex(32)
        cur.execute("INSERT INTO users (email,name,password,token) VALUES (%s,%s,%s,%s) RETURNING id",
                    (email.lower(), name, hash_password(password), token))
        uid = cur.fetchone()[0]; cur.close(); conn.close()
        return {"token": token, "user_id": uid, "name": name}
    except psycopg2.errors.UniqueViolation:
        raise HTTPException(status_code=400, detail="Email already registered")

@app.post("/login")
def login(email: str = Form(...), password: str = Form(...)):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM users WHERE email=%s AND password=%s",
                (email.lower(), hash_password(password)))
    user = cur.fetchone(); cur.close(); conn.close()
    if not user:
        raise HTTPException(status_code=401, detail="Invalid email or password")
    return {"token": user['token'], "name": user['name'], "user_id": user['id']}

@app.post("/upload")
async def upload_fit(file: UploadFile = File(...), notes: str = Form(default=""),
                     user: dict = Depends(get_current_user)):
    data    = await file.read()
    metrics = parse_fit_bytes(data)
    NON_CYCLING_FIT_SPORTS = {"running", "walking", "hiking", "swimming"}
    if metrics.get('sport') in NON_CYCLING_FIT_SPORTS:
        raise HTTPException(status_code=400,
            detail=f"This file is a {metrics['sport']} activity, not a ride — only cycling activities are tracked here.")
    streams = metrics.pop('streams', None)
    conn = get_db(); cur = conn.cursor()
    new_end_time = None
    if metrics.get('start_time') and metrics.get('elapsed_h'):
        try:
            new_end_time = (datetime.fromisoformat(metrics['start_time']) + timedelta(hours=metrics['elapsed_h'])).isoformat()
        except Exception:
            new_end_time = None

    duplicate_of_id = None
    if metrics.get('start_time') and new_end_time:
        cur.execute("""SELECT id FROM rides WHERE user_id=%s
            AND ABS(ride_date - %s::date) <= 1
            AND start_time IS NOT NULL AND elapsed_h IS NOT NULL
            AND start_time <= %s::timestamp
            AND %s::timestamp <= (start_time + (elapsed_h * INTERVAL '1 hour'))
            LIMIT 1""",
            (user['id'], metrics['ride_date'], new_end_time, metrics['start_time']))
        overlap_match = cur.fetchone()
        if overlap_match:
            duplicate_of_id = overlap_match[0]

    if duplicate_of_id is None:
        cur.execute("""SELECT id FROM rides WHERE user_id=%s
            AND ABS(ride_date - %s::date) <= 1
            AND (start_time IS NULL OR elapsed_h IS NULL)
            AND ABS(COALESCE(dist_mi,0)-%s)<0.5 AND ABS(COALESCE(duration_h,0)-%s)<0.1""",
            (user['id'], metrics['ride_date'], metrics.get('dist_mi') or 0, metrics.get('duration_h') or 0))
        existing = cur.fetchone()
        if existing:
            cur.close(); conn.close()
            return {"ride_id": existing[0], "metrics": metrics, "coaching": "Already in your database.", "duplicate": True}

    cur.execute("""INSERT INTO rides (user_id,ride_date,start_time,name,dist_mi,duration_h,
        avg_power,norm_power,avg_hr,max_hr,avg_cadence,max_cadence,
        p5,p15,p30,p300,elev_gain_ft,elev_loss_ft,calories,avg_lr_balance,
        training_stress_score,intensity_factor,elapsed_h,ride_type,is_virtual,temp_c,notes,possible_duplicate_of)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
        (user['id'], metrics['ride_date'], metrics.get('start_time'), metrics.get('name'),
         metrics.get('dist_mi'), metrics.get('duration_h'),
         metrics.get('avg_power'), metrics.get('norm_power'),
         metrics.get('avg_hr'), metrics.get('max_hr'),
         metrics.get('avg_cadence'), metrics.get('max_cadence'),
         metrics.get('p5'), metrics.get('p15'), metrics.get('p30'), metrics.get('p300'),
         metrics.get('elev_gain_ft'), metrics.get('elev_loss_ft'), metrics.get('calories'),
         metrics.get('avg_lr_balance'), metrics.get('training_stress_score'),
         metrics.get('intensity_factor'), metrics.get('elapsed_h'), metrics.get('ride_type','General'),
         metrics.get('is_virtual', False), metrics.get('temp_c'), notes, duplicate_of_id))
    ride_id = cur.fetchone()[0]
    if streams:
        cur.execute("INSERT INTO ride_streams (ride_id, streams) VALUES (%s,%s)",
                    (ride_id, psycopg2.extras.Json(streams)))
    cur.close(); conn.close()
    coaching = await get_coaching_summary(user, metrics, ride_id=ride_id)
    if coaching and coaching != "AI coaching unavailable.":
        conn3 = get_db(); cur3 = conn3.cursor()
        cur3.execute("UPDATE rides SET coaching_synopsis=%s WHERE id=%s", (coaching, ride_id))
        cur3.close(); conn3.close()
    return {"ride_id": ride_id, "metrics": metrics, "coaching": coaching}

@app.get("/rides")
def get_rides(user: dict = Depends(get_current_user)):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM rides WHERE user_id=%s ORDER BY ride_date DESC LIMIT 200", (user['id'],))
    rides = [dict(r) for r in cur.fetchall()]; cur.close(); conn.close()
    return {"rides": rides, "count": len(rides)}

@app.get("/rides/export")
def export_rides(user: dict = Depends(get_current_user)):
    """Full ride history as a downloadable .xlsx workbook."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""SELECT r.*, e.name AS equipment_name FROM rides r
        LEFT JOIN equipment e ON r.equipment_id = e.id
        WHERE r.user_id=%s ORDER BY r.ride_date DESC""", (user['id'],))
    rides = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rides"

    headers = [
        "Date", "Name", "Type", "Distance (mi)", "Moving Time (h)", "Elapsed Time (h)",
        "Avg Power (W)", "Normalized Power (W)", "Avg HR (bpm)", "Max HR (bpm)",
        "Avg Cadence (rpm)", "Max Cadence (rpm)", "5s Best (W)", "15s Best (W)",
        "30s Best (W)", "5-min Best (W)", "Elevation Gain (ft)", "Elevation Loss (ft)",
        "Calories", "TSS", "IF", "L/R Balance (% right)", "Equipment", "Temp (C)",
        "Virtual", "Notes"
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rides:
        ws.append([
            str(r.get('ride_date')) if r.get('ride_date') else None,
            r.get('name'), r.get('ride_type'), r.get('dist_mi'), r.get('duration_h'),
            r.get('elapsed_h'), r.get('avg_power'), r.get('norm_power'), r.get('avg_hr'),
            r.get('max_hr'), r.get('avg_cadence'), r.get('max_cadence'), r.get('p5'),
            r.get('p15'), r.get('p30'), r.get('p300'), r.get('elev_gain_ft'),
            r.get('elev_loss_ft'), r.get('calories'), r.get('training_stress_score'),
            r.get('intensity_factor'), r.get('avg_lr_balance'), r.get('equipment_name'),
            r.get('temp_c'), ('Yes' if r.get('is_virtual') else 'No'), r.get('notes'),
        ])

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 30)
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "cycling_rides_export_" + get_local_today(None).isoformat() + ".xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=" + filename}
    )

@app.post("/notes")
def add_note(note: str = Form(...), user: dict = Depends(get_current_user)):
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO coaching_notes (user_id,note) VALUES (%s,%s)", (user['id'],note))
    cur.close(); conn.close()
    return {"status": "saved"}

@app.get("/notes")
def get_notes(user: dict = Depends(get_current_user)):
    """List saved personal notes, newest first — lets you verify a note actually saved."""
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, note, created_at FROM coaching_notes WHERE user_id=%s ORDER BY created_at DESC LIMIT 50", (user['id'],))
    notes = [dict(r) for r in cur.fetchall()]; cur.close(); conn.close()
    return {"notes": notes, "count": len(notes)}

@app.get("/equipment")
def get_equipment(user: dict = Depends(get_current_user)):
    """List the rider's bike/setup roster, built once (via the profile
    interview or added directly) and picked from per-ride rather than
    typed fresh each time."""
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT id, name FROM equipment WHERE user_id=%s ORDER BY name", (user['id'],))
    rows = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return {"equipment": rows}

@app.post("/equipment")
def add_equipment(name: str = Form(...), user: dict = Depends(get_current_user)):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cur.execute("INSERT INTO equipment (user_id, name) VALUES (%s,%s) ON CONFLICT (user_id, name) DO NOTHING RETURNING id",
                    (user['id'], name.strip()))
        row = cur.fetchone()
        if not row:
            cur.execute("SELECT id FROM equipment WHERE user_id=%s AND name=%s", (user['id'], name.strip()))
            row = cur.fetchone()
    finally:
        cur.close(); conn.close()
    return {"status": "saved", "id": row['id'], "name": name.strip()}

@app.post("/rides/{ride_id}/equipment")
def set_ride_equipment(ride_id: int, equipment_id: int = Form(...), user: dict = Depends(get_current_user)):
    """Set which bike/setup was used for a specific ride — a quick pick
    from the roster, separate from the Post-Ride Debrief note field."""
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id FROM equipment WHERE id=%s AND user_id=%s", (equipment_id, user['id']))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Equipment entry not found")
    cur.execute("UPDATE rides SET equipment_id=%s WHERE id=%s AND user_id=%s", (equipment_id, ride_id, user['id']))
    updated = cur.rowcount
    cur.close(); conn.close()
    if not updated:
        raise HTTPException(status_code=404, detail="Ride not found")
    return {"status": "saved", "ride_id": ride_id, "equipment_id": equipment_id}

@app.get("/profile")
def get_profile(user: dict = Depends(get_current_user)):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM profiles WHERE user_id=%s", (user['id'],))
    profile = cur.fetchone(); cur.close(); conn.close()
    return {"profile": dict(profile) if profile else None, "name": user['name']}

@app.post("/profile")
async def save_profile(
    age: str = Form(default=""),
    weight_lbs: str = Form(default=""),
    location: str = Form(default=""),
    fitness_level: str = Form(default=""),
    ftp: str = Form(default=""),
    annual_goal_mi: str = Form(default=""),
    other_goals: str = Form(default=""),
    health_notes: str = Form(default=""),
    injuries: str = Form(default=""),
    heat_tolerance: str = Form(default=""),
    medical_clearance: str = Form(default="false"),
    user: dict = Depends(get_current_user)
):
    conn = get_db(); cur = conn.cursor()
    cur.execute("""
        INSERT INTO profiles (user_id, age, weight_lbs, location, fitness_level, ftp,
            annual_goal_mi, other_goals, health_notes, injuries, heat_tolerance,
            medical_clearance, interview_complete, updated_at)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,true,NOW())
        ON CONFLICT (user_id) DO UPDATE SET
            age=EXCLUDED.age, weight_lbs=EXCLUDED.weight_lbs,
            location=EXCLUDED.location, fitness_level=EXCLUDED.fitness_level,
            ftp=EXCLUDED.ftp, annual_goal_mi=EXCLUDED.annual_goal_mi,
            other_goals=EXCLUDED.other_goals, health_notes=EXCLUDED.health_notes,
            injuries=EXCLUDED.injuries, heat_tolerance=EXCLUDED.heat_tolerance,
            medical_clearance=EXCLUDED.medical_clearance,
            interview_complete=true, updated_at=NOW()
    """, (
        user['id'],
        int(age) if age.strip() else None,
        float(weight_lbs) if weight_lbs.strip() else None,
        location or None, fitness_level or None,
        int(ftp) if ftp.strip() else None,
        int(annual_goal_mi) if annual_goal_mi.strip() else None,
        other_goals or None, health_notes or None,
        injuries or None, heat_tolerance or None,
        medical_clearance.lower() == 'true'
    ))
    cur.close(); conn.close()
    return {"status": "saved"}

@app.post("/profile/ftp")
def update_ftp_only(ftp: int = Form(...), user: dict = Depends(get_current_user)):
    """Targeted FTP-only update — deliberately separate from POST
    /profile above, which re-saves the WHOLE profile record from its
    full form and would silently wipe every other field (age, weight,
    goals, health notes, etc.) if called with only ftp set, since every
    other field defaults to an empty string there. This is the safe
    path for the ride-detail page's "apply this ride's estimated FTP"
    button, or any other single-field FTP update — one column, one
    targeted UPDATE, nothing else touched. Requires an existing profile
    row (created via the interview or POST /profile) to attach to."""
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE profiles SET ftp=%s, updated_at=NOW() WHERE user_id=%s", (ftp, user['id']))
    updated = cur.rowcount
    cur.close(); conn.close()
    if not updated:
        raise HTTPException(status_code=404, detail="No profile on file yet — complete the profile interview first")
    return {"status": "saved", "ftp": ftp}

@app.post("/interview")
async def ai_interview(
    message: str = Form(...),
    history: str = Form(default="[]"),
    user: dict = Depends(get_current_user)
):
    """Conversational AI entrance interview."""
    if not ANTHROPIC_KEY:
        return {"reply": "AI unavailable.", "profile_update": {}}
    
    import json as _json
    
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM profiles WHERE user_id=%s", (user['id'],))
    profile = cur.fetchone()
    cur.execute("SELECT note FROM coaching_notes WHERE user_id=%s ORDER BY created_at DESC LIMIT 5", (user['id'],))
    notes = [r['note'] for r in cur.fetchall()]
    cur.close(); conn.close()

    profile_ctx = ""
    if profile:
        profile_ctx = "\n\nEXISTING RIDER PROFILE (already on file — do not ask for this again):\n"
        if profile.get('age'):           profile_ctx += "- Age: " + str(profile['age']) + "\n"
        if profile.get('weight_lbs'):    profile_ctx += "- Weight: " + str(profile['weight_lbs']) + " lbs\n"
        if profile.get('location'):      profile_ctx += "- Location: " + str(profile['location']) + "\n"
        if profile.get('fitness_level'): profile_ctx += "- Fitness: " + str(profile['fitness_level']) + "\n"
        if profile.get('ftp'):           profile_ctx += "- FTP: " + str(profile['ftp']) + "W\n"
        if profile.get('annual_goal_mi'):profile_ctx += "- Annual goal: " + str(profile['annual_goal_mi']) + " mi\n"
        if profile.get('other_goals'):   profile_ctx += "- Goals: " + str(profile['other_goals']) + "\n"
        if profile.get('health_notes'):  profile_ctx += "- Health: " + str(profile['health_notes']) + "\n"
        if profile.get('injuries'):      profile_ctx += "- Injuries: " + str(profile['injuries']) + "\n"
        if profile.get('heat_tolerance'):profile_ctx += "- Heat tolerance: " + str(profile['heat_tolerance']) + "\n"
    if notes:
        profile_ctx += "\nPERSONAL NOTES: " + "; ".join(notes) + "\n"

    try:
        hist = _json.loads(history)
    except:
        hist = []

    system_prompt = """You are a friendly cycling coach conducting an ongoing coaching conversation with an athlete.
Your goal is to gather key information naturally through conversation:
- Age and weight
- Where they ride (city/region/climate — heat, altitude, terrain)
- Fitness level and cycling experience  
- FTP if they know it, or riding history
- Primary goals (mileage target, events, fitness, weight loss)
- Any injuries, recent illnesses, or medical conditions
- Heat tolerance and any history of heat-related issues
- Whether they have medical clearance if they mention serious conditions
- What bikes or setups they ride — if they mention more than one (e.g. road vs. gravel tires on the same bike, a separate TT or mountain bike, riding a trainer indoors), ask them to list each distinct setup; if they only have one bike with no variation, a single entry or skipping this is fine

IMPORTANT RULES:
- If they mention any serious cardiac conditions, recent surgery, chest pain during exercise, or uncontrolled medical conditions: ALWAYS say they should consult their doctor before continuing and ask if they have medical clearance.
- If they mention wanting to lose weight: acknowledge it warmly but note that cycling supports overall health — direct specific dietary advice to a nutritionist.
- If they mention recent COVID, flu, mono, or similar illness: briefly note the post-viral performance dip and adjust expectations.
- Keep responses conversational, warm, 2-4 sentences max.
- After gathering enough info (3-4 exchanges), summarize what you've learned and ask if there's anything else important to share.
- End by saying their profile has been saved and coaching will be personalized to them." + profile_ctx + "

At the END of your response, on a new line, output a JSON object (and ONLY the JSON, no other text on that line) with any profile fields you extracted:
{"age":null,"weight_lbs":null,"location":null,"fitness_level":null,"ftp":null,"annual_goal_mi":null,"other_goals":null,"health_notes":null,"injuries":null,"heat_tolerance":null,"medical_clearance":false,"equipment":null}
Only include fields where you extracted real information. Use null for unknown fields. For "equipment", if distinct bikes/setups were mentioned, use a list of short names (e.g. ["Road bike - 28mm tires","Gravel setup - 45mm tires","Indoor trainer"]); otherwise null."""

    messages = []
    if not hist:
        messages.append({
            "role": "assistant",
            "content": "Hi " + user['name'] + "! I'm your cycling coach. Before we dive into your rides, I'd love to learn a bit about you. Tell me — how long have you been cycling, and what got you into it?"
        })
    
    for h in hist:
        messages.append(h)
    messages.append({"role": "user", "content": message})

    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01"},
                json={"model": "claude-sonnet-4-6", "max_tokens": 500,
                      "system": system_prompt, "messages": messages},
                timeout=30
            )
            full_reply = resp.json()['content'][0]['text']
    except Exception as e:
        return {"reply": "Sorry, I had trouble connecting. Please try again.", "profile_update": {}}

    profile_update = {}
    lines = full_reply.strip().split('\n')
    reply_text = full_reply
    for line in reversed(lines):
        line = line.strip()
        if line.startswith('{') and line.endswith('}'):
            try:
                extracted = _json.loads(line)
                profile_update = {k: v for k, v in extracted.items() if v is not None and v != False}
                reply_text = '\n'.join(lines[:-1]).strip()
                break
            except:
                pass

    if profile_update:
        equipment_list = profile_update.pop('equipment', None)
        if equipment_list and isinstance(equipment_list, list):
            conn_e = get_db(); cur_e = conn_e.cursor()
            for eq_name in equipment_list:
                if eq_name and isinstance(eq_name, str):
                    cur_e.execute(
                        "INSERT INTO equipment (user_id, name) VALUES (%s,%s) ON CONFLICT (user_id, name) DO NOTHING",
                        (user['id'], eq_name.strip())
                    )
            cur_e.close(); conn_e.close()

    if profile_update:
        conn = get_db(); cur = conn.cursor()
        fields = list(profile_update.keys())
        vals = [profile_update[f] for f in fields]
        set_clause = ', '.join(f + '=%s' for f in fields)
        cur.execute(
            "INSERT INTO profiles (user_id, " + ', '.join(fields) + ") VALUES (%s" + ',%s'*len(fields) + ") "
            "ON CONFLICT (user_id) DO UPDATE SET " + set_clause + ", updated_at=NOW()",
            [user['id']] + vals + vals
        )
        cur.close(); conn.close()

    return {"reply": reply_text, "profile_update": profile_update}

# ── Document Import ───────────────────────────────────────────────────────

IMPORT_MAX_STORE_CHARS  = 20000
IMPORT_MAX_PROMPT_CHARS = 6000

@app.post("/coaching/import")
async def import_doc(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload a text/markdown document as context for the coaching chat.
    Not for medical records, lab results, or other health/PHI documents."""
    data = await file.read()
    try:
        text = data.decode('utf-8', errors='ignore')
    except Exception:
        raise HTTPException(status_code=400, detail="Could not read as text. .txt and .md files only for now.")
    if not text.strip():
        raise HTTPException(status_code=400, detail="File appears to be empty.")
    truncated = len(text) > IMPORT_MAX_STORE_CHARS
    text = text[:IMPORT_MAX_STORE_CHARS]
    conn = get_db(); cur = conn.cursor()
    cur.execute("INSERT INTO imported_docs (user_id, filename, content) VALUES (%s,%s,%s) RETURNING id",
                (user['id'], file.filename, text))
    doc_id = cur.fetchone()[0]; cur.close(); conn.close()
    return {"status": "imported", "id": doc_id, "filename": file.filename,
            "chars": len(text), "truncated": truncated}

@app.get("/coaching/imports")
def list_imports(user: dict = Depends(get_current_user)):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""SELECT id, filename, LEFT(content,150) AS preview, created_at
        FROM imported_docs WHERE user_id=%s ORDER BY created_at DESC LIMIT 20""", (user['id'],))
    docs = [dict(r) for r in cur.fetchall()]; cur.close(); conn.close()
    return {"imports": docs, "count": len(docs)}

@app.delete("/coaching/imports/{doc_id}")
def delete_import(doc_id: int, user: dict = Depends(get_current_user)):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM imported_docs WHERE id=%s AND user_id=%s", (doc_id, user['id']))
    deleted = cur.rowcount
    cur.close(); conn.close()
    if not deleted:
        raise HTTPException(status_code=404, detail="Not found")
    return {"status": "deleted"}

# ── Post-ride Coaching Chat ──────────────────────────────────────────────────

@app.post("/coaching/chat")
async def coaching_chat(
    message: str = Form(...),
    history: str = Form(default="[]"),
    user: dict = Depends(get_current_user)
):
    """Ongoing post-ride coaching conversation. Distinct from /interview —
    this is not intake, it's a coach who already knows the rider talking
    through their recent rides, recovery, and trends."""
    if not ANTHROPIC_KEY:
        return {"reply": "AI coaching unavailable."}

    import json as _json

    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM profiles WHERE user_id=%s", (user['id'],))
    profile = cur.fetchone()
    cur.execute("SELECT * FROM rides WHERE user_id=%s ORDER BY ride_date DESC LIMIT 5", (user['id'],))
    recent = [dict(r) for r in cur.fetchall()]
    cur.execute("""SELECT COUNT(*) AS n, COALESCE(SUM(dist_mi),0) AS mi,
        COALESCE(SUM(duration_h),0) AS hrs, COALESCE(SUM(elev_gain_ft),0) AS elev
        FROM rides WHERE user_id=%s AND ride_date >= %s AND ride_date < %s""",
        (user['id'], f'{YEAR}-01-01', f'{YEAR+1}-01-01'))
    ytd = cur.fetchone()
    cur.execute("SELECT note FROM coaching_notes WHERE user_id=%s ORDER BY created_at DESC LIMIT 5", (user['id'],))
    notes = [r['note'] for r in cur.fetchall()]
    cur.execute("SELECT filename, content FROM imported_docs WHERE user_id=%s ORDER BY created_at DESC LIMIT 1", (user['id'],))
    imported = cur.fetchone()
    cur.execute("SELECT entry_date, summary FROM coaching_memory_log WHERE user_id=%s ORDER BY entry_date DESC LIMIT 30", (user['id'],))
    memory_log = [dict(r) for r in cur.fetchall()]
    cur.execute("SELECT theme, content FROM coaching_memory_themes WHERE user_id=%s", (user['id'],))
    memory_themes = {r['theme']: r['content'] for r in cur.fetchall()}
    cur.close(); conn.close()

    goal = int(profile['annual_goal_mi']) if profile and profile.get('annual_goal_mi') else ANNUAL_GOAL
    total_mi = float(ytd['mi'] or 0)
    local_now = get_local_now(profile.get('timezone') if profile else None)
    local_today = local_now.date()
    current_dt_ctx = "\nCURRENT DATE/TIME (rider's actual local time): " + local_now.strftime('%A, %B %d, %Y, %I:%M %p') + "\n"
    day_of_year = local_today.timetuple().tm_yday
    pace_mi = goal * day_of_year / 366
    pace_diff = round(total_mi - pace_mi, 1)
    pct_complete = round(total_mi / goal * 100, 1) if goal else 0

    ytd_ctx = (
        "\nYEAR-TO-DATE PROGRESS (as of " + local_today.strftime('%Y-%m-%d') + "):\n"
        + "- Total this year: " + str(round(total_mi,1)) + " mi across " + str(ytd['n']) + " rides, "
        + str(round(float(ytd['hrs'] or 0),1)) + " hours, " + str(round(float(ytd['elev'] or 0))) + " ft climbed\n"
        + "- Annual goal: " + str(goal) + " mi (" + str(pct_complete) + "% complete)\n"
        + "- Pace: " + ("ahead of" if pace_diff >= 0 else "behind") + " schedule by " + str(abs(pace_diff)) + " mi\n"
    )

    profile_ctx = ""
    if profile:
        profile_ctx = "\nRIDER PROFILE:\n"
        if profile.get('age'):            profile_ctx += "- Age: " + str(profile['age']) + "\n"
        if profile.get('weight_lbs'):     profile_ctx += "- Weight: " + str(profile['weight_lbs']) + " lbs\n"
        if profile.get('location'):       profile_ctx += "- Location: " + str(profile['location']) + "\n"
        if profile.get('fitness_level'):  profile_ctx += "- Fitness: " + str(profile['fitness_level']) + "\n"
        if profile.get('ftp'):            profile_ctx += "- FTP: " + str(profile['ftp']) + "W\n"
        if profile.get('annual_goal_mi'): profile_ctx += "- Annual goal: " + str(profile['annual_goal_mi']) + " mi\n"
        if profile.get('other_goals'):    profile_ctx += "- Goals: " + str(profile['other_goals']) + "\n"
        if profile.get('health_notes'):   profile_ctx += "- Health: " + str(profile['health_notes']) + "\n"
        if profile.get('injuries'):       profile_ctx += "- Injuries: " + str(profile['injuries']) + "\n"
        if profile.get('heat_tolerance'): profile_ctx += "- Heat tolerance: " + str(profile['heat_tolerance']) + "\n"

    rides_ctx = ""
    if recent:
        rides_ctx = "\nRECENT RIDES (most recent first):\n"
        for r in recent:
            rides_ctx += (
                "- " + str(r.get('ride_date',''))[:10] + ": " + str(r.get('dist_mi','?')) + "mi, "
                + str(r.get('duration_h','?')) + "h, avg pwr " + str(r.get('avg_power','?')) + "W"
                + " (NP " + str(r.get('norm_power','?')) + "W), "
                + "avg HR " + str(r.get('avg_hr','?')) + " (max " + str(r.get('max_hr','?')) + "), "
                + "elev " + str(r.get('elev_gain_ft','?')) + "ft, "
                + "bests 5s/15s/30s/5min: " + str(r.get('p5','?')) + "/" + str(r.get('p15','?'))
                + "/" + str(r.get('p30','?')) + "/" + str(r.get('p300','?')) + "W"
                + (", L/R balance " + str(r.get('avg_lr_balance')) + "% right" if r.get('avg_lr_balance') is not None else "")
                + (", virtual" if r.get('is_virtual') else "") + "\n"
            )

    notes_ctx = ""
    if notes:
        notes_ctx = "\nPERSONAL NOTES: " + "; ".join(notes) + "\n"

    imported_ctx = ""
    if imported:
        imported_ctx = (
            "\nIMPORTED CONTEXT (" + str(imported['filename']) + "):\n"
            + str(imported['content'])[:IMPORT_MAX_PROMPT_CHARS] + "\n"
        )

    memory_log_ctx = ""
    if memory_log:
        memory_log_ctx = "\nCOACHING MEMORY — DATED LOG (most recent first):\n"
        for m in memory_log:
            memory_log_ctx += "- " + str(m['entry_date']) + ": " + str(m['summary']) + "\n"

    memory_themes_ctx = ""
    if memory_themes:
        memory_themes_ctx = "\nCOACHING MEMORY — STANDING PATTERNS:\n"
        for key, label in MEMORY_THEMES.items():
            if memory_themes.get(key):
                memory_themes_ctx += "- " + label + ": " + str(memory_themes[key]) + "\n"

    system_prompt = (
        "You are an ongoing cycling coach chatting with an athlete after their rides. "
        "This is NOT an intake interview — you already know them from the profile and ride "
        "data below. Talk about how their recent ride(s) felt, recovery, trends versus their "
        "goals, and give specific, practical guidance for what's next. You have their exact "
        "year-to-date mileage and pace vs. goal below — use those real numbers, never ask the "
        "rider to tell you their own totals."
        + current_dt_ctx + profile_ctx + ytd_ctx + rides_ctx + notes_ctx + imported_ctx + memory_log_ctx + memory_themes_ctx +
        "\n\nIMPORTANT RULES:\n"
        "- If asked what the date or time is, answer directly and confidently from the "
        "CURRENT DATE/TIME line above — that's the rider's real local time, not a guess.\n"
        "- If they mention chest pain, serious cardiac symptoms, or any acute medical concern: "
        "tell them clearly to stop and see a doctor. Do not downplay it.\n"
        "- If they mention wanting to lose weight, acknowledge it warmly but redirect specific "
        "dietary advice to a nutritionist — you can speak to training load, not diet plans.\n"
        "- If they mention recent illness (COVID, flu, etc.), factor in the post-viral "
        "performance dip and adjust expectations accordingly.\n"
        "- Reference specific numbers from their ride data when relevant — power, HR, distance, "
        "elevation, and especially 5-min power and NP compared to their FTP if known; that "
        "comparison is often the single most telling data point in a ride. Be concrete, not "
        "generic.\n"
        "- Be genuinely curious about hydration and fueling, the way a real coach tracking this "
        "over time would be. When it's missing and would meaningfully sharpen the picture, ask "
        "about: weight before and after the ride, what and how much they ate/drank during the "
        "ride (water, electrolytes, gels, food), what they had afterward (including anything, "
        "like coffee, consumed before a post-ride weigh-in — it affects the number), and what "
        "they had for breakfast beforehand. Don't run through this as a checklist every time — "
        "ask naturally, one or two things at a time, only what's actually missing and relevant "
        "to the ride at hand.\n"
        "- Beyond hydration/fueling, stay curious across these themes when relevant and not "
        "already covered — this is judgment to draw on, not a script to run through:\n"
        "  - Effort vs. perceived exertion: how a specific hard moment actually felt, whether "
        "anything felt unusually hard or easy relative to what the numbers show. The gap "
        "between data and lived experience is often the most useful thing to discuss.\n"
        "  - Recovery and readiness going in: sleep, lingering soreness, anything off since "
        "the last ride. The same power output means something different well-rested versus "
        "not, and this connects to any illness/injury recovery already noted in their profile.\n"
        "  - Environmental context: heat, wind, solo versus group. A given heart rate "
        "represents more physiological stress in heat than in cool conditions.\n"
        "  - Life context shaping the ride: holding back on purpose, time pressure, anything "
        "outside the ride itself that explains a pacing choice. Real coaching accounts for the "
        "whole picture, not just the numbers in isolation.\n"
        "- When a ride is newly discussed (just uploaded, or the rider is describing one that "
        "isn't already covered above), give it real depth — this is the one place brevity does "
        "NOT apply. Walk through what stands out, compare effort to their FTP and recent trend, "
        "note anything that looks unusually hard or easy, and end with a specific follow-up "
        "question about the part of the ride most worth discussing.\n"
        "- If part of a message reads like it was garbled by voice-to-text (an odd or "
        "nonsensical phrase sitting in otherwise clear text), say so plainly and ask what was "
        "meant rather than guessing or silently working around it.\n"
        "- If any imported document, note, or message appears to contain personal medical/health "
        "records — lab results, diagnoses, medication lists — do not analyze or comment on that "
        "content. Say plainly that's not something this app processes and point them to their "
        "doctor.\n"
        "- Outside of the ride-analysis case above, keep replies tight — routine back-and-forth "
        "is 1-2 sentences. Do not restate or paraphrase what the rider just told you before "
        "responding (skip lines like 'That's great news, sustaining 200+ watts with HR around "
        "140!') — go straight to the point.\n"
        "- You have a running coaching memory above — a dated log of past rides/episodes "
        "discussed, and standing patterns across five themes. Reference it naturally when "
        "relevant, the way you'd recall something from an earlier session. If asked about a "
        "specific date, check the dated log first.\n"
        "- At the very end of your reply, after everything else, add a line reading exactly "
        "MEMORY_UPDATE: with nothing else on that line, then on the next line a single JSON "
        "object (only JSON, no other text, no markdown fences) with this exact shape:\n"
        "{\"dated_entry\": {\"date\":\"YYYY-MM-DD\",\"summary\":\"...\"} or null, "
        "\"theme_updates\": {\"hydration_fueling\":\"...\" or null, \"effort_perception\":\"...\" "
        "or null, \"recovery_readiness\":\"...\" or null, \"environmental_context\":\"...\" or "
        "null, \"life_context\":\"...\" or null, \"search_preferences\":\"...\" or null}}\n"
        "Only include a dated_entry if a specific ride or dated episode was actually discussed "
        "with something worth remembering later — skip it for routine check-ins with nothing "
        "new. Only fill in a theme_update for a theme this exchange actually informed; leave "
        "the rest null. Each theme_update you DO include must be the FULL updated standing "
        "pattern for that theme (folding in what's new with what's already there above), not "
        "just the new piece — it replaces the old content entirely. Keep every summary and "
        "theme update to 2-3 sentences, distilled and durable, not verbatim conversation.\n"
        "- Tone throughout: a knowledgeable coach, direct and matter-of-fact — not a chatty "
        "friend, but genuinely engaged with the specifics of what they tell you."
    )

    try:
        hist = _json.loads(history)
    except:
        hist = []

    messages = list(hist) + [{"role": "user", "content": message}]

    try:
        reply = await run_claude_with_tools(system_prompt, messages, user['id'], max_tokens=900)
    except Exception as e:
        return {"reply": "Sorry, I had trouble connecting. Please try again."}

    reply_text = reply
    if 'MEMORY_UPDATE:' in reply:
        parts = reply.split('MEMORY_UPDATE:', 1)
        reply_text = parts[0].strip()
        try:
            mem_update = extract_json_object(parts[1].strip())
        except Exception:
            mem_update = {}
        if mem_update:
            conn2 = get_db(); cur2 = conn2.cursor()
            de = mem_update.get('dated_entry')
            if de and de.get('date') and de.get('summary'):
                cur2.execute("""
                    INSERT INTO coaching_memory_log (user_id, entry_date, summary)
                    VALUES (%s,%s,%s)
                    ON CONFLICT (user_id, entry_date) DO UPDATE SET summary=EXCLUDED.summary
                """, (user['id'], de['date'], de['summary']))
            tu = mem_update.get('theme_updates') or {}
            for theme_key in MEMORY_THEMES:
                content = tu.get(theme_key)
                if content:
                    cur2.execute("""
                        INSERT INTO coaching_memory_themes (user_id, theme, content, updated_at)
                        VALUES (%s,%s,%s,NOW())
                        ON CONFLICT (user_id, theme) DO UPDATE SET content=EXCLUDED.content, updated_at=NOW()
                    """, (user['id'], theme_key, content))
            cur2.close(); conn2.close()

    return {"reply": reply_text}

# ── Coaching Memory (dated log + standing themes) ────────────────────────────

@app.get("/coaching/memory")
def get_memory(user: dict = Depends(get_current_user)):
    """View what's currently stored in coaching memory — the dated log and the
    five standing-pattern threads. Each dated_log entry now also carries a
    `flagged` bool — true when a ride on that date is flagged as a possible
    duplicate (possible_duplicate_of IS NOT NULL) — so the Ride History list
    can show a badge in-context, not just the one dedicated review card on
    the Dashboard, which is easy to miss if a rider doesn't happen to check
    it. Read-only enrichment; doesn't change what's stored."""
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT l.id, l.entry_date, l.summary,
            EXISTS(
                SELECT 1 FROM rides r
                WHERE r.user_id = %s AND r.ride_date = l.entry_date
                    AND r.possible_duplicate_of IS NOT NULL
            ) AS flagged
        FROM coaching_memory_log l
        WHERE l.user_id = %s
        ORDER BY l.entry_date DESC LIMIT 50
    """, (user['id'], user['id']))
    log = [dict(r) for r in cur.fetchall()]
    cur.execute("SELECT theme, content, updated_at FROM coaching_memory_themes WHERE user_id=%s", (user['id'],))
    themes = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return {"dated_log": log, "themes": themes}

@app.post("/coaching/memory/log/{entry_id}")
def edit_memory_log_entry(
    entry_id: int,
    entry_date: str = Form(default=None),
    summary: str = Form(default=None),
    user: dict = Depends(get_current_user)
):
    """Fix a dated log entry — e.g. a wrong date from a timezone slip during
    import. Only the fields provided are changed; the other stays as-is."""
    if not entry_date and not summary:
        raise HTTPException(status_code=400, detail="Provide entry_date and/or summary to update")
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id FROM coaching_memory_log WHERE id=%s AND user_id=%s", (entry_id, user['id']))
    if not cur.fetchone():
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Entry not found")
    try:
        if entry_date:
            cur.execute("UPDATE coaching_memory_log SET entry_date=%s WHERE id=%s", (entry_date, entry_id))
        if summary:
            cur.execute("UPDATE coaching_memory_log SET summary=%s WHERE id=%s", (summary, entry_id))
    except psycopg2.errors.UniqueViolation:
        conn.rollback(); cur.close(); conn.close()
        raise HTTPException(status_code=409, detail="Another entry already exists for that date — delete or edit that one first")
    cur.close(); conn.close()
    return {"status": "updated", "id": entry_id}

@app.delete("/coaching/memory/log/{entry_id}")
def delete_memory_log_entry(entry_id: int, user: dict = Depends(get_current_user)):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM coaching_memory_log WHERE id=%s AND user_id=%s", (entry_id, user['id']))
    deleted = cur.rowcount
    cur.close(); conn.close()
    if not deleted:
        raise HTTPException(status_code=404, detail="Entry not found")
    return {"status": "deleted"}

@app.post("/coaching/memory/seed")
async def seed_memory(text: str = Form(...), user: dict = Depends(get_current_user)):
    """One-time import: extract dated log entries and standing-pattern updates from
    historical conversation content (e.g. past coaching sessions elsewhere) and
    seed them into memory, so the ongoing coach starts from real history instead
    of zero. Uses the same distillation approach as the ongoing memory updates —
    this is a retroactive application of it, not a separate mechanism."""
    if not ANTHROPIC_KEY:
        raise HTTPException(status_code=503, detail="AI unavailable")

    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT theme, content FROM coaching_memory_themes WHERE user_id=%s", (user['id'],))
    existing_themes = {r['theme']: r['content'] for r in cur.fetchall()}
    cur.close(); conn.close()

    existing_ctx = ""
    if existing_themes:
        existing_ctx = "\nEXISTING STANDING PATTERNS (fold new content into these, don't discard):\n"
        for key, label in MEMORY_THEMES.items():
            if existing_themes.get(key):
                existing_ctx += "- " + label + ": " + str(existing_themes[key]) + "\n"

    extract_prompt = (
        "You are extracting structured coaching memory from historical conversation content "
        "between a cyclist and their coach. Read the raw content below and extract:\n"
        "1. Dated log entries — one per distinct ride or dated episode actually discussed, "
        "2-3 sentences each, distilled from what was said, not verbatim quotes.\n"
        "2. Updated standing patterns for these five themes, based on everything below "
        "collectively: hydration_fueling, effort_perception, recovery_readiness, "
        "environmental_context, life_context. Only include a theme if the content actually "
        "informs it — leave the rest null.\n"
        + existing_ctx +
        "\nOutput ONLY a JSON object, no other text, no markdown fences, shaped exactly like:\n"
        "{\"dated_entries\": [{\"date\":\"YYYY-MM-DD\",\"summary\":\"...\"}], "
        "\"theme_updates\": {\"hydration_fueling\":\"...\" or null, \"effort_perception\":\"...\" "
        "or null, \"recovery_readiness\":\"...\" or null, \"environmental_context\":\"...\" or "
        "null, \"life_context\":\"...\" or null}}\n\n"
        "RAW CONTENT:\n" + text[:150000]
    )

    try:
        async with httpx.AsyncClient() as client:
            resp = await client.post(
                "https://api.anthropic.com/v1/messages",
                headers={"x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01"},
                json={"model": "claude-sonnet-4-6", "max_tokens": 3000,
                      "messages": [{"role": "user", "content": extract_prompt}]},
                timeout=90
            )
            raw = resp.json()['content'][0]['text']
    except Exception as e:
        raise HTTPException(status_code=502, detail="Extraction request failed: " + str(e))

    try:
        parsed = extract_json_object(raw)
    except Exception:
        raise HTTPException(status_code=502, detail="Could not parse extraction result as JSON")

    entries = parsed.get('dated_entries') or []
    theme_updates = parsed.get('theme_updates') or {}

    conn = get_db(); cur = conn.cursor()
    added = 0
    for e in entries:
        if e.get('date') and e.get('summary'):
            cur.execute("""
                INSERT INTO coaching_memory_log (user_id, entry_date, summary)
                VALUES (%s,%s,%s)
                ON CONFLICT (user_id, entry_date) DO UPDATE SET summary=EXCLUDED.summary
            """, (user['id'], e['date'], e['summary']))
            added += 1
    updated_themes = []
    for theme_key in MEMORY_THEMES:
        content = theme_updates.get(theme_key)
        if content:
            cur.execute("""
                INSERT INTO coaching_memory_themes (user_id, theme, content, updated_at)
                VALUES (%s,%s,%s,NOW())
                ON CONFLICT (user_id, theme) DO UPDATE SET content=EXCLUDED.content, updated_at=NOW()
            """, (user['id'], theme_key, content))
            updated_themes.append(theme_key)
    cur.close(); conn.close()

    return {"status": "seeded", "dated_entries_added": added, "themes_updated": updated_themes}

# ── Strava Integration ───────────────────────────────────────────────────────

@app.get("/debug/dashboard")
def debug_dashboard(user: dict = Depends(get_current_user)):
    """Debug endpoint to show dashboard error details."""
    import traceback
    try:
        conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM profiles WHERE user_id=%s", (user['id'],))
        profile = cur.fetchone()
        user_goal = int(profile['annual_goal_mi']) if profile and profile.get('annual_goal_mi') else ANNUAL_GOAL
        cur.execute("SELECT * FROM rides WHERE user_id=%s AND ride_date >= %s AND ride_date < %s ORDER BY ride_date ASC LIMIT 5",
            (user['id'], f'{YEAR}-01-01', f'{YEAR+1}-01-01'))
        rides = [dict(r) for r in cur.fetchall()]; cur.close(); conn.close()
        result = build_full_dashboard(rides, user['name'], annual_goal=user_goal, user_timezone=profile.get('timezone') if profile else None)
        return {"status": "ok", "html_length": len(result), "rides": len(rides), "goal": user_goal}
    except Exception as e:
        return {"status": "error", "error": str(e), "traceback": traceback.format_exc()}

@app.get("/admin/users")
def admin_list_users(user: dict = Depends(get_current_user)):
    """Roster of everyone signed up — restricted to ADMIN_EMAILS."""
    if user['email'] not in ADMIN_EMAILS:
        raise HTTPException(status_code=403, detail="Not authorized")
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT u.id, u.name, u.email, u.created_at,
            (SELECT COUNT(*) FROM rides r WHERE r.user_id=u.id) AS ride_count,
            (SELECT MAX(ride_date) FROM rides r WHERE r.user_id=u.id) AS last_ride,
            EXISTS(SELECT 1 FROM strava_tokens st WHERE st.user_id=u.id) AS strava_connected,
            EXISTS(SELECT 1 FROM profiles p WHERE p.user_id=u.id AND p.interview_complete=true) AS profile_complete
        FROM users u ORDER BY u.created_at DESC
    """)
    users = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return {"users": users, "count": len(users)}

@app.get("/strava/connect")
def strava_connect(_auth: str = ""):
    """Redirect user to Strava OAuth page. Token passed as _auth query param."""
    from urllib.parse import urlencode
    from fastapi.responses import RedirectResponse, HTMLResponse
    if not _auth:
        return HTMLResponse("<h2>Missing auth token. Please try again from the app.</h2>")
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id FROM users WHERE token=%s", (_auth,))
    user = cur.fetchone(); cur.close(); conn.close()
    if not user:
        return HTMLResponse("<h2>Invalid session. Please log in again.</h2>")
    params = {
        "client_id":       STRAVA_CLIENT_ID,
        "redirect_uri":    STRAVA_REDIRECT_URI,
        "response_type":   "code",
        "approval_prompt": "force",
        "scope":           "activity:read_all",
        "state":           _auth
    }
    return RedirectResponse(STRAVA_AUTH_URL + "?" + urlencode(params))

@app.get("/strava/callback")
async def strava_callback(code: str, state: str, error: str = None):
    """Handle Strava OAuth callback — exchange code for tokens."""
    from fastapi.responses import HTMLResponse
    if error:
        return HTMLResponse("<h2>Strava connection cancelled.</h2><p>You can close this window.</p>")
    
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM users WHERE token=%s", (state,))
    user = cur.fetchone()
    if not user:
        cur.close(); conn.close()
        return HTMLResponse("<h2>Invalid session. Please try again.</h2>")

    async with httpx.AsyncClient() as client:
        resp = await client.post(STRAVA_TOKEN_URL, data={
            "client_id":     STRAVA_CLIENT_ID,
            "client_secret": STRAVA_CLIENT_SECRET,
            "code":          code,
            "grant_type":    "authorization_code"
        })
        data = resp.json()

    if "access_token" not in data:
        cur.close(); conn.close()
        return HTMLResponse("<h2>Strava connection failed.</h2><p>" + str(data) + "</p>")

    cur.execute("""
        INSERT INTO strava_tokens (user_id, athlete_id, access_token, refresh_token, expires_at)
        VALUES (%s,%s,%s,%s,%s)
        ON CONFLICT (user_id) DO UPDATE SET
            athlete_id=EXCLUDED.athlete_id,
            access_token=EXCLUDED.access_token,
            refresh_token=EXCLUDED.refresh_token,
            expires_at=EXCLUDED.expires_at
    """, (user['id'], data.get('athlete',{}).get('id'),
          data['access_token'], data['refresh_token'], data['expires_at']))
    cur.close(); conn.close()

    return HTMLResponse("""
        <html><body style="font-family:Inter,sans-serif;text-align:center;padding:40px;">
        <h2 style="color:#27AE60;">✓ Strava Connected!</h2>
        <p>Your Strava account is now linked. You can close this window and return to the app.</p>
        <script>setTimeout(()=>window.close(),3000);</script>
        </body></html>
    """)

@app.get("/strava/status")
def strava_status(user: dict = Depends(get_current_user)):
    """Check if user has Strava connected."""
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT athlete_id, last_sync FROM strava_tokens WHERE user_id=%s", (user['id'],))
    token = cur.fetchone(); cur.close(); conn.close()
    return {"connected": token is not None, "last_sync": str(token['last_sync']) if token and token['last_sync'] else None}

@app.post("/strava/sync")
async def strava_sync(
    days_back: int = Form(default=90),
    user: dict = Depends(get_current_user)
):
    """Pull recent activities from Strava and store them."""
    import time
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM strava_tokens WHERE user_id=%s", (user['id'],))
    token_row = cur.fetchone()
    if not token_row:
        cur.close(); conn.close()
        raise HTTPException(status_code=400, detail="Strava not connected")

    access_token = token_row['access_token']
    if token_row['expires_at'] and int(time.time()) > token_row['expires_at'] - 300:
        async with httpx.AsyncClient() as client:
            resp = await client.post(STRAVA_TOKEN_URL, data={
                "client_id":     STRAVA_CLIENT_ID,
                "client_secret": STRAVA_CLIENT_SECRET,
                "grant_type":    "refresh_token",
                "refresh_token": token_row['refresh_token']
            })
            new_tokens = resp.json()
        if "access_token" in new_tokens:
            access_token = new_tokens['access_token']
            cur2 = conn.cursor()
            cur2.execute("UPDATE strava_tokens SET access_token=%s, refresh_token=%s, expires_at=%s WHERE user_id=%s",
                        (access_token, new_tokens['refresh_token'], new_tokens['expires_at'], user['id']))
            cur2.close()

    days_back_ts  = int(time.time()) - (days_back * 86400)
    year_start_ts = int(datetime(YEAR, 1, 1).timestamp())
    after_ts = max(days_back_ts, year_start_ts)
    imported = 0; skipped = 0; errors = 0; out_of_range = 0; flagged = 0; non_cycling = 0
    page = 1

    async with httpx.AsyncClient() as client:
        while True:
            resp = await client.get(
                "https://www.strava.com/api/v3/athlete/activities",
                headers={"Authorization": "Bearer " + access_token},
                params={"after": after_ts, "per_page": 50, "page": page}
            )
            activities = resp.json()
            if not activities or not isinstance(activities, list):
                break

            for act in activities:
                try:
                    activity_type = act.get('sport_type') or act.get('type') or ''
                    if activity_type not in CYCLING_ACTIVITY_TYPES:
                        non_cycling += 1
                        continue
                    start_local_raw = act.get('start_date_local','')
                    act_date = start_local_raw[:10]
                    start_time_val = start_local_raw[:19] if len(start_local_raw) >= 19 else None
                    if act_date < f'{YEAR}-01-01' or act_date >= f'{YEAR+1}-01-01':
                        out_of_range += 1
                        continue
                    dist_mi  = round((act.get('distance') or 0) / 1609.34, 2)
                    dur_h    = round((act.get('moving_time') or 0) / 3600, 2)
                    elapsed_h = round((act.get('elapsed_time') or 0) / 3600, 2) if act.get('elapsed_time') else None
                    sport    = act.get('sport_type','').lower()
                    is_virt  = act.get('trainer', False) or 'virtual' in sport or 'zwift' in (act.get('name','') or '').lower()

                    cur3 = conn.cursor()
                    new_end_time = None
                    if start_time_val and elapsed_h:
                        try:
                            new_end_time = (datetime.fromisoformat(start_time_val) + timedelta(hours=elapsed_h)).isoformat()
                        except Exception:
                            new_end_time = None

                    duplicate_of_id = None
                    if start_time_val and new_end_time:
                        cur3.execute("""SELECT id FROM rides WHERE user_id=%s
                            AND ABS(ride_date - %s::date) <= 1
                            AND start_time IS NOT NULL AND elapsed_h IS NOT NULL
                            AND start_time <= %s::timestamp
                            AND %s::timestamp <= (start_time + (elapsed_h * INTERVAL '1 hour'))
                            LIMIT 1""",
                            (user['id'], act_date, new_end_time, start_time_val))
                        overlap_match = cur3.fetchone()
                        if overlap_match:
                            duplicate_of_id = overlap_match[0]

                    if duplicate_of_id is None:
                        cur3.execute("""SELECT id FROM rides WHERE user_id=%s
                            AND ABS(ride_date - %s::date) <= 1
                            AND (start_time IS NULL OR elapsed_h IS NULL)
                            AND ABS(COALESCE(dist_mi,0)-%s)<0.5 AND ABS(COALESCE(duration_h,0)-%s)<0.1""",
                            (user['id'], act_date, dist_mi, dur_h))
                        if cur3.fetchone():
                            cur3.close(); skipped += 1; continue

                    stream_resp = await client.get(
                        f"https://www.strava.com/api/v3/activities/{act['id']}/streams",
                        headers={"Authorization": "Bearer " + access_token},
                        params={"keys": "watts,heartrate,cadence,altitude,distance,time,velocity_smooth", "key_by_type": "true"}
                    )
                    raw_streams = stream_resp.json()

                    def stream_vals(key):
                        s = raw_streams.get(key,{})
                        return s.get('data',[]) if isinstance(s, dict) else []

                    powers   = [v for v in stream_vals('watts')     if v and v > 0]
                    hrs      = [v for v in stream_vals('heartrate')  if v]
                    cads     = [v for v in stream_vals('cadence')    if v]
                    alts     = stream_vals('altitude')

                    def best_avg(vals, n):
                        if not vals or len(vals) < n: return max(vals) if vals else None
                        return round(max(sum(vals[i:i+n])/n for i in range(len(vals)-n+1)))

                    np_val = None
                    if powers and len(powers) > 30:
                        sm = [sum(powers[max(0,i-29):i+1])/len(powers[max(0,i-29):i+1]) for i in range(len(powers))]
                        np_val = round((sum(x**4 for x in sm)/len(sm))**0.25)

                    elev_gain_m = 0.0; elev_loss_m = 0.0
                    if alts and len(alts) > 1:
                        for i in range(1, len(alts)):
                            d = alts[i] - alts[i-1]
                            if d > 0: elev_gain_m += d
                            elif d < 0: elev_loss_m += -d
                    elev_ft = round(elev_gain_m * 3.28084) if elev_gain_m else (round((act.get('total_elevation_gain') or 0) * 3.28084))
                    elev_loss_ft = round(elev_loss_m * 3.28084) if elev_loss_m else None

                    avg_power = round(sum(powers)/len(powers)) if powers else None
                    avg_hr    = round(sum(hrs)/len(hrs)) if hrs else None
                    avg_cad   = round(sum(cads)/len(cads)) if cads else None

                    ride_type = classify_ride(dist_mi, dur_h, avg_power, is_virt)

                    cur3.execute("""INSERT INTO rides (user_id,ride_date,start_time,name,dist_mi,duration_h,
                        avg_power,norm_power,avg_hr,max_hr,avg_cadence,max_cadence,
                        p5,p15,p30,p300,elev_gain_ft,elev_loss_ft,calories,elapsed_h,ride_type,is_virtual,temp_c,notes,possible_duplicate_of)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                        (user['id'], act_date, start_time_val, act.get('name','Activity'),
                         dist_mi, dur_h, avg_power, np_val,
                         avg_hr, act.get('max_heartrate'),
                         avg_cad, max(cads) if cads else None,
                         best_avg(powers,5), best_avg(powers,15), best_avg(powers,30), best_avg(powers,300),
                         elev_ft, elev_loss_ft, act.get('calories'), elapsed_h, ride_type, is_virt,
                         act.get('average_temp'), None, duplicate_of_id))
                    new_ride_id = cur3.fetchone()[0]

                    stream_data = {
                        'time_offset_s': stream_vals('time'),
                        'distance':      stream_vals('distance'),
                        'altitude':      alts,
                        'speed':         stream_vals('velocity_smooth'),
                        'power':         stream_vals('watts'),
                        'heart_rate':    stream_vals('heartrate'),
                        'cadence':       stream_vals('cadence'),
                    }
                    if any(stream_data.values()):
                        cur3.execute("INSERT INTO ride_streams (ride_id, streams) VALUES (%s,%s)",
                                    (new_ride_id, psycopg2.extras.Json(stream_data)))
                    cur3.close()
                    imported += 1
                    if duplicate_of_id:
                        flagged += 1
                except Exception as e:
                    errors += 1

            if len(activities) < 50:
                break
            page += 1

    cur.execute("UPDATE strava_tokens SET last_sync=NOW() WHERE user_id=%s", (user['id'],))
    cur.close(); conn.close()

    return {"imported": imported, "skipped": skipped, "errors": errors, "out_of_range": out_of_range, "flagged": flagged, "non_cycling": non_cycling,
            "message": f"Synced {imported} new activities from Strava ({skipped} already existed, "
                       f"{out_of_range} outside {YEAR}" + (f", {flagged} possible duplicate{'s' if flagged != 1 else ''} to review" if flagged else "")
                       + (f", {non_cycling} non-cycling activit{'ies' if non_cycling != 1 else 'y'} skipped" if non_cycling else "") + ")"}

@app.delete("/rides/clear")
def clear_rides(user: dict = Depends(get_current_user)):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM rides WHERE user_id=%s", (user['id'],))
    cur.close(); conn.close()
    return {"status": "all rides cleared"}

@app.delete("/rides/{ride_id}")
def delete_ride(ride_id: int, user: dict = Depends(get_current_user)):
    """Delete a single ride. ride_streams cleans up automatically
    (ON DELETE CASCADE) since it's keyed off ride_id."""
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM rides WHERE id=%s AND user_id=%s", (ride_id, user['id']))
    deleted = cur.rowcount
    cur.close(); conn.close()
    if not deleted:
        raise HTTPException(status_code=404, detail="Ride not found")
    return {"status": "deleted", "id": ride_id}

@app.get("/rides/audit-slow-pace")
def audit_slow_pace(user: dict = Depends(get_current_user)):
    """Flags rides with a suspiciously slow average pace — a real
    heuristic for finding walks that got imported as rides before the
    activity-type filter existed (v2.10.0), since the app never stored
    what kind of activity each ride actually was, so there's no direct
    way to look this up. Walking pace (2-4 mph) is categorically
    different from even a slow, casual ride (6+ mph), so a <6 mph
    threshold should catch genuine walks — but this is a heuristic,
    not a certainty: a very hilly or stop-heavy ride could occasionally
    land here too. Surfaces candidates for manual review; deletes
    nothing on its own."""
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT id, ride_date, name, dist_mi, duration_h,
               ROUND((dist_mi / NULLIF(duration_h, 0))::numeric, 1) AS avg_mph
        FROM rides
        WHERE user_id=%s AND dist_mi IS NOT NULL AND duration_h IS NOT NULL AND duration_h > 0
            AND (dist_mi / duration_h) < 6
        ORDER BY ride_date DESC
    """, (user['id'],))
    candidates = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return {"candidates": candidates, "count": len(candidates)}

@app.get("/rides/audit-duplicates")
def audit_duplicates(user: dict = Depends(get_current_user)):
    """One-time diagnostic for the v2.9.1 fix — finds the specific
    pattern the bug created: a freshly re-synced ride (has start_time)
    paired with an older, pre-existing ride (no start_time, predates
    that column) for what's almost certainly the same physical
    activity — same date, near-identical distance and duration.
    Tighter tolerance than the normal dedup check (0.3mi/3min, not
    0.5mi/6min) since these should be exact re-imports of the same
    source data, not just similar rides. Read-only — surfaces
    candidates for review, deletes nothing itself."""
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT
            old.id AS old_id, old.ride_date AS old_date, old.name AS old_name,
            old.dist_mi AS old_dist, old.duration_h AS old_duration,
            old.avg_power AS old_avg_power, old.avg_hr AS old_avg_hr,
            new.id AS new_id, new.ride_date AS new_date, new.name AS new_name,
            new.dist_mi AS new_dist, new.duration_h AS new_duration,
            new.avg_power AS new_avg_power, new.avg_hr AS new_avg_hr
        FROM rides old
        JOIN rides new ON new.user_id = old.user_id
            AND new.id != old.id
            AND old.start_time IS NULL
            AND new.start_time IS NOT NULL
            AND ABS(old.ride_date - new.ride_date) <= 1
            AND ABS(COALESCE(old.dist_mi,0) - COALESCE(new.dist_mi,0)) < 0.3
            AND ABS(COALESCE(old.duration_h,0) - COALESCE(new.duration_h,0)) < 0.05
        WHERE old.user_id = %s
        ORDER BY old.ride_date DESC
    """, (user['id'],))
    pairs = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return {"pairs": pairs, "count": len(pairs)}

@app.get("/rides/flagged")
def get_flagged_rides(user: dict = Depends(get_current_user)):
    """Rides imported despite their time window overlapping an existing
    ride's — flagged for the rider to resolve (delete the wrong one, or
    confirm both are genuinely separate), rather than the app silently
    guessing which recording is the real one."""
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT f.id AS flagged_id, f.ride_date AS flagged_date, f.name AS flagged_name,
               f.dist_mi AS flagged_dist, f.duration_h AS flagged_duration,
               f.avg_power AS flagged_avg_power, f.avg_hr AS flagged_avg_hr,
               o.id AS original_id, o.ride_date AS original_date, o.name AS original_name,
               o.dist_mi AS original_dist, o.duration_h AS original_duration,
               o.avg_power AS original_avg_power, o.avg_hr AS original_avg_hr
        FROM rides f
        JOIN rides o ON f.possible_duplicate_of = o.id
        WHERE f.user_id=%s
        ORDER BY f.ride_date DESC
    """, (user['id'],))
    pairs = [dict(r) for r in cur.fetchall()]
    cur.close(); conn.close()
    return {"flagged": pairs, "count": len(pairs)}

@app.post("/rides/{ride_id}/clear-review")
def clear_review(ride_id: int, user: dict = Depends(get_current_user)):
    """Confirms a flagged ride is genuinely separate from the one it
    overlapped, not a duplicate — clears the flag, keeps both rides."""
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE rides SET possible_duplicate_of=NULL WHERE id=%s AND user_id=%s", (ride_id, user['id']))
    updated = cur.rowcount
    cur.close(); conn.close()
    if not updated:
        raise HTTPException(status_code=404, detail="Ride not found")
    return {"status": "cleared", "id": ride_id}

def _downsample(arr, target=400):
    """Thin an array down for charting — a 2-3 hour ride can be 7,000+
    samples, too many to plot smoothly, and a human reading the shape of
    a line doesn't need per-second resolution to see it. Stats elsewhere
    still use the full-resolution data; this is charts only."""
    if not arr:
        return []
    n = len(arr)
    step = max(1, n // target)
    return arr[::step]

def compute_power_curve(power_stream):
    """Continuous best-average-power across a log-spaced set of durations
    (1s out to the full ride length), for the ride-detail page's dedicated
    Power Curve chart — a finer-resolution companion to the four fixed
    checkpoints (5s/15s/30s/5-min) used on the long-term Dashboard trend
    charts, which stay exactly as-is; this is scoped to one specific ride.

    Deliberately treats a missing/None reading as 0 watts, same as a
    genuine coasting sample — a power curve should reflect real sustained
    output including recovery dips, not silently splice them out. This is
    a different convention from p5/p15/p30/p300 elsewhere in this file,
    which historically filter zero/coasting readings out before computing
    bests (see parse_fit_bytes) — that's a pre-existing, deliberately
    unchanged behavior for those fixed checkpoints; the power curve is a
    new, separate computation free to make its own call, and "best average
    power over a sustained duration" is the standard definition used by
    every other tool that draws this kind of curve.

    Assumes ~1 sample/second recording, same convention already used by
    best_avg() and tool_find_push_segments() elsewhere in this file — the
    stored streams don't carry a reliable per-sample interval, and this
    matches how the rest of the app already treats them.

    Uses a prefix-sum rolling-window max per checkpoint duration — O(n)
    per checkpoint rather than a naive O(n) recompute from scratch inside
    an O(n) outer loop over every possible duration (which would be
    O(n^2) and, for a 3+ hour ride at ~10,800 samples, far too slow to
    compute on every ride-detail page load).
    """
    if not power_stream:
        return []
    vals = [p if p is not None else 0 for p in power_stream]
    n = len(vals)
    if n < 5:
        return []

    prefix = [0] * (n + 1)
    for i, v in enumerate(vals):
        prefix[i + 1] = prefix[i] + v

    # Log-ish spaced checkpoints, matching the density used by every other
    # tool that draws this curve (dense at the short/anaerobic end, sparse
    # at the long/aerobic end) — capped at the ride's own length below.
    CHECKPOINTS = [
        1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 12, 15, 20, 25, 30, 40, 50, 60, 75, 90,
        120, 150, 180, 240, 300, 420, 600, 900, 1200, 1800, 2400, 3000,
        3600, 4500, 5400, 7200, 9000, 10800, 14400, 18000, 21600
    ]
    curve = []
    for d in CHECKPOINTS:
        if d > n:
            break
        best_sum = max(prefix[i + d] - prefix[i] for i in range(n - d + 1))
        curve.append({"duration_s": d, "watts": round(best_sum / d)})
    return curve

def estimate_ftp_from_curve(power_curve):
    """Rough FTP estimate from a single ride's power curve — 95% of the
    ride's best 20-minute power, the standard field estimate. Not a
    lab-measured or structured-test FTP; always presented to the rider
    as a suggestion to confirm, never silently written over what's on
    file. Returns None when the ride isn't long enough for a 20-min
    checkpoint (compute_power_curve() only includes a duration once the
    ride has at least that many samples).

    An earlier version of this also tried a 60-minute-direct fallback
    for rides without a 20-min checkpoint but long enough for a 60-min
    one — caught in testing as dead code: compute_power_curve() finds
    the best window of each duration anywhere in the ride, so any ride
    with enough samples for a 3600s checkpoint necessarily also has
    enough for a 1200s one. That branch could never actually fire and
    was removed rather than left in as unreachable code."""
    if not power_curve:
        return None
    by_dur = {c['duration_s']: c['watts'] for c in power_curve}
    if 1200 in by_dur:
        return {'watts': round(by_dur[1200] * 0.95), 'basis': "95% of 20-min best"}
    return None

def fmt_duration_label(s):
    """Human-readable label for a duration in seconds — 45s, 5min, 1.5hr
    — shared by the power-curve chart's tick formatting logic (JS side)
    and the coaching tool's checkpoint labels (Python side), so both
    describe the same durations the same way."""
    if s < 60:
        return f"{s}s"
    if s < 3600:
        m = s / 60
        return f"{int(m)}min" if m % 1 == 0 else f"{m:g}min"
    h = s / 3600
    return f"{int(h)}hr" if h % 1 == 0 else f"{h:g}hr"

def build_ride_detail_html(ride, streams, profile_ftp=None):
    def esc(s):
        return html.escape(str(s)) if s is not None else ''
    def j(v):
        return json.dumps(v)

    ride_date = ride.get('ride_date')
    date_str = ride_date.strftime('%B %d, %Y') if hasattr(ride_date, 'strftime') else str(ride_date)

    def stat(label, value, unit=''):
        if value is None or value == '':
            return ''
        return "<div class='dstat'><div class='dlabel'>" + esc(label) + "</div><div class='dvalue'>" + esc(value) + (" " + esc(unit) if unit else '') + "</div></div>"

    stats_html = (
        "<div class='dstats-grid'>"
        + stat("Distance", ride.get('dist_mi'), "mi")
        + stat("Moving Time", ride.get('duration_h'), "h")
        + stat("Elevation Gain", ride.get('elev_gain_ft'), "ft")
        + stat("Elevation Loss", ride.get('elev_loss_ft'), "ft")
        + stat("Avg Power", ride.get('avg_power'), "W")
        + stat("Normalized Power", ride.get('norm_power'), "W")
        + stat("Avg HR", ride.get('avg_hr'), "bpm")
        + stat("Max HR", ride.get('max_hr'), "bpm")
        + stat("Avg Cadence", ride.get('avg_cadence'), "rpm")
        + stat("Calories", ride.get('calories'))
        + stat("Training Stress", ride.get('training_stress_score'))
        + stat("Intensity Factor", ride.get('intensity_factor'))
        + stat("L/R Balance", ride.get('avg_lr_balance'), "% right")
        + stat("Equipment", ride.get('equipment_name'))
        + "</div>"
    )

    has_streams = bool(streams and streams.get('distance'))
    charts_html = ""
    chart_js = ""
    ftp_banner_html = ""

    if has_streams:
        distances_raw = streams.get('distance') or []
        n = len(distances_raw)
        distances_mi = [round(d/1609.34, 2) if d is not None else None for d in distances_raw]

        def series(key, transform=None):
            vals = (streams.get(key) or [])[:n]
            vals = vals + [None] * (n - len(vals))
            if transform:
                vals = [transform(v) if v is not None else None for v in vals]
            return vals

        labels_ds = _downsample(distances_mi)
        power_ds = _downsample(series('power'))
        hr_ds = _downsample(series('heart_rate'))
        cadence_ds = _downsample(series('cadence'))
        alt_ds = _downsample(series('altitude', lambda a: round(a * 3.28084, 1)))

        # Full-resolution data, kept separate from the downsampled chart
        # arrays above — a multi-hour ride can be 7,000+ samples, too many
        # to render smoothly, but a drag-selected window needs real
        # precision to compute accurate stats for just that stretch, not
        # an approximation from ~400 thinned points.
        full_dist = distances_mi
        full_power = series('power')
        full_hr = series('heart_rate')
        full_cadence = series('cadence')
        full_alt_ft = series('altitude', lambda a: round(a * 3.28084, 1))

        # L/R balance is FIT-upload-only (Strava's API doesn't expose it) —
        # only show this chart when there's real data, same graceful-absence
        # approach used for the whole chart section when streams are missing.
        lr_raw = streams.get('left_right_balance') or []
        has_lr = any(v is not None for v in lr_raw)
        lr_ds = _downsample(series('left_right_balance')) if has_lr else []

        # Power curve — continuous best-average-power across every duration,
        # not just the four fixed dashboard checkpoints. Its own x-axis
        # (duration, log scale) is a completely different domain from the
        # mile-based charts above, so it's deliberately NOT part of the
        # shared drag-select-to-recompute group below — dragging on it
        # wouldn't mean the same thing as dragging on a mile-based chart.
        power_curve = compute_power_curve(streams.get('power')) if streams.get('power') else []
        has_power_curve = len(power_curve) >= 3  # need at least a few points for a meaningful curve

        # FTP estimate banner — only shown when a real estimate is
        # computable AND it differs meaningfully (5%+) from what's on
        # file in the profile, so this doesn't nag on every ride. Never
        # auto-applies; always a suggestion with an explicit confirm
        # step (POST /profile/ftp), never silently overwriting a real
        # FTP test result.
        ftp_banner_html = ""
        ftp_estimate = estimate_ftp_from_curve(power_curve) if has_power_curve else None
        if ftp_estimate:
            est_watts = ftp_estimate['watts']
            show_banner = (profile_ftp is None) or (abs(est_watts - profile_ftp) / max(profile_ftp, 1) >= 0.05)
            if show_banner:
                current_txt = (str(profile_ftp) + "W on file") if profile_ftp else "no FTP on file yet"
                ftp_banner_html = (
                    "<div class='dftp-banner'>"
                    + "<b>Estimated FTP from this ride: " + str(est_watts) + "W</b> ("
                    + esc(ftp_estimate['basis']) + ") &nbsp; — currently " + esc(current_txt) + ". "
                    + "<a href=\"#\" onclick=\"window.parent.applyEstimatedFtp && window.parent.applyEstimatedFtp("
                    + str(est_watts) + ");return false;\">Update profile FTP →</a>"
                    + "</div>"
                )

        charts_html = (
            "<div class='dselect-bar' id='selectBar'>Drag across any chart below to see stats for just that stretch.</div>"
            "<div class='dchart-card'><h3>Altitude Profile</h3><div class='dchart-wrap'><canvas id='altChart'></canvas><div class='dselect-overlay' id='altChartOv'></div></div></div>"
            "<div class='dchart-card'><h3>Power</h3><div class='dchart-wrap'><canvas id='powerChart'></canvas><div class='dselect-overlay' id='powerChartOv'></div></div></div>"
            "<div class='dchart-card'><h3>Heart Rate</h3><div class='dchart-wrap'><canvas id='hrChart'></canvas><div class='dselect-overlay' id='hrChartOv'></div></div></div>"
            "<div class='dchart-card'><h3>Cadence</h3><div class='dchart-wrap'><canvas id='cadChart'></canvas><div class='dselect-overlay' id='cadChartOv'></div></div></div>"
            + ("<div class='dchart-card'><h3>Power Curve</h3><div class='dchart-wrap'><canvas id='powerCurveChart'></canvas></div></div>" if has_power_curve else "")
            + ("<div class='dchart-card'><h3>Left/Right Power Balance</h3><div class='dchart-wrap'><canvas id='lrChart'></canvas><div class='dselect-overlay' id='lrChartOv'></div></div></div>" if has_lr else "")
        )
        # ── v2.12.0 fix ──────────────────────────────────────────────────
        # Two real bugs, same root cause: the x-axis on every chart here
        # was left on Chart.js's default CATEGORY scale (labels/data passed
        # as two parallel arrays, no scales.x.type set) instead of a real
        # LINEAR scale. getValueForPixel() on a category scale returns the
        # *index* into the label array, not the actual mile value the
        # label displays — so pixelToValue()/computeSelectionStats() were
        # silently reading index numbers as if they were miles. Near the
        # start of a ride the index and the mile value happen to be close
        # in magnitude, which is why dragging there looked plausible while
        # dragging further out did not. Fixed by switching every chart to
        # scales.x.type:'linear' and passing each dataset as real {x,y}
        # point objects (built by the small pts() helper below) instead of
        # a shared category-label array. computeSelectionStats() and
        # pixelToValue() themselves needed no changes — they were already
        # written correctly for real mile values, they just weren't being
        # fed them.
        # Also fixed: charts had no explicit height, so Chart.js's
        # responsive sizing had nothing to anchor to and kept growing —
        # only one chart fit on screen at a time. .dchart-wrap now has a
        # fixed height (matches the 260px cap already used successfully
        # on the main Dashboard's charts) with maintainAspectRatio:false
        # so the canvas fills it instead of dictating its own size.
        chart_js = (
            "const X=" + j(labels_ds) + ";"
            "function pts(xs,ys){return xs.map((x,i)=>({x:x,y:ys[i]}));}"
            "const opts={responsive:true,maintainAspectRatio:false,animation:false,"
            "elements:{point:{radius:0},line:{tension:0.2}},"
            "scales:{x:{type:'linear',title:{display:true,text:'Miles'},ticks:{maxTicksLimit:8}}}};"
            "new Chart(document.getElementById('altChart'),{type:'line',data:{"
            "datasets:[{data:pts(X," + j(alt_ds) + "),borderColor:'#9333ea',backgroundColor:'rgba(147,51,234,0.1)',fill:true}]},"
            "options:Object.assign({},opts,{scales:Object.assign({},opts.scales,{y:{title:{display:true,text:'ft'}}})})});"
            "new Chart(document.getElementById('powerChart'),{type:'line',data:{"
            "datasets:[{data:pts(X," + j(power_ds) + "),borderColor:'#2563eb'}]},"
            "options:Object.assign({},opts,{scales:Object.assign({},opts.scales,{y:{title:{display:true,text:'W'}}})})});"
            "new Chart(document.getElementById('hrChart'),{type:'line',data:{"
            "datasets:[{data:pts(X," + j(hr_ds) + "),borderColor:'#dc2626'}]},"
            "options:Object.assign({},opts,{scales:Object.assign({},opts.scales,{y:{title:{display:true,text:'bpm'}}})})});"
            "new Chart(document.getElementById('cadChart'),{type:'line',data:{"
            "datasets:[{data:pts(X," + j(cadence_ds) + "),borderColor:'#059669'}]},"
            "options:Object.assign({},opts,{scales:Object.assign({},opts.scales,{y:{title:{display:true,text:'rpm'}}})})});"
            + ("new Chart(document.getElementById('lrChart'),{type:'line',data:{"
               "datasets:[{data:pts(X," + j(lr_ds) + "),borderColor:'#ea580c',label:'% right'}]},"
               "options:Object.assign({},opts,{plugins:{legend:{display:false}},scales:Object.assign({},opts.scales,{y:{title:{display:true,text:'% right'},suggestedMin:30,suggestedMax:70}})})});"
               if has_lr else "")
            + (
                # Power curve — its own log-scale duration axis, its own tick/
                # tooltip formatting (seconds -> "5s"/"1m"/"20m"/"1h"), and
                # deliberately NOT added to chartIds below — its x-axis is
                # duration, not miles, so it doesn't participate in the
                # shared mile-based drag-select-to-recompute interaction.
                "function fmtDur(s){"
                "if(s<60)return s+'s';"
                "if(s<3600){var m=s/60;return (m%1===0?m:m.toFixed(1))+'m';}"
                "var h=s/3600;return (h%1===0?h:h.toFixed(1))+'h';}"
                "new Chart(document.getElementById('powerCurveChart'),{type:'line',data:{"
                "datasets:[{data:" + j([{"x": p["duration_s"], "y": p["watts"]} for p in power_curve]) + ","
                "borderColor:'#7c3aed',backgroundColor:'rgba(124,58,237,0.08)',fill:true,"
                "pointRadius:0,pointHitRadius:20,pointHoverRadius:5,pointHoverBackgroundColor:'#7c3aed',"
                "tension:0.15,borderWidth:2}]},"
                # v2.13.1 fix: the tooltip originally only fired on a pixel-
                # perfect hit against an invisible (radius-0) point on the
                # line — workable with a mouse if you're careful, effectively
                # impossible with a finger on mobile. interaction mode
                # 'nearest' + axis:'x' + intersect:false means any hover/tap
                # near a given duration finds the closest point along the
                # x-axis, no exact pixel match needed; pointHitRadius:20
                # widens each point's own hit target too, on top of that.
                "options:{responsive:true,maintainAspectRatio:false,animation:false,"
                "interaction:{mode:'nearest',axis:'x',intersect:false},"
                "plugins:{legend:{display:false},tooltip:{callbacks:{"
                "title:function(items){return fmtDur(items[0].parsed.x);},"
                "label:function(ctx){return ctx.parsed.y+'W';}}}},"
                "scales:{x:{type:'logarithmic',title:{display:true,text:'Duration'},"
                "ticks:{callback:function(v){"
                "var nice=[1,5,15,30,60,300,600,1200,3600,7200,10800,21600];"
                "if(nice.indexOf(v)===-1)return null;return fmtDur(v);}}},"
                "y:{title:{display:true,text:'Watts'},beginAtZero:true}}}"
                "});"
                if has_power_curve else ""
            )
            + (
                "const fullData={dist:" + j(full_dist) + ",power:" + j(full_power) + ",hr:" + j(full_hr)
                + ",cadence:" + j(full_cadence) + ",alt:" + j(full_alt_ft) + "};"
                "const chartIds=" + j(['altChart','powerChart','hrChart','cadChart'] + (['lrChart'] if has_lr else [])) + ";"
                "const charts={};chartIds.forEach(id=>{charts[id]=Chart.getChart(id);});"
                "function computeSelectionStats(startMi,endMi){"
                "const lo=Math.min(startMi,endMi),hi=Math.max(startMi,endMi);"
                "const idxs=[];for(let i=0;i<fullData.dist.length;i++){if(fullData.dist[i]!=null&&fullData.dist[i]>=lo&&fullData.dist[i]<=hi)idxs.push(i);}"
                "if(!idxs.length)return null;"
                "const avg=(arr,exZero)=>{const v=idxs.map(i=>arr[i]).filter(x=>x!=null&&(!exZero||x>0));return v.length?Math.round(v.reduce((a,b)=>a+b,0)/v.length):null;};"
                "const mx=(arr)=>{const v=idxs.map(i=>arr[i]).filter(x=>x!=null);return v.length?Math.round(Math.max(...v)):null;};"
                "let elev=0;for(let k=1;k<idxs.length;k++){const a=fullData.alt[idxs[k-1]],b=fullData.alt[idxs[k]];if(a!=null&&b!=null&&b>a)elev+=(b-a);}"
                "return{distance:Math.round((hi-lo)*100)/100,avgPower:avg(fullData.power,true),maxPower:mx(fullData.power),"
                "avgHr:avg(fullData.hr,false),maxHr:mx(fullData.hr),avgCadence:avg(fullData.cadence,true),elevGain:Math.round(elev)};"
                "}"
                "function updateOverlays(startMi,endMi){"
                "chartIds.forEach(id=>{const c=charts[id],ov=document.getElementById(id+'Ov');if(!c||!ov)return;"
                "const p1=c.scales.x.getPixelForValue(startMi),p2=c.scales.x.getPixelForValue(endMi);"
                "ov.style.left=Math.min(p1,p2)+'px';ov.style.width=Math.abs(p2-p1)+'px';ov.style.display='block';});"
                "}"
                "function clearOverlays(){chartIds.forEach(id=>{const ov=document.getElementById(id+'Ov');if(ov)ov.style.display='none';});}"
                "function showSelectionStats(s){const bar=document.getElementById('selectBar');"
                "if(!s){bar.innerHTML='Drag across any chart below to see stats for just that stretch.';return;}"
                "bar.innerHTML='<b>'+s.distance+' mi selected</b> &nbsp; '"
                "+(s.avgPower!=null?'Avg Power: '+s.avgPower+'W (max '+s.maxPower+'W) &nbsp; ':'')"
                "+(s.avgHr!=null?'Avg HR: '+s.avgHr+'bpm (max '+s.maxHr+'bpm) &nbsp; ':'')"
                "+(s.avgCadence!=null?'Avg Cadence: '+s.avgCadence+'rpm &nbsp; ':'')"
                "+(s.elevGain?'Elev Gain: '+s.elevGain+'ft &nbsp; ':'')"
                "+'<a href=\"#\" onclick=\"clearSelection();return false;\" style=\"margin-left:8px;\">Clear</a>';}"
                "function clearSelection(){clearOverlays();showSelectionStats(null);}"
                "let dragState={active:false,startVal:null,chartId:null};"
                "function pixelToValue(chart,canvas,clientX){const r=canvas.getBoundingClientRect();return chart.scales.x.getValueForPixel(clientX-r.left);}"
                "function handleDragMove(clientX){if(!dragState.active)return;const c=charts[dragState.chartId],cv=document.getElementById(dragState.chartId);if(!c||!cv)return;"
                "updateOverlays(dragState.startVal,pixelToValue(c,cv,clientX));}"
                "function handleDragEnd(clientX){if(!dragState.active)return;const c=charts[dragState.chartId],cv=document.getElementById(dragState.chartId);dragState.active=false;if(!c||!cv)return;"
                "const val=pixelToValue(c,cv,clientX);"
                "if(Math.abs(val-dragState.startVal)<0.05){clearSelection();return;}"
                "updateOverlays(dragState.startVal,val);showSelectionStats(computeSelectionStats(dragState.startVal,val));}"
                "chartIds.forEach(id=>{const cv=document.getElementById(id);if(!cv)return;"
                "cv.addEventListener('mousedown',e=>{const c=charts[id];if(!c)return;dragState={active:true,startVal:pixelToValue(c,cv,e.clientX),chartId:id};});"
                "cv.addEventListener('touchstart',e=>{const c=charts[id];if(!c||!e.touches[0])return;dragState={active:true,startVal:pixelToValue(c,cv,e.touches[0].clientX),chartId:id};},{passive:true});"
                "});"
                "document.addEventListener('mousemove',e=>handleDragMove(e.clientX));"
                "document.addEventListener('touchmove',e=>{if(e.touches[0])handleDragMove(e.touches[0].clientX);},{passive:true});"
                "document.addEventListener('mouseup',e=>handleDragEnd(e.clientX));"
                "document.addEventListener('touchend',e=>handleDragEnd((e.changedTouches[0]||{}).clientX));"
              if has_streams else "")
        )
    else:
        charts_html = "<div class='dchart-card'><p class='dno-data'>No detailed chart data for this ride — only rides uploaded/synced since raw-data storage was added have this.</p></div>"

    return (
        "<!DOCTYPE html><html><head><meta charset='utf-8'>"
        + "<style>"
        + "body{font-family:-apple-system,Segoe UI,Roboto,sans-serif;background:#f7f7f8;margin:0;padding:16px;color:#222;}"
        + "h1{font-size:1.2rem;margin:0 0 2px;} .ddate{font-size:0.8rem;color:#888;margin-bottom:16px;}"
        + "h3{font-size:0.85rem;margin:0 0 10px;color:#444;}"
        + ".dstats-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(120px,1fr));gap:10px;margin-bottom:20px;}"
        + ".dstat{background:#fff;border-radius:8px;padding:10px 12px;box-shadow:0 1px 4px rgba(0,0,0,0.06);border-left:3px solid #2563eb;}"
        + ".dlabel{font-size:0.65rem;text-transform:uppercase;color:#999;letter-spacing:.03em;}"
        + ".dvalue{font-size:1.05rem;font-weight:600;margin-top:2px;}"
        + ".dchart-card{background:#fff;border-radius:8px;padding:14px;box-shadow:0 1px 4px rgba(0,0,0,0.06);margin-bottom:14px;}"
        + ".dno-data{font-size:0.85rem;color:#888;text-align:center;padding:20px 0;}"
        + ".dchart-wrap{position:relative;height:260px;}"
        + ".dchart-wrap canvas{max-height:260px;}"
        + ".dselect-overlay{position:absolute;top:0;bottom:0;display:none;background:rgba(37,99,235,0.12);border-left:2px solid #2563eb;border-right:2px solid #2563eb;pointer-events:none;}"
        + ".dselect-bar{background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:10px 14px;margin-bottom:14px;font-size:0.8rem;color:#1e40af;}"
        + ".dftp-banner{background:#f5f3ff;border:1px solid #ddd6fe;border-radius:8px;padding:10px 14px;margin-bottom:14px;font-size:0.8rem;color:#5b21b6;}"
        + ".dftp-banner a{color:#7c3aed;font-weight:600;}"
        + ".dhead-row{display:flex;justify-content:space-between;align-items:flex-start;gap:10px;flex-wrap:wrap;}"
        + ".dexport-link{font-size:0.75rem;color:#3b82f6;font-weight:500;text-decoration:none;white-space:nowrap;}"
        + "canvas{cursor:crosshair;}"
        + "</style></head><body>"
        + "<div class='dhead-row'>"
        + "<div><h1>" + esc(ride.get('name', 'Ride')) + "</h1><div class='ddate'>" + esc(date_str) + "</div></div>"
        + "<a href=\"#\" class='dexport-link' onclick=\"window.parent.exportRideTcx && window.parent.exportRideTcx(" + str(ride.get('id')) + ");return false;\">📥 Export TCX</a>"
        + "</div>"
        + ftp_banner_html
        + stats_html
        + charts_html
        + "<script src='https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js'></script>"
        + "<script>" + chart_js + "</script>"
        + "</body></html>"
    )

@app.get("/rides/by-date/{ride_date}")
def get_ride_by_date(ride_date: str, user: dict = Depends(get_current_user)):
    """Look up a ride by exact date — bridges the coaching memory's
    dated log (which has no ride_id, only a date) through to a ride's
    detail page. If multiple rides share a date, prefers one that has
    stream data, since that gives a more useful detail page than the
    "no chart data" fallback."""
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""
        SELECT r.id FROM rides r
        LEFT JOIN ride_streams rs ON rs.ride_id = r.id
        WHERE r.user_id=%s AND r.ride_date=%s
        ORDER BY (rs.ride_id IS NOT NULL) DESC, r.id DESC
        LIMIT 1
    """, (user['id'], ride_date))
    row = cur.fetchone()
    cur.close(); conn.close()
    if not row:
        raise HTTPException(status_code=404, detail="No ride found for this date")
    return {"ride_id": row['id']}

def build_tcx(ride, streams):
    """Builds a Garmin TCX v2 export for one ride from its stored raw
    stream data.

    GPX was considered and ruled out: GPX trackpoints require a real
    lat/lon per the spec, and this app has never stored GPS coordinates
    — FIT and Strava streams here only ever captured power, HR,
    cadence, altitude, distance, and speed. Writing a GPX with fake or
    zeroed coordinates would misrepresent the ride's actual route to
    whatever tool imports it, which is worse than not offering GPX at
    all. TCX's Trackpoint schema, unlike GPX's, makes Position
    genuinely optional — Time, DistanceMeters, AltitudeMeters,
    HeartRateBpm, Cadence, and a Garmin TPX extension for Watts are all
    valid without it — so this is a real, standards-compliant export
    that TrainingPeaks, Garmin Connect, WKO, and others can read, just
    without a route/map. TCX *import* into this app is a natural
    follow-up but a separate, larger piece (its own parser alongside
    the existing FIT/Strava paths) — not attempted here.
    """
    def esc(s):
        return html.escape(str(s)) if s is not None else ''

    start_time = ride.get('start_time')
    if start_time and hasattr(start_time, 'isoformat'):
        start_dt = start_time
    elif ride.get('ride_date'):
        start_dt = datetime.combine(ride['ride_date'], datetime.min.time())
    else:
        start_dt = datetime.utcnow()

    distances = streams.get('distance') or []
    n = len(distances)
    altitudes = streams.get('altitude') or []
    powers = streams.get('power') or []
    hrs = streams.get('heart_rate') or []
    cadences = streams.get('cadence') or []

    # Per-point absolute timestamps — FIT-sourced streams carry real ISO
    # timestamps ('timestamp'); Strava-sourced streams only carry a
    # seconds-from-start offset ('time_offset_s', their native format,
    # kept as-is on import). Handle both rather than assuming one; fall
    # back to a ~1Hz assumption (matching the convention used elsewhere
    # in this file, e.g. best_avg()) if neither is present.
    raw_ts = streams.get('timestamp')
    time_offsets = streams.get('time_offset_s')

    def point_time(i):
        if raw_ts and i < len(raw_ts) and raw_ts[i]:
            try:
                return datetime.fromisoformat(str(raw_ts[i]).replace('Z', ''))
            except Exception:
                pass
        if time_offsets and i < len(time_offsets) and time_offsets[i] is not None:
            return start_dt + timedelta(seconds=time_offsets[i])
        return start_dt + timedelta(seconds=i)

    def g(arr, i):
        return arr[i] if i < len(arr) else None

    trackpoints = []
    for i in range(n):
        t = point_time(i)
        parts = ["<Trackpoint>", "<Time>" + t.strftime('%Y-%m-%dT%H:%M:%SZ') + "</Time>"]
        d = g(distances, i)
        if d is not None:
            parts.append("<DistanceMeters>" + str(round(d, 1)) + "</DistanceMeters>")
        a = g(altitudes, i)
        if a is not None:
            parts.append("<AltitudeMeters>" + str(round(a, 1)) + "</AltitudeMeters>")
        h_ = g(hrs, i)
        if h_ is not None:
            parts.append("<HeartRateBpm><Value>" + str(int(h_)) + "</Value></HeartRateBpm>")
        c = g(cadences, i)
        if c is not None:
            parts.append("<Cadence>" + str(int(c)) + "</Cadence>")
        p = g(powers, i)
        if p is not None:
            parts.append(
                "<Extensions><TPX xmlns=\"http://www.garmin.com/xmlschemas/ActivityExtension/v2\">"
                "<Watts>" + str(int(p)) + "</Watts></TPX></Extensions>"
            )
        parts.append("</Trackpoint>")
        trackpoints.append(''.join(parts))

    total_time_s = round(ride['duration_h'] * 3600) if ride.get('duration_h') else n
    if distances and distances[-1] is not None:
        total_dist_m = distances[-1]
    else:
        total_dist_m = (ride.get('dist_mi') or 0) * 1609.34
    calories = ride.get('calories') or 0

    xml = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n"
        "<TrainingCenterDatabase xmlns=\"http://www.garmin.com/xmlschemas/TrainingCenterDatabase/v2\" "
        "xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" "
        "xsi:schemaLocation=\"http://www.garmin.com/xmlschemas/TrainingCenterDatabase/v2 "
        "http://www.garmin.com/xmlschemas/TrainingCenterDatabasev2.xsd\">"
        "<Activities><Activity Sport=\"Biking\">"
        "<Id>" + start_dt.strftime('%Y-%m-%dT%H:%M:%SZ') + "</Id>"
        "<Lap StartTime=\"" + start_dt.strftime('%Y-%m-%dT%H:%M:%SZ') + "\">"
        "<TotalTimeSeconds>" + str(total_time_s) + "</TotalTimeSeconds>"
        "<DistanceMeters>" + str(round(total_dist_m, 1)) + "</DistanceMeters>"
        "<Calories>" + str(int(calories)) + "</Calories>"
        "<Intensity>Active</Intensity>"
        "<TriggerMethod>Manual</TriggerMethod>"
        "<Track>" + ''.join(trackpoints) + "</Track>"
        "</Lap>"
        "<Notes>" + esc(ride.get('name', 'Ride'))
        + " — exported from Cycling Coach. No GPS data (never captured by this app); "
        "power/HR/cadence/elevation/distance only.</Notes>"
        "</Activity></Activities>"
        "</TrainingCenterDatabase>"
    )
    return xml

@app.get("/rides/{ride_id}/export-tcx")
def export_ride_tcx(ride_id: int, user: dict = Depends(get_current_user)):
    """Single-ride TCX export — TrainingPeaks/Garmin Connect/WKO-
    compatible, built from this app's own stored raw stream data. See
    build_tcx() for why TCX rather than GPX (no GPS ever captured)."""
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM rides WHERE id=%s AND user_id=%s", (ride_id, user['id']))
    ride = cur.fetchone()
    if not ride:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Ride not found")
    cur.execute("SELECT streams FROM ride_streams WHERE ride_id=%s", (ride_id,))
    srow = cur.fetchone()
    cur.close(); conn.close()
    if not srow or not srow.get('streams'):
        raise HTTPException(status_code=400,
            detail="No detailed stream data available for this ride — only rides uploaded/synced since raw-data storage was added have this.")
    xml = build_tcx(dict(ride), srow['streams'])
    ride_date = ride.get('ride_date')
    date_str = ride_date.isoformat() if hasattr(ride_date, 'isoformat') else str(ride_date)
    filename = "ride_" + date_str + "_" + str(ride_id) + ".tcx"
    return Response(
        content=xml,
        media_type="application/vnd.garmin.tcx+xml",
        headers={"Content-Disposition": "attachment; filename=" + filename}
    )

@app.get("/rides/{ride_id}/detail-page", response_class=HTMLResponse)
def get_ride_detail_page(ride_id: int, user: dict = Depends(get_current_user)):
    conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("""SELECT r.*, e.name AS equipment_name FROM rides r
        LEFT JOIN equipment e ON r.equipment_id = e.id
        WHERE r.id=%s AND r.user_id=%s""", (ride_id, user['id']))
    ride = cur.fetchone()
    if not ride:
        cur.close(); conn.close()
        raise HTTPException(status_code=404, detail="Ride not found")
    cur.execute("SELECT streams FROM ride_streams WHERE ride_id=%s", (ride_id,))
    srow = cur.fetchone()
    cur.execute("SELECT ftp FROM profiles WHERE user_id=%s", (user['id'],))
    prow = cur.fetchone()
    cur.close(); conn.close()
    streams = srow['streams'] if srow else None
    profile_ftp = prow['ftp'] if prow and prow.get('ftp') else None
    html_out = build_ride_detail_html(dict(ride), streams, profile_ftp=profile_ftp)
    return HTMLResponse(content=html_out)

@app.get("/dashboard", response_class=HTMLResponse)
def get_dashboard(user: dict = Depends(get_current_user)):
    try:
        conn = get_db(); cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM profiles WHERE user_id=%s", (user['id'],))
        profile = cur.fetchone()
        user_goal = int(profile['annual_goal_mi']) if profile and profile.get('annual_goal_mi') else ANNUAL_GOAL
        cur.execute("""SELECT * FROM rides WHERE user_id=%s
            AND ride_date >= %s AND ride_date < %s
            ORDER BY ride_date ASC""",
            (user['id'], f'{YEAR}-01-01', f'{YEAR+1}-01-01'))
        rides = [dict(r) for r in cur.fetchall()]; cur.close(); conn.close()
        html = build_full_dashboard(rides, user['name'], annual_goal=user_goal, user_timezone=profile.get('timezone') if profile else None)
        return HTMLResponse(content=html)
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print("DASHBOARD ERROR:", tb)
        return HTMLResponse(content="<pre style='color:red;padding:20px;'>" + tb + "</pre>", status_code=500)
